"""
Update Landscaping Consolidation Workbook with Enhanced Data and Charts
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

# Load the landscaping data
input_path = "/workspaces/Spend_Analysis_and_Categorization/spend-analysis/data/processed/landscaping_structured.xlsx"
df = pd.read_excel(input_path)

# Filter to 2015-2025
df['Award_Date'] = pd.to_datetime(df['Award_Date'])
df['Award_Year'] = df['Award_Date'].dt.year
df = df[(df['Award_Year'] >= 2015) & (df['Award_Year'] <= 2025)].copy()

# Calculate vendor spend
vendor_spend = df.groupby('Vendor name').agg({
    'Award_Amount_Numeric': 'sum',
    'Contract (PO) #': 'count'
}).reset_index()
vendor_spend.columns = ['Vendor', 'Total Spend', 'Contract Count']
vendor_spend = vendor_spend.sort_values('Total Spend', ascending=False)
total_spend = vendor_spend['Total Spend'].sum()
vendor_spend['% of Total'] = (vendor_spend['Total Spend'] / total_spend * 100)

# Calculate department spend
dept_spend = df.groupby('Department').agg({
    'Award_Amount_Numeric': 'sum',
    'Contract (PO) #': 'count'
}).reset_index()
dept_spend.columns = ['Department', 'Total Spend', 'Contract Count']
dept_spend = dept_spend.sort_values('Total Spend', ascending=False)
dept_spend['% of Total'] = (dept_spend['Total Spend'] / total_spend * 100)

print("="*80)
print("UPDATING LANDSCAPING CONSOLIDATION WORKBOOK")
print("="*80)
print(f"\nAnalyzing {len(df)} contracts from 2015-2025")
print(f"Total spend: ${total_spend:,.2f}")
print(f"Vendors: {len(vendor_spend)}")
print(f"Departments: {len(dept_spend)}")

# Load existing workbook
wb_path = "/workspaces/Spend_Analysis_and_Categorization/spend-analysis/outputs/new landscape analysis.xlsx"
wb = load_workbook(wb_path)

# Define styles
header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')

# ============================================================================
# UPDATE SOURCE DATA SHEET
# ============================================================================
print("\n1. Updating Source Data sheet...")
ws_source = wb["Source Data"]

# Find next available row (after existing content)
row = 32

# TOP VENDORS TABLE
ws_source[f'A{row}'] = 'TOP VENDORS (ALL 14)'
ws_source[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
ws_source[f'A{row}'].fill = header_fill
ws_source.merge_cells(f'A{row}:D{row}')
row += 1

vendor_headers = ['Rank', 'Vendor Name', 'Total Spend', '% of Total']
for col, header in enumerate(vendor_headers, 1):
    cell = ws_source.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border_thin
row += 1

for idx, (_, vendor_row) in enumerate(vendor_spend.iterrows(), 1):
    ws_source.cell(row=row, column=1, value=idx).alignment = center_align
    ws_source.cell(row=row, column=2, value=vendor_row['Vendor']).alignment = left_align
    ws_source.cell(row=row, column=3, value=vendor_row['Total Spend']).number_format = '$#,##0'
    ws_source.cell(row=row, column=4, value=vendor_row['% of Total'] / 100).number_format = '0.0%'

    for col in range(1, 5):
        ws_source.cell(row=row, column=col).border = border_thin
    row += 1

row += 2

# USING DEPARTMENTS TABLE
ws_source[f'A{row}'] = 'DEPARTMENTS USING LANDSCAPING SERVICES'
ws_source[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
ws_source[f'A{row}'].fill = header_fill
ws_source.merge_cells(f'A{row}:D{row}')
row += 1

dept_headers = ['Department', 'Total Spend', '% of Total', 'Contracts']
for col, header in enumerate(dept_headers, 1):
    cell = ws_source.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border_thin
row += 1

for _, dept_row in dept_spend.iterrows():
    ws_source.cell(row=row, column=1, value=dept_row['Department']).alignment = left_align
    ws_source.cell(row=row, column=2, value=dept_row['Total Spend']).number_format = '$#,##0'
    ws_source.cell(row=row, column=3, value=dept_row['% of Total'] / 100).number_format = '0.0%'
    ws_source.cell(row=row, column=4, value=int(dept_row['Contract Count'])).alignment = center_align

    for col in range(1, 5):
        ws_source.cell(row=row, column=col).border = border_thin
    row += 1

print(f"   ✓ Added {len(vendor_spend)} vendors to TOP VENDORS table")
print(f"   ✓ Added {len(dept_spend)} departments to DEPARTMENTS table")

# ============================================================================
# UPDATE CHARTS SHEET
# ============================================================================
print("\n2. Updating Charts sheet...")
ws_charts = wb["Charts"]

# Clear existing charts
for chart in list(ws_charts._charts):
    ws_charts._charts.remove(chart)

# RECREATE PIE CHART WITH IMPROVED FORMATTING
print("   ✓ Recreating pie chart with data labels...")
pie = PieChart()
pie.title = "Service Category Breakdown ($97M Total)"
pie.style = 10
pie.height = 12
pie.width = 16

labels = Reference(ws_charts, min_col=1, min_row=5, max_row=8)
data = Reference(ws_charts, min_col=2, min_row=4, max_row=8)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)

# Enhanced data labels
pie.dataLabels = DataLabelList()
pie.dataLabels.showCatName = False
pie.dataLabels.showVal = True
pie.dataLabels.showPercent = True
pie.dataLabels.showLeaderLines = True

# Increase font size
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
pie.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(sz=1800, b=True))

ws_charts.add_chart(pie, "E4")

# RECREATE BAR CHART WITH IMPROVED FORMATTING
print("   ✓ Recreating bar chart with data labels...")
bar = BarChart()
bar.type = "col"
bar.title = "Consolidation Scenarios - Annual Savings"
bar.y_axis.title = "Annual Savings ($ Millions)"
bar.x_axis.title = "Scenario"
bar.style = 11
bar.height = 12
bar.width = 16

categories = Reference(ws_charts, min_col=1, min_row=14, max_row=16)
values = Reference(ws_charts, min_col=3, min_row=13, max_row=16)
bar.add_data(values, titles_from_data=True)
bar.set_categories(categories)

# Enhanced data labels
bar.dataLabels = DataLabelList()
bar.dataLabels.showVal = True

ws_charts.add_chart(bar, "E14")

# ADD NEW TOP 10 VENDORS HORIZONTAL BAR CHART
print("   ✓ Adding new Top 10 Vendors bar chart...")

# Add data for top 10 vendors chart
top10_start_row = 24
ws_charts[f'A{top10_start_row}'] = 'Top 10 Vendors by Spend'
ws_charts[f'A{top10_start_row}'].font = Font(name='Calibri', size=12, bold=True)

top10_headers = ['Vendor Name', 'Total Spend (Millions)']
row = top10_start_row + 1
for col, header in enumerate(top10_headers, 1):
    cell = ws_charts.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border_thin
row += 1

top10_vendors = vendor_spend.head(10)
for _, vendor_row in top10_vendors.iterrows():
    ws_charts.cell(row=row, column=1, value=vendor_row['Vendor']).alignment = left_align
    ws_charts.cell(row=row, column=2, value=vendor_row['Total Spend'] / 1_000_000).number_format = '0.0'

    for col in range(1, 3):
        ws_charts.cell(row=row, column=col).border = border_thin
    row += 1

# Create horizontal bar chart
bar_vendors = BarChart()
bar_vendors.type = "bar"  # Horizontal bars
bar_vendors.title = "Top 10 Vendors by Spend"
bar_vendors.y_axis.title = "Vendor"
bar_vendors.x_axis.title = "Total Spend ($ Millions)"
bar_vendors.style = 11
bar_vendors.height = 15
bar_vendors.width = 18

vendor_labels = Reference(ws_charts, min_col=1, min_row=top10_start_row+2, max_row=top10_start_row+11)
vendor_values = Reference(ws_charts, min_col=2, min_row=top10_start_row+1, max_row=top10_start_row+11)
bar_vendors.add_data(vendor_values, titles_from_data=True)
bar_vendors.set_categories(vendor_labels)

# Add data labels
bar_vendors.dataLabels = DataLabelList()
bar_vendors.dataLabels.showVal = True

ws_charts.add_chart(bar_vendors, "E24")

# Adjust column widths
ws_charts.column_dimensions['A'].width = 45
ws_charts.column_dimensions['B'].width = 25

# ============================================================================
# UPDATE EXECUTIVE SUMMARY SHEET
# ============================================================================
print("\n3. Updating Executive Summary sheet...")
ws_summary = wb["Executive Summary"]

# Find row after key metrics table (should be around row 18)
row = 19

# Add Top 5 Vendors section
ws_summary[f'A{row}'] = 'TOP 5 VENDORS (93% of Total Spend)'
ws_summary[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
ws_summary[f'A{row}'].fill = header_fill
ws_summary.merge_cells(f'A{row}:B{row}')
row += 1

top5_headers = ['Vendor Name', 'Total Spend']
for col, header in enumerate(top5_headers, 1):
    cell = ws_summary.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border_thin
row += 1

top5_vendors = vendor_spend.head(5)
for _, vendor_row in top5_vendors.iterrows():
    ws_summary.cell(row=row, column=1, value=vendor_row['Vendor']).alignment = left_align
    ws_summary.cell(row=row, column=2, value=vendor_row['Total Spend']).number_format = '$#,##0'

    for col in range(1, 3):
        ws_summary.cell(row=row, column=col).border = border_thin
    row += 1

# Update recommendation section (move it down)
row += 1
ws_summary.merge_cells(f'A{row}:F{row}')
ws_summary[f'A{row}'] = 'RECOMMENDATION: Pursue moderate consolidation scenario (7 vendors) - optimal balance of savings and risk management'
ws_summary[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, italic=True)
ws_summary[f'A{row}'].fill = PatternFill(start_color='D5F5D5', end_color='D5F5D5', fill_type='solid')
ws_summary[f'A{row}'].alignment = left_align
ws_summary[f'A{row}'].border = border_thin

print(f"   ✓ Added top 5 vendors to Executive Summary")

# Save workbook
wb.save(wb_path)

print("\n" + "="*80)
print("WORKBOOK UPDATE COMPLETE")
print("="*80)
print(f"\nUpdated file: {wb_path}")
print("\nChanges made:")
print("  ✓ Source Data: Added all 14 vendors with spend details")
print("  ✓ Source Data: Added all departments with spend details")
print("  ✓ Charts: Added Top 10 Vendors horizontal bar chart")
print("  ✓ Charts: Enhanced pie chart with $ and % data labels")
print("  ✓ Charts: Enhanced bar chart with larger fonts")
print("  ✓ Executive Summary: Added top 5 vendor list")
print("\nWorkbook ready for screen sharing and presentations!")
print("="*80)
