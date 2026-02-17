"""
Clean Landscaping Consolidation Workbook - Remove Charts, Focus on Data Tables
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load the landscaping data
input_path = "/workspaces/Spend_Analysis_and_Categorization/spend-analysis/data/processed/landscaping_structured.xlsx"
df = pd.read_excel(input_path)

# Filter to 2015-2025
df['Award_Date'] = pd.to_datetime(df['Award_Date'])
df['Award_Year'] = df['Award_Date'].dt.year
df = df[(df['Award_Year'] >= 2015) & (df['Award_Year'] <= 2025)].copy()

# Calculate vendor spend
vendor_spend = df.groupby('Vendor name').agg({
    'Award_Amount_Numeric': 'sum',
    'Contract (PO) #': 'count'
}).reset_index()
vendor_spend.columns = ['Vendor', 'Total Spend', 'Contract Count']
vendor_spend = vendor_spend.sort_values('Total Spend', ascending=False)
total_spend = vendor_spend['Total Spend'].sum()
vendor_spend['% of Total'] = (vendor_spend['Total Spend'] / total_spend * 100)

# Calculate department spend
dept_spend = df.groupby('Department').agg({
    'Award_Amount_Numeric': 'sum',
    'Contract (PO) #': 'count'
}).reset_index()
dept_spend.columns = ['Department', 'Total Spend', 'Contract Count']
dept_spend = dept_spend.sort_values('Total Spend', ascending=False)
dept_spend['% of Total'] = (dept_spend['Total Spend'] / total_spend * 100)

# Calculate service category spend
service_spend = df.groupby('Service_Category').agg({
    'Award_Amount_Numeric': 'sum',
    'Contract (PO) #': 'count'
}).reset_index()
service_spend.columns = ['Service Category', 'Total Spend', 'Contract Count']
service_spend = service_spend.sort_values('Total Spend', ascending=False)
service_spend['% of Total'] = (service_spend['Total Spend'] / total_spend * 100)

print("="*80)
print("CLEANING WORKBOOK - DATA TABLES ONLY")
print("="*80)
print(f"\nData Summary:")
print(f"  Vendors: {len(vendor_spend)}")
print(f"  Departments: {len(dept_spend)}")
print(f"  Service Categories: {len(service_spend)}")
print(f"  Total Spend: ${total_spend:,.2f}")

# Load existing workbook
wb_path = "/workspaces/Spend_Analysis_and_Categorization/spend-analysis/outputs/new landscape analysis.xlsx"
wb = load_workbook(wb_path)

# Define styles
header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')
right_align = Alignment(horizontal='right', vertical='center')

# ============================================================================
# REMOVE CHARTS SHEET
# ============================================================================
print("\n1. Removing Charts sheet...")
if "Charts" in wb.sheetnames:
    del wb["Charts"]
    print("   ✓ Charts sheet removed")

# ============================================================================
# CLEAN AND REBUILD SOURCE DATA SHEET
# ============================================================================
print("\n2. Rebuilding Source Data sheet with clean tables...")

# Remove old Source Data sheet and create new one
if "Source Data" in wb.sheetnames:
    del wb["Source Data"]

ws = wb.create_sheet("Source Data")

# Title
ws.merge_cells('A1:E1')
ws['A1'] = 'Landscaping Services - Source Data (2015-2025)'
ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='1F4788')
ws['A1'].alignment = center_align
ws.row_dimensions[1].height = 25

row = 3

# ============================================================================
# TABLE 1: TOP 14 VENDORS
# ============================================================================
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = 'TOP 14 VENDORS BY SPEND'
ws[f'A{row}'].font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
ws[f'A{row}'].fill = header_fill
ws[f'A{row}'].alignment = center_align
row += 1

# Headers
vendor_headers = ['Rank', 'Vendor Name', 'Total Spend', '% of Total']
for col, header in enumerate(vendor_headers, 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border_thin
row += 1

# Data
for idx, (_, vendor_row) in enumerate(vendor_spend.iterrows(), 1):
    ws.cell(row=row, column=1, value=idx).alignment = center_align
    ws.cell(row=row, column=2, value=vendor_row['Vendor']).alignment = left_align
    ws.cell(row=row, column=3, value=vendor_row['Total Spend']).number_format = '$#,##0'
    ws.cell(row=row, column=3).alignment = right_align
    ws.cell(row=row, column=4, value=vendor_row['% of Total'] / 100).number_format = '0.0%'
    ws.cell(row=row, column=4).alignment = center_align

    for col in range(1, 5):
        ws.cell(row=row, column=col).border = border_thin
    row += 1

print(f"   ✓ Added TOP 14 VENDORS table")

row += 2

# ============================================================================
# TABLE 2: USING DEPARTMENTS
# ============================================================================
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = 'USING DEPARTMENTS/AREAS'
ws[f'A{row}'].font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
ws[f'A{row}'].fill = header_fill
ws[f'A{row}'].alignment = center_align
row += 1

# Headers
dept_headers = ['Rank', 'Department/Area', 'Total Spend', '% of Total']
for col, header in enumerate(dept_headers, 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border_thin
row += 1

# Data
for idx, (_, dept_row) in enumerate(dept_spend.iterrows(), 1):
    ws.cell(row=row, column=1, value=idx).alignment = center_align
    ws.cell(row=row, column=2, value=dept_row['Department']).alignment = left_align
    ws.cell(row=row, column=3, value=dept_row['Total Spend']).number_format = '$#,##0'
    ws.cell(row=row, column=3).alignment = right_align
    ws.cell(row=row, column=4, value=dept_row['% of Total'] / 100).number_format = '0.0%'
    ws.cell(row=row, column=4).alignment = center_align

    for col in range(1, 5):
        ws.cell(row=row, column=col).border = border_thin
    row += 1

print(f"   ✓ Added USING DEPARTMENTS table")

row += 2

# ============================================================================
# TABLE 3: SERVICE CATEGORIES
# ============================================================================
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = 'SERVICE CATEGORIES'
ws[f'A{row}'].font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
ws[f'A{row}'].fill = header_fill
ws[f'A{row}'].alignment = center_align
row += 1

# Headers
service_headers = ['Rank', 'Service Category', 'Total Spend', '% of Total']
for col, header in enumerate(service_headers, 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border_thin
row += 1

# Data
for idx, (_, service_row) in enumerate(service_spend.iterrows(), 1):
    ws.cell(row=row, column=1, value=idx).alignment = center_align
    ws.cell(row=row, column=2, value=service_row['Service Category']).alignment = left_align
    ws.cell(row=row, column=3, value=service_row['Total Spend']).number_format = '$#,##0'
    ws.cell(row=row, column=3).alignment = right_align
    ws.cell(row=row, column=4, value=service_row['% of Total'] / 100).number_format = '0.0%'
    ws.cell(row=row, column=4).alignment = center_align

    for col in range(1, 5):
        ws.cell(row=row, column=col).border = border_thin
    row += 1

print(f"   ✓ Added SERVICE CATEGORIES table")

# Set column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 55
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 12

# ============================================================================
# REORDER SHEETS
# ============================================================================
print("\n3. Reordering sheets...")
# Move Source Data to second position (after Executive Summary)
sheet_order = ["Executive Summary", "Source Data", "Methodology & Interpretation"]
ordered_sheets = []
for sheet_name in sheet_order:
    if sheet_name in wb.sheetnames:
        ordered_sheets.append(wb[sheet_name])

wb._sheets = ordered_sheets
print("   ✓ Sheet order: Executive Summary → Source Data → Methodology & Interpretation")

# Save workbook
wb.save(wb_path)

print("\n" + "="*80)
print("WORKBOOK CLEANED - DATA TABLES ONLY")
print("="*80)
print(f"\nUpdated file: {wb_path}")
print("\nFinal Structure:")
print("  Sheet 1: Executive Summary")
print("  Sheet 2: Source Data")
print("    ✓ Table 1: Top 14 Vendors (ranked, with $ and %)")
print("    ✓ Table 2: Using Departments (ranked, with $ and %)")
print("    ✓ Table 3: Service Categories (ranked, with $ and %)")
print("  Sheet 3: Methodology & Interpretation")
print("\nAll tables formatted with:")
print("  • Blue headers with white text")
print("  • Borders on all cells")
print("  • Dollar amounts with accounting format ($#,##0)")
print("  • Percentages with 1 decimal place (0.0%)")
print("  • Proper alignment (left for text, right for numbers, center for %)")
print("\nReady for manual chart creation!")
print("="*80)
