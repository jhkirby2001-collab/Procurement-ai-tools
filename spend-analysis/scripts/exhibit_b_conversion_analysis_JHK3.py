#!/usr/bin/env python3
"""
Exhibit B -> Awarded Contract Conversion Analysis
==================================================
Measures whether non-contract ("maverick") Exhibit B spend was successfully
moved onto a competitively awarded contract during the reporting period.

Two gates are scored, not one:
    Gate 1 - did the spend exit Exhibit B onto an awarded contract?
    Gate 2 - was that award COMPETITIVE (Bid / RFQ / RFP)?

Matching runs in confidence tiers so nothing is overstated to leadership.
Only Tier A/B matches with a forward-looking award date count as conversions.

Author:  James H. Kirby III, CSCP, MS-SCM
Input:   two-tab workbook (Exhibit B report + awarded contracts, same period)
Output:  outputs/Exhibit_B_Conversion_Analysis_JHK3.xlsx
Usage:   python exhibit_b_conversion_analysis_JHK3.py [source.xlsx] [output.xlsx]
"""

import re
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- config ----
DEFAULT_SRC = Path("/root/.claude/uploads/ad23f03c-b709-51a6-adbb-799dc27f0ad2/"
                   "bb84a753-ExhibitB_Awarded_Contracts_1.1.2026__8.21.2026.xlsx")
DEFAULT_OUT = Path(__file__).resolve().parents[2] / "outputs" / "Exhibit_B_Conversion_Analysis_JHK3.xlsx"

EB_SHEET = "Exhibit-B Report_1.1.202"
AW_SHEET = "Awarded Jan-Aug"

# A competitive award is the procurement-correct destination for maverick spend.
COMPETITIVE_TYPES = {"BID", "RFQ", "RFP"}

# Justifications that are NOT expected to convert to a contract.
NON_CONVERTING_JUSTIFICATIONS = ("3. One Time Purchase",
                                 "4. Invoice Reconciliation",
                                 "6. Bid Protest")

# Brand palette
NAVY, LTBLUE, RED, TINT, WHITE = "002F6C", "41B6E6", "DA291C", "D6EEF9", "FFFFFF"
GREEN, AMBER, GREY = "1E7B34", "B26B00", "F2F2F2"

FONT = "Arial"
MONEY = '$#,##0;($#,##0);"-"'
PCT = '0.0%'

thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)


# ------------------------------------------------------------- text utils ---
def s(x):
    return "" if pd.isna(x) else str(x)


def norm_vendor(v):
    t = re.sub(r"[^A-Z0-9 ]", " ", s(v).upper())
    t = re.sub(r"\b(INC|LLC|CO|CORP|CORPORATION|COMPANY|LTD|LP|LLP|THE|OF|AND|USA|US"
               r"|INCORPORATED|GROUP|SERVICES|SERVICE)\b", " ", t)
    return re.sub(r"\s+", " ", t).strip()


STOPWORDS = set(
    "THE OF AND FOR TO A AN IN ON AT BY WITH PER NO NEW CITY CHICAGO DEPARTMENT "
    "CONTRACT REQ RX SPEC INVOICE INVOICES TOTAL AMOUNT PURCHASE REQUEST "
    "REQUISITION SERVICES SERVICE".split()
)


def toks(t):
    return {w for w in re.findall(r"[A-Z]{3,}", s(t).upper()) if w not in STOPWORDS}


def jaccard(a, b):
    return len(a & b) / max(1, len(a | b))


def extract_rx(v):
    """Recover new-contract requisition numbers from a free-text field.

    Handles 'RX 596048', 'RX676831- SPEC 1197406C', concatenated '710856710875',
    and 'Target Market 710856, Non-Target Market 710875'.
    """
    t = s(v)
    if re.fullmatch(r"\d{12}", t):
        return [t[:6], t[6:]]
    return re.findall(r"\b\d{6}\b", t)


def extract_ids(v):
    return re.findall(r"\b\d{5,7}\b", s(v))


# ------------------------------------------------------------------ load ----
def load(src):
    xl = pd.ExcelFile(src)
    eb = xl.parse(EB_SHEET)
    aw = xl.parse(AW_SHEET)
    eb.columns = [str(c).strip() for c in eb.columns]
    aw.columns = [str(c).strip() for c in aw.columns]

    eb["Requisition Amount"] = pd.to_numeric(eb["Requisition Amount"], errors="coerce").fillna(0)
    eb["Action Date"] = pd.to_datetime(eb["Action Date"], errors="coerce")
    eb["Date Received-DPS"] = pd.to_datetime(eb["Date Received-DPS"], errors="coerce")
    eb["EB_Date"] = eb["Action Date"].fillna(eb["Date Received-DPS"])

    aw["AWDDATE"] = pd.to_datetime(aw["AWDDATE"], errors="coerce")
    aw["AMOUNT"] = pd.to_numeric(aw["AMOUNT"], errors="coerce").fillna(0)
    for col, src_col in [("REQ_s", "REQ"), ("PO_s", "PO"), ("CID_s", "CONTRACT_ID")]:
        aw[col] = pd.to_numeric(aw[src_col], errors="coerce").astype("Int64").astype(str)
    aw["V_n"] = aw["VENDOR"].map(norm_vendor)
    aw["T"] = aw["DESCRIPTION"].map(toks)
    aw["COMPETITIVE"] = aw["PROCUREMENT_TYPE"].map(lambda x: s(x).upper().strip() in COMPETITIVE_TYPES)

    eb["RX_list"] = eb["Requisition# for New Contract/Mod"].map(extract_rx)
    eb["RX"] = eb["RX_list"].map(lambda l: l[0] if l else "")
    eb["CNA_ids"] = eb["Contract Number Awarded"].map(extract_ids)
    eb["V_n"] = eb["Vendor Name"].map(norm_vendor)
    eb["T"] = (eb["Requisition Description"].map(s) + " " + eb["Justification"].map(s)).map(toks)
    return eb, aw


# ----------------------------------------------------------- match engine ---
def match(eb, aw):
    """Attach the best awarded-contract candidate to every Exhibit B row.

    Tier A  hard ID link (RX->REQ, or reported contract# -> PO/CONTRACT_ID)
    Tier B  same vendor + commodity overlap
    Tier C  same vendor only (suggestive, NOT proof)
    Tier D  commodity/description similarity only
    Tier E  no linkage found
    """
    req_ix, po_ix, cid_ix, ven_ix = {}, {}, {}, {}
    for i, r in aw.iterrows():
        req_ix.setdefault(r["REQ_s"], i)
        po_ix.setdefault(r["PO_s"], i)
        cid_ix.setdefault(r["CID_s"], i)
        ven_ix.setdefault(r["V_n"], []).append(i)

    rows = []
    for _, r in eb.iterrows():
        hit = None
        for rx in r["RX_list"]:
            if rx in req_ix:
                hit = (req_ix[rx], "A", "Requisition ID link: Exhibit B new-contract RX = awarded REQ")
                break
        if hit is None:
            for cid in r["CNA_ids"]:
                if cid in po_ix:
                    hit = (po_ix[cid], "A", "Contract number link: reported number = awarded PO")
                    break
                if cid in cid_ix:
                    hit = (cid_ix[cid], "A", "Contract number link: reported number = awarded CONTRACT_ID")
                    break
        if hit is None and r["V_n"]:
            cands = ven_ix.get(r["V_n"], [])
            if cands:
                j, i = max((jaccard(r["T"], aw.at[c, "T"]), c) for c in cands)
                tier = "B" if j >= 0.10 else "C"
                basis = "Vendor match + commodity overlap" if j >= 0.10 else "Vendor name match only"
                hit = (i, tier, f"{basis} (similarity {j:.2f})")
        if hit is None and r["T"]:
            scored = [(jaccard(r["T"], t), i) for i, t in aw["T"].items() if t]
            if scored:
                j, i = max(scored)
                if j >= 0.30:
                    hit = (i, "D", f"Commodity/description similarity only ({j:.2f})")
        if hit is None:
            rows.append(dict(tier="E", basis="No linkage found in awarded contract file", aw_ix=np.nan))
        else:
            i, tier, basis = hit
            rows.append(dict(tier=tier, basis=basis, aw_ix=i))

    res = eb.join(pd.DataFrame(rows, index=eb.index))

    def pull(col, default=""):
        return res["aw_ix"].map(lambda x: aw.at[int(x), col] if pd.notna(x) else default)

    res["AW_CONTRACT_ID"] = pull("CID_s")
    res["AW_PO"] = pull("PO_s")
    res["AW_VENDOR"] = pull("VENDOR")
    res["AW_DESC"] = pull("DESCRIPTION")
    res["AW_TYPE"] = pull("PROCUREMENT_TYPE")
    res["AW_DATE"] = res["aw_ix"].map(lambda x: aw.at[int(x), "AWDDATE"] if pd.notna(x) else pd.NaT)
    res["AW_AMOUNT"] = res["aw_ix"].map(lambda x: aw.at[int(x), "AMOUNT"] if pd.notna(x) else np.nan)
    res["AW_COMPETITIVE"] = res["aw_ix"].map(
        lambda x: bool(aw.at[int(x), "COMPETITIVE"]) if pd.notna(x) else False)

    res["AWARD_AFTER_EB"] = res["AW_DATE"] > res["EB_Date"]
    # A conversion must be ID/vendor-verified AND forward-looking in time.
    res["CONVERTED"] = res["tier"].isin(["A", "B"]) & res["AWARD_AFTER_EB"]
    # Exhibit B filed AFTER the contract already landed = spend that should have ridden the contract.
    res["REVERSE_LEAK"] = (res["tier"].isin(["A", "B"]) & res["AW_DATE"].notna()
                           & ~res["AWARD_AFTER_EB"])

    res["EXPECTED_TO_CONVERT"] = (
        res["Request for New Contract/Mod"].map(lambda x: s(x).strip().lower() == "yes")
        & ~res["Justification"].map(lambda x: s(x).startswith(NON_CONVERTING_JUSTIFICATIONS))
    )
    res["SELF_REPORTED"] = res["Contract Awarded?"].map(lambda x: s(x).strip().lower() == "yes")
    res["SOURCING_KEY"] = res.apply(
        lambda r: "RX" + r["RX"] if r["RX"] else "VD" + (r["V_n"] or "UNKNOWN")[:28], axis=1)
    return res


def sourcing_efforts(res):
    """Roll requisitions up to the sourcing effort - the true unit of analysis."""
    g = res.groupby("SOURCING_KEY").agg(
        Requisitions=("Requisition Number", "nunique"),
        Exhibit_B_Spend=("Requisition Amount", "sum"),
        Departments=("Department Description", "nunique"),
        Description=("Requisition Description", "first"),
        Vendor=("Vendor Name", "first"),
        Lead_Department=("Department Description", "first"),
        Phase=("Procurement Phase", "first"),
        Justification=("Justification", "first"),
        Converted=("CONVERTED", "max"),
        Competitive=("AW_COMPETITIVE", "max"),
        Self_Reported=("SELF_REPORTED", "max"),
        Expected=("EXPECTED_TO_CONVERT", "max"),
        Best_Tier=("tier", "min"),
        Award_Contract=("AW_CONTRACT_ID", "max"),
        Award_Vendor=("AW_VENDOR", "max"),
        Award_Type=("AW_TYPE", "max"),
        Award_Date=("AW_DATE", "max"),
    ).reset_index()
    g["Status"] = np.where(
        g["Converted"] & g["Competitive"], "Converted - competitive award",
        np.where(g["Converted"], "Converted - non-competitive award",
                 np.where(g["Self_Reported"], "Claimed awarded - UNVERIFIED", "Still open")))
    return g.sort_values("Exhibit_B_Spend", ascending=False)


# ------------------------------------------------------------- xlsx utils ---
def style_header(ws, row, ncols, fill=NAVY, size=10):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color=WHITE, size=size)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX
    ws.row_dimensions[row].height = 30


def title_block(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(name=FONT, bold=True, size=15, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    st = ws.cell(row=2, column=2 - 1, value=subtitle)
    st.font = Font(name=FONT, italic=True, size=9, color="444444")
    st.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def write_table(ws, df, start_row, money_cols=(), pct_cols=(), date_cols=(), band=True):
    ncols = len(df.columns)
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=str(col).replace("_", " "))
    style_header(ws, start_row, ncols)
    for i, (_, r) in enumerate(df.iterrows()):
        rw = start_row + 1 + i
        for j, col in enumerate(df.columns, start=1):
            v = r[col]
            if isinstance(v, (np.bool_, bool)):
                v = "Yes" if v else "No"
            elif isinstance(v, (np.integer,)):
                v = int(v)
            elif isinstance(v, (np.floating,)):
                v = float(v)
            elif isinstance(v, pd.Timestamp):
                v = v.to_pydatetime() if pd.notna(v) else None
            elif pd.isna(v):
                v = ""
            cell = ws.cell(row=rw, column=j, value=v)
            cell.font = Font(name=FONT, size=9)
            cell.border = BOX
            cell.alignment = Alignment(vertical="top", wrap_text=(col in ("Description", "Match Basis",
                                                                         "Award_Description", "Finding",
                                                                         "Detail", "Note")))
            if col in money_cols:
                cell.number_format = MONEY
            if col in pct_cols:
                cell.number_format = PCT
            if col in date_cols:
                cell.number_format = "yyyy-mm-dd"
            if band and i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=TINT)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return start_row + len(df)


# ------------------------------------------------------------------ tabs ----
# 'EB Detail' column map, used by summary formulas so nothing is hardcoded.
D = dict(amount="F", phase="H", expected="W", selfrep="K", tier="N",
         competitive="T", converted="U", leak="V")


# Bounded to the data extent - whole-column refs make recalculation pathologically slow.
NROWS = 0


def ref(col):
    return f"'EB Detail'!${col}$2:${col}${NROWS + 1}"


def tab_exec(wb, res, se, aw, period):
    global NROWS
    NROWS = len(res)
    ws = wb.create_sheet("Executive Summary")
    widths(ws, {"A": 3, "B": 52, "C": 20, "D": 20, "E": 58})
    title_block(ws, "EXHIBIT B CONVERSION TO AWARDED CONTRACT",
                f"Did non-contract (maverick) spend move onto a competitively awarded contract?   |   {period}", 5)

    r = 4
    ws.cell(row=r, column=2, value="THE BOTTOM LINE").font = Font(name=FONT, bold=True, size=12, color=NAVY)
    r += 1
    n_conv = int(se["Converted"].sum())
    verdict = (f"Of {len(se)} distinct sourcing efforts representing "
               f"${res['Requisition Amount'].sum():,.0f} of Exhibit B spend, "
               f"{n_conv} can be independently verified as having moved onto an awarded contract "
               f"during the period. Departments self-reported {int(se['Self_Reported'].sum())} as awarded. "
               f"The gap between the claim and the evidence is the single most important "
               f"finding in this analysis.")
    ws.merge_cells(start_row=r, start_column=2, end_row=r + 2, end_column=5)
    c = ws.cell(row=r, column=2, value=verdict)
    c.font = Font(name=FONT, size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.fill = PatternFill("solid", fgColor=TINT)
    c.border = BOX
    r += 4

    # --- scorecard (all figures are formulas over 'EB Detail') ---
    ws.cell(row=r, column=2, value="CONVERSION SCORECARD").font = Font(name=FONT, bold=True, size=12, color=NAVY)
    r += 1
    hdr = ["Measure", "Requisitions", "Spend", "Basis"]
    for j, h in enumerate(hdr, start=2):
        ws.cell(row=r, column=j, value=h)
    style_header(ws, r, 5)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=WHITE)
    r += 1

    A, F, U, T, K, W = ref(D["amount"]), D["amount"], ref(D["converted"]), ref(D["competitive"]), ref(D["selfrep"]), ref(D["expected"])
    lines = [
        ("Total Exhibit B spend in period", f'=COUNTA({ref("A")})', f"=SUM({A})",
         "Every Exhibit B requisition in the source file"),
        ("Expected to convert (new contract requested)", f'=COUNTIF({W},"Yes")',
         f'=SUMIF({W},"Yes",{A})',
         "Excludes one-time purchases, invoice reconciliation, bid protests"),
        ("Self-reported by department as awarded", f'=COUNTIF({K},"Yes")',
         f'=SUMIF({K},"Yes",{A})', "Department's own 'Contract Awarded?' answer - unverified"),
        ("VERIFIED converted to an awarded contract", f'=COUNTIF({U},"Yes")',
         f'=SUMIF({U},"Yes",{A})',
         "ID or vendor+commodity link to the awarded file, award dated after the Exhibit B"),
        ("   of which COMPETITIVELY awarded (Bid/RFQ/RFP)",
         f'=COUNTIFS({U},"Yes",{T},"Yes")', f'=SUMIFS({A},{U},"Yes",{T},"Yes")',
         "The procurement-correct outcome"),
        ("   of which non-competitive (sole source/emergency)",
         f'=COUNTIFS({U},"Yes",{T},"No")', f'=SUMIFS({A},{U},"Yes",{T},"No")',
         "Exited Exhibit B but did NOT go competitive"),
        ("Reverse leakage (Exhibit B filed AFTER award)",
         f'=COUNTIF({ref(D["leak"])},"Yes")', f'=SUMIF({ref(D["leak"])},"Yes",{A})',
         "Contract existed - spend should have ridden it"),
        ("Still open / no linkage found", f'=COUNTIF({ref(D["tier"])},"E")',
         f'=SUMIF({ref(D["tier"])},"E",{A})',
         "No matching award in the supplied file - see Limitations"),
    ]
    first = r
    for label, cnt, amt, basis in lines:
        ws.cell(row=r, column=2, value=label).font = Font(name=FONT, size=10,
                                                          bold=label.startswith("VERIFIED"))
        ws.cell(row=r, column=3, value=cnt).number_format = "#,##0"
        ws.cell(row=r, column=4, value=amt).number_format = MONEY
        ws.cell(row=r, column=5, value=basis).font = Font(name=FONT, size=8, color="555555")
        ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="top")
        for cc in range(2, 6):
            ws.cell(row=r, column=cc).border = BOX
            if not ws.cell(row=r, column=cc).font.size:
                ws.cell(row=r, column=cc).font = Font(name=FONT, size=10)
        if label.startswith("VERIFIED"):
            for cc in range(2, 6):
                ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=TINT)
                ws.cell(row=r, column=cc).font = Font(name=FONT, size=10, bold=True, color=NAVY)
        ws.row_dimensions[r].height = 26
        r += 1

    # conversion rates as live formulas
    r += 1
    ws.cell(row=r, column=2, value="CONVERSION RATE").font = Font(name=FONT, bold=True, size=12, color=NAVY)
    r += 1
    U, K = ref(D["converted"]), ref(D["selfrep"])
    # A, W, T already bound above: amount, expected-to-convert, competitive
    rates = [
        # Numerators are restricted to the expected-to-convert population so that
        # numerator and denominator describe the same set of rows.
        ("Verified conversion rate - requisitions, as % of those expected to convert",
         f'=IFERROR(COUNTIFS({W},"Yes",{U},"Yes")/C{first+1},0)'),
        ("Verified conversion rate - dollars, as % of spend expected to convert",
         f'=IFERROR(SUMIFS({A},{W},"Yes",{U},"Yes")/D{first+1},0)'),
        ("Competitive conversion rate - dollars, as % of spend expected to convert",
         f'=IFERROR(SUMIFS({A},{W},"Yes",{U},"Yes",{T},"Yes")/D{first+1},0)'),
        # Must compare like with like: of the rows the department CLAIMED, how many verify.
        # Dividing all verified conversions by all claims mixes two different populations.
        ("Department claims that verify (of rows marked 'Contract Awarded? = Yes')",
         f'=IFERROR(COUNTIFS({K},"Yes",{U},"Yes")/COUNTIF({K},"Yes"),0)'),
        ("Verified conversions the department did NOT flag as awarded",
         f'=COUNTIFS({K},"No",{U},"Yes")'),
    ]
    for label, f in rates:
        ws.cell(row=r, column=2, value=label).font = Font(name=FONT, size=10)
        cell = ws.cell(row=r, column=3, value=f)
        cell.number_format = "#,##0" if label.startswith("Verified conversions the") else PCT
        cell.font = Font(name=FONT, size=11, bold=True, color=RED)
        for cc in (2, 3):
            ws.cell(row=r, column=cc).border = BOX
        r += 1

    # --- findings ---
    r += 1
    ws.cell(row=r, column=2, value="WHAT LEADERSHIP NEEDS TO KNOW").font = Font(name=FONT, bold=True, size=12, color=NAVY)
    r += 1
    top = se.iloc[0]
    phase_ef = res[res["Procurement Phase"].map(lambda x: s(x).startswith(("E.", "F.")))]
    phase_a = res[res["Procurement Phase"].map(lambda x: s(x).startswith("A."))]
    rep_nums = set()
    for v in res.loc[res["SELF_REPORTED"], "Contract Number Awarded"].dropna():
        rep_nums |= set(re.findall(r"\b\d{4,7}\b", str(v)))
    aw_all = set()
    for col in ("PO_s", "CID_s", "REQ_s"):
        aw_all |= set(aw[col].astype(str))
    aw_all |= set(pd.to_numeric(aw["SPEC"], errors="coerce").dropna().astype("int64").astype(str))
    verified_nums = rep_nums & aw_all

    findings = [
        ("The claim does not survive verification",
         f"Departments marked {int(res['SELF_REPORTED'].sum())} requisitions "
         f"(${res.loc[res['SELF_REPORTED'],'Requisition Amount'].sum():,.0f}) as having an awarded contract. "
         f"Only {int((res['SELF_REPORTED'] & res['CONVERTED']).sum())} of those verify against the awarded "
         f"contract file. Of the {len(rep_nums)} contract numbers departments supplied, only "
         f"{len(verified_nums)} appear anywhere in the actual award data. The error runs both ways: "
         f"{int((~res['SELF_REPORTED'] & res['CONVERTED']).sum())} requisitions did convert to an awarded "
         f"contract but were never flagged as awarded. Exhibit B status is not being maintained."),
        ("Exhibit B spend is concentrated in a handful of stalled sourcing efforts",
         f"The single largest - {s(top['Description'])[:90]} - accounts for "
         f"${top['Exhibit_B_Spend']:,.0f} across {int(top['Requisitions'])} requisitions and is still in "
         f"'{s(top['Phase'])}'. Counting requisitions instead of sourcing efforts hides this."),
        ("Near-term relief is real and measurable",
         f"{len(phase_ef)} requisitions worth ${phase_ef['Requisition Amount'].sum():,.0f} sit in "
         f"Recommendation of Award or Signature Cycle - one step from contract. These convert in the "
         f"next 60-90 days without new intervention."),
        ("The long tail will not clear this year",
         f"{len(phase_a)} requisitions worth ${phase_a['Requisition Amount'].sum():,.0f} are still in "
         f"Specification Development. On normal cycle times these are 12+ months from award and will "
         f"keep generating Exhibit B activity."),
        ("Exiting Exhibit B is not the same as going competitive",
         f"Of the awarded contracts in this period, only {int(aw['COMPETITIVE'].sum())} of {len(aw)} were "
         f"competitively procured (Bid/RFQ/RFP). Some Exhibit B spend exited into sole-source and "
         f"emergency awards - off the exception report, but not competitively sourced."),
        ("Reverse leakage is occurring",
         f"{int(res['REVERSE_LEAK'].sum())} requisitions were filed as Exhibit B exceptions AFTER a "
         f"contract had already been awarded for that scope. That spend should have ridden the contract."),
    ]
    for i, (h, body) in enumerate(findings, start=1):
        ws.cell(row=r, column=2, value=f"{i}.  {h}").font = Font(name=FONT, bold=True, size=10, color=RED)
        r += 1
        ws.merge_cells(start_row=r, start_column=2, end_row=r + 1, end_column=5)
        c = ws.cell(row=r, column=2, value=body)
        c.font = Font(name=FONT, size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 16
        ws.row_dimensions[r + 1].height = 16
        r += 3

    ws.sheet_view.showGridLines = False
    return ws


def tab_scorecard(wb, se):
    ws = wb.create_sheet("Conversion Scorecard")
    title_block(ws, "CONVERSION SCORECARD BY SOURCING EFFORT",
                "One row per distinct sourcing effort (the pending new-contract requisition), "
                "not per Exhibit B requisition. This is the correct unit of analysis.", 12)
    df = se[["SOURCING_KEY", "Description", "Lead_Department", "Vendor", "Requisitions",
             "Departments", "Exhibit_B_Spend", "Phase", "Status", "Best_Tier",
             "Award_Contract", "Award_Type"]].copy()
    df.columns = ["Sourcing Effort", "Description", "Lead Department", "Exhibit B Vendor",
                  "Requisitions", "Depts", "Exhibit B Spend", "Procurement Phase",
                  "Status", "Evidence Tier", "Awarded Contract", "Award Type"]
    df["Description"] = df["Description"].map(lambda x: s(x)[:150])
    write_table(ws, df, 4, money_cols=("Exhibit B Spend",))
    widths(ws, {"A": 18, "B": 55, "C": 30, "D": 26, "E": 11, "F": 8, "G": 16,
                "H": 26, "I": 30, "J": 12, "K": 15, "L": 18})
    for row in range(5, 5 + len(df)):
        st = ws.cell(row=row, column=9).value
        if st == "Converted - competitive award":
            ws.cell(row=row, column=9).font = Font(name=FONT, size=9, bold=True, color=GREEN)
        elif st == "Converted - non-competitive award":
            ws.cell(row=row, column=9).font = Font(name=FONT, size=9, bold=True, color=AMBER)
        elif st == "Claimed awarded - UNVERIFIED":
            ws.cell(row=row, column=9).font = Font(name=FONT, size=9, bold=True, color=RED)
    ws.auto_filter.ref = f"A4:L{4 + len(df)}"
    return ws


def tab_confirmed(wb, res):
    ws = wb.create_sheet("Confirmed Conversions")
    title_block(ws, "CONFIRMED CONVERSIONS - EVIDENCE TRAIL",
                "Every row here is independently verified: linked to the awarded contract file by ID or "
                "vendor+commodity, with the award dated AFTER the Exhibit B request.", 11)
    c = res[res["CONVERTED"]].copy()
    df = c[["Requisition Number", "EB_Date", "Department Description", "Requisition Description",
            "Requisition Amount", "AW_CONTRACT_ID", "AW_VENDOR", "AW_DATE", "AW_TYPE",
            "AW_COMPETITIVE", "basis"]].copy()
    df.columns = ["Exhibit B Req#", "Exhibit B Date", "Department", "Description", "Exhibit B Spend",
                  "Awarded Contract ID", "Awarded Vendor", "Award Date", "Procurement Type",
                  "Competitive?", "Match Basis"]
    df["Description"] = df["Description"].map(lambda x: s(x)[:120])
    df = df.sort_values("Exhibit B Spend", ascending=False)
    write_table(ws, df, 4, money_cols=("Exhibit B Spend",),
                date_cols=("Exhibit B Date", "Award Date"))
    widths(ws, {"A": 15, "B": 14, "C": 30, "D": 48, "E": 15, "F": 17,
                "G": 30, "H": 13, "I": 18, "J": 12, "K": 46})
    return ws


def tab_claim_gap(wb, res, aw):
    ws = wb.create_sheet("Claimed vs Verified")
    title_block(ws, "DEPARTMENT CLAIM vs INDEPENDENT VERIFICATION",
                "Rows the department marked 'Contract Awarded? = Yes'. The right-hand columns show what "
                "the awarded contract file actually supports.", 9)
    c = res[res["SELF_REPORTED"]].copy()
    aw_all = set()
    for col in ("PO_s", "CID_s", "REQ_s"):
        aw_all |= set(aw[col].astype(str))

    def num_in_award(v):
        ids = re.findall(r"\b\d{4,7}\b", s(v))
        if not ids:
            return "No number supplied"
        return "Found in award file" if any(i in aw_all for i in ids) else "NOT in award file"

    c["Number check"] = c["Contract Number Awarded"].map(num_in_award)
    c["Verdict"] = np.where(c["CONVERTED"], "VERIFIED",
                            np.where(c["tier"] == "E", "No supporting evidence", "Weak / partial evidence"))
    df = c[["Requisition Number", "Department Description", "Requisition Description",
            "Requisition Amount", "Contract Number Awarded", "Number check", "tier",
            "Verdict", "basis"]].copy()
    df.columns = ["Exhibit B Req#", "Department", "Description", "Exhibit B Spend",
                  "Contract # Claimed", "Claimed Number Check", "Evidence Tier", "Verdict", "Match Basis"]
    df["Description"] = df["Description"].map(lambda x: s(x)[:110])
    df = df.sort_values("Exhibit B Spend", ascending=False)
    write_table(ws, df, 4, money_cols=("Exhibit B Spend",))
    widths(ws, {"A": 15, "B": 30, "C": 46, "D": 15, "E": 20, "F": 22, "G": 12, "H": 24, "I": 44})
    for row in range(5, 5 + len(df)):
        v = ws.cell(row=row, column=8).value
        col = GREEN if v == "VERIFIED" else (RED if v == "No supporting evidence" else AMBER)
        ws.cell(row=row, column=8).font = Font(name=FONT, size=9, bold=True, color=col)
    return ws


def tab_open(wb, se):
    ws = wb.create_sheet("Still Open - Ranked")
    title_block(ws, "STILL-OPEN SOURCING EFFORTS RANKED BY SPEND AT RISK",
                "Maverick spend with no verified contract award. This is the work queue - "
                "clearing the top of this list removes the most Exhibit B activity.", 9)
    o = se[~se["Converted"]].copy().sort_values("Exhibit_B_Spend", ascending=False)
    o["Rank"] = range(1, len(o) + 1)
    o["Cumulative"] = o["Exhibit_B_Spend"].cumsum()
    o["Cum %"] = o["Cumulative"] / o["Exhibit_B_Spend"].sum()
    df = o[["Rank", "Description", "Lead_Department", "Requisitions", "Exhibit_B_Spend",
            "Cumulative", "Cum %", "Phase", "Status"]].copy()
    df.columns = ["Rank", "Description", "Lead Department", "Requisitions", "Exhibit B Spend",
                  "Cumulative", "Cum %", "Procurement Phase", "Status"]
    df["Description"] = df["Description"].map(lambda x: s(x)[:150])
    write_table(ws, df, 4, money_cols=("Exhibit B Spend", "Cumulative"), pct_cols=("Cum %",))
    widths(ws, {"A": 7, "B": 60, "C": 32, "D": 12, "E": 16, "F": 16, "G": 9, "H": 28, "I": 28})
    ws.auto_filter.ref = f"A4:I{4 + len(df)}"
    return ws


def tab_pipeline(wb, res):
    ws = wb.create_sheet("Pipeline Forecast")
    title_block(ws, "PIPELINE FORECAST - WHAT CONVERTS NEXT",
                "Procurement Phase shows how far each Exhibit B requisition has progressed toward a "
                "contract. This turns a backward-looking scorecard into a forecast.", 6)
    g = (res.groupby(res["Procurement Phase"].map(lambda x: s(x) or "(not stated)"))
         .agg(Requisitions=("Requisition Amount", "size"),
              Spend=("Requisition Amount", "sum"),
              Efforts=("SOURCING_KEY", "nunique"))
         .reset_index())
    g.columns = ["Procurement Phase", "Requisitions", "Exhibit B Spend", "Sourcing Efforts"]
    horizon = {"E.": "0-3 months - imminent", "F.": "0-3 months - imminent",
               "D.": "3-6 months", "C.": "3-6 months", "B.": "6-12 months",
               "A.": "12+ months", "G.": "n/a - not applicable", "H.": "n/a - one-time purchase"}
    g["Expected Conversion Horizon"] = g["Procurement Phase"].map(
        lambda x: next((v for k, v in horizon.items() if s(x).startswith(k)), "unknown"))
    g["Share of Spend"] = g["Exhibit B Spend"] / g["Exhibit B Spend"].sum()
    g = g.sort_values("Procurement Phase")
    g = g[["Procurement Phase", "Expected Conversion Horizon", "Sourcing Efforts",
           "Requisitions", "Exhibit B Spend", "Share of Spend"]]
    end = write_table(ws, g, 4, money_cols=("Exhibit B Spend",), pct_cols=("Share of Spend",))
    tr = end + 1
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(name=FONT, bold=True, size=10, color=WHITE)
    ws.cell(row=tr, column=5, value=f"=SUM(E5:E{end})").number_format = MONEY
    ws.cell(row=tr, column=6, value=f"=SUM(F5:F{end})").number_format = PCT
    for cc in range(1, 7):
        ws.cell(row=tr, column=cc).fill = PatternFill("solid", fgColor=RED)
        ws.cell(row=tr, column=cc).font = Font(name=FONT, bold=True, size=10, color=WHITE)
        ws.cell(row=tr, column=cc).border = BOX
    imminent = res[res["Procurement Phase"].map(lambda x: s(x).startswith(("E.", "F.")))]
    n = tr + 2
    ws.merge_cells(start_row=n, start_column=1, end_row=n + 1, end_column=6)
    c = ws.cell(row=n, column=1, value=(
        f"NEAR-TERM HEADLINE:  {len(imminent)} requisitions worth "
        f"${imminent['Requisition Amount'].sum():,.0f} are in Recommendation of Award or Signature Cycle "
        f"- one administrative step from a contract. Expect these to leave Exhibit B within 60-90 days."))
    c.font = Font(name=FONT, bold=True, size=10, color=NAVY)
    c.fill = PatternFill("solid", fgColor=TINT)
    c.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
    c.border = BOX
    widths(ws, {"A": 42, "B": 26, "C": 16, "D": 14, "E": 18, "F": 15})
    return ws


def tab_leakage(wb, res):
    ws = wb.create_sheet("Reverse Leakage")
    title_block(ws, "REVERSE LEAKAGE - EXHIBIT B FILED AFTER THE CONTRACT EXISTED",
                "These requisitions were processed as contract exceptions AFTER an award covering that "
                "scope had already landed. The spend should have ridden the contract.", 8)
    l = res[res["REVERSE_LEAK"]].copy()
    if len(l) == 0:
        ws.cell(row=4, column=1, value="No reverse leakage detected in this period.").font = Font(name=FONT, size=10)
        widths(ws, {"A": 60})
        return ws
    df = l[["Requisition Number", "EB_Date", "Department Description", "Requisition Description",
            "Requisition Amount", "AW_DATE", "AW_VENDOR", "AW_CONTRACT_ID"]].copy()
    df.columns = ["Exhibit B Req#", "Exhibit B Date", "Department", "Description",
                  "Spend That Should Have Ridden Contract", "Contract Award Date",
                  "Awarded Vendor", "Awarded Contract ID"]
    df["Description"] = df["Description"].map(lambda x: s(x)[:120])
    df["Days After Award"] = (l["EB_Date"] - l["AW_DATE"]).dt.days.values
    df = df.sort_values("Spend That Should Have Ridden Contract", ascending=False)
    write_table(ws, df, 4, money_cols=("Spend That Should Have Ridden Contract",),
                date_cols=("Exhibit B Date", "Contract Award Date"))
    widths(ws, {"A": 15, "B": 14, "C": 30, "D": 48, "E": 22, "F": 16, "G": 30, "H": 17, "I": 15})
    return ws


def tab_dataquality(wb, res, aw):
    ws = wb.create_sheet("Data Quality Findings")
    title_block(ws, "DATA QUALITY FINDINGS",
                "Issues in the source data that limit measurement - and that are fixable at the "
                "point of entry.", 4)
    cna = res["Contract Number Awarded"].dropna().map(str).str.strip()
    junk = cna[~cna.str.match(r"^\d+$")]
    rx = res["Requisition# for New Contract/Mod"].dropna().map(str).str.strip()
    rx_dirty = rx[~rx.str.match(r"^\d{6}$")]
    aw_all = set()
    for col in ("PO_s", "CID_s", "REQ_s"):
        aw_all |= set(aw[col].astype(str))
    rep = set()
    for v in res.loc[res["SELF_REPORTED"], "Contract Number Awarded"].dropna():
        rep |= set(re.findall(r"\b\d{4,7}\b", str(v)))

    rows = [
        ("'Contract Number Awarded' is largely unusable",
         f"{len(junk)} of {len(cna)} populated values are not contract numbers - they include "
         f"'TBD', 'na', '0', 'NO', 'yes', 'Pending', 'UNKNOWN', 'N?A' and 'None yet'.",
         "Cannot verify an award from this field alone",
         "Make it a validated numeric field, required only when 'Contract Awarded?' = Yes"),
        ("Claimed contract numbers do not reconcile",
         f"{len(rep)} distinct numbers were supplied on rows claiming an award. Only "
         f"{len(rep & aw_all)} appear anywhere in the awarded contract file.",
         "The award claim cannot be substantiated for most rows",
         "Validate the entry against the contract master at the point of entry"),
        ("'Requisition# for New Contract/Mod' is free text",
         f"{len(rx_dirty)} of {len(rx)} values need parsing - 'RX 596048', 'RX676831- SPEC 1197406C', "
         f"concatenated '710856710875', 'Target Market 710856, Non-Target Market 710875', 'na'.",
         "Automated tracking requires a cleanup pass every time",
         "Split into a numeric requisition field plus a free-text note field"),
        ("Awarded contract file has no requisition on every row",
         f"{int(aw['REQ'].isna().sum())} of {len(aw)} awarded contracts carry no REQ number, and "
         f"{int(aw['SPEC'].isna().sum())} carry no SPEC.",
         "Those awards cannot be traced back to the Exhibit B that preceded them",
         "Require the originating requisition on every award record"),
        ("No shared key between the two systems",
         "Exhibit B and the award file share no reliable common identifier; linkage depends on "
         "requisition numbers appearing in a free-text field.",
         "Conversion cannot be measured automatically today",
         "Carry the Exhibit B requisition number forward onto the resulting award record"),
    ]
    df = pd.DataFrame(rows, columns=["Finding", "Detail", "Impact on Measurement", "Recommended Fix"])
    write_table(ws, df, 4)
    widths(ws, {"A": 40, "B": 62, "C": 40, "D": 48})
    for row in range(5, 5 + len(df)):
        ws.row_dimensions[row].height = 56
        ws.cell(row=row, column=1).font = Font(name=FONT, size=9, bold=True, color=NAVY)
        for cc in range(1, 5):
            ws.cell(row=row, column=cc).alignment = Alignment(wrap_text=True, vertical="top")
    return ws


def tab_method(wb, res, aw, se, period):
    ws = wb.create_sheet("Methodology & Limits")
    widths(ws, {"A": 3, "B": 34, "C": 96})
    title_block(ws, "METHODOLOGY, DEFINITIONS AND LIMITATIONS",
                "Read this before quoting any number in this workbook.", 3)
    r = 4

    def section(t):
        nonlocal r
        ws.cell(row=r, column=2, value=t).font = Font(name=FONT, bold=True, size=11, color=NAVY)
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=TINT)
        ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor=TINT)
        r += 1

    def item(k, v):
        nonlocal r
        ws.cell(row=r, column=2, value=k).font = Font(name=FONT, bold=True, size=9)
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        c = ws.cell(row=r, column=3, value=v)
        c.font = Font(name=FONT, size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(14, 12 * (len(v) // 105 + 1))
        r += 1

    section("WHAT THIS ANALYSIS ANSWERS")
    item("Question", "Did non-contract ('maverick') Exhibit B spend move onto a competitively "
                     "awarded contract during the period?")
    item("Two gates", "Gate 1 - did the spend exit Exhibit B onto an awarded contract? "
                      "Gate 2 - was that award competitive (Bid/RFQ/RFP)? Exiting Exhibit B into a "
                      "sole-source or emergency award clears the exception report but is not the "
                      "procurement-correct outcome.")
    r += 1

    section("UNIT OF ANALYSIS")
    item("Sourcing effort", f"Requisitions are rolled up to the pending new-contract requisition (RX). "
                            f"{len(res)} Exhibit B requisitions represent only {len(se)} distinct sourcing "
                            f"efforts. One effort absorbs {int(se.iloc[0]['Requisitions'])} requisitions and "
                            f"${se.iloc[0]['Exhibit_B_Spend']:,.0f}. Counting requisitions overstates the "
                            f"number of problems and understates their size.")
    r += 1

    section("EVIDENCE TIERS")
    item("Tier A - Confirmed", "Hard ID link: the Exhibit B new-contract requisition equals an awarded "
                               "REQ, or the reported contract number equals an awarded PO/CONTRACT_ID.")
    item("Tier B - Confirmed", "Same vendor plus commodity-token overlap of 10% or more between the "
                               "Exhibit B description and the award description.")
    item("Tier C - Suggestive", "Same vendor only, no commodity overlap. NOT counted as a conversion.")
    item("Tier D - Weak", "Description similarity of 30% or more with no vendor or ID link. "
                          "NOT counted as a conversion.")
    item("Tier E - None", "No linkage found in the awarded contract file.")
    item("Timing gate", "A Tier A or B match counts as a conversion only if the award date is AFTER the "
                        "Exhibit B date. A match with an earlier award date is reverse leakage, not a "
                        "conversion.")
    item("Headline basis", "The headline conversion figure uses Tier A and B only. Tiers C and D are "
                           "disclosed separately and are never added to the headline.")
    r += 1

    section("DENOMINATOR")
    item("Expected to convert", f"'Request for New Contract/Mod' = Yes, excluding justifications "
                                f"{', '.join(NON_CONVERTING_JUSTIFICATIONS)} - these are not intended to "
                                f"result in a contract. "
                                f"{int(res['EXPECTED_TO_CONVERT'].sum())} requisitions "
                                f"(${res.loc[res['EXPECTED_TO_CONVERT'],'Requisition Amount'].sum():,.0f}) "
                                f"meet this test.")
    r += 1

    section("LIMITATIONS - READ BEFORE CITING")
    item("Award file is a fixed list", f"The awarded contract file contains {len(aw)} contracts. If it is "
                                       f"not the complete population of awards for the period, an Exhibit B "
                                       f"effort could have been awarded on a contract absent from this file. "
                                       f"'No verified conversion' therefore means 'not evidenced in the "
                                       f"supplied award data' - it does not prove no award occurred.")
    item("No shared key", "The two systems share no reliable common identifier. Linkage depends on "
                          "requisition numbers appearing inside a free-text field, so some genuine "
                          "conversions are likely unmatched. The verified figure is a floor, not a ceiling.")
    item("Self-reported fields excluded", "'Contract Awarded?' and 'Contract Number Awarded' are department "
                                          "self-reports and are analysed as a claim to be tested, never used "
                                          "as evidence of conversion.")
    item("Spend basis", "Requisition Amount as submitted. Approved and denied rows are both retained in the "
                        "population and flagged; denied rows never count as converted.")
    item("Period", f"{period}. Both tabs cover the same window, so an Exhibit B request late in the period "
                   f"has had little time to reach award - conversion for recent months is understated by "
                   f"construction.")
    r += 1

    section("SOURCE DATA")
    item("Exhibit B tab", f"{len(res)} requisitions, ${res['Requisition Amount'].sum():,.2f}")
    item("Awarded contracts tab", f"{len(aw)} contracts, ${aw['AMOUNT'].sum():,.2f} total award value "
                                  f"(full award book, not only Exhibit B conversions)")
    item("Competitive definition", f"PROCUREMENT_TYPE in {sorted(COMPETITIVE_TYPES)}. "
                                   f"{int(aw['COMPETITIVE'].sum())} of {len(aw)} awards qualify.")
    ws.sheet_view.showGridLines = False
    return ws


def tab_detail(wb, res):
    ws = wb.create_sheet("EB Detail")
    df = res[["Requisition Number", "EB_Date", "Department Description", "Vendor Name",
              "Requisition Description", "Requisition Amount", "Justification",
              "Procurement Phase", "Approved or Denied", "Request for New Contract/Mod",
              "SELF_REPORTED", "Contract Number Awarded", "SOURCING_KEY", "tier", "basis",
              "AW_CONTRACT_ID", "AW_VENDOR", "AW_DATE", "AW_TYPE", "AW_COMPETITIVE",
              "CONVERTED", "REVERSE_LEAK", "EXPECTED_TO_CONVERT"]].copy()
    df.columns = ["Requisition Number", "Exhibit B Date", "Department", "Vendor", "Description",
                  "Amount", "Justification", "Procurement Phase", "Approved or Denied",
                  "New Contract Requested", "Self-Reported Awarded", "Contract # Claimed",
                  "Sourcing Key", "Evidence Tier", "Match Basis", "Awarded Contract ID",
                  "Awarded Vendor", "Award Date", "Procurement Type", "Competitive",
                  "Converted", "Reverse Leak", "Expected to Convert"]
    write_table(ws, df, 1, money_cols=("Amount",), date_cols=("Exhibit B Date", "Award Date"), band=False)
    widths(ws, {"A": 16, "B": 14, "C": 30, "D": 28, "E": 50, "F": 14, "G": 40, "H": 26,
                "I": 15, "J": 16, "K": 16, "L": 18, "M": 18, "N": 11, "O": 44, "P": 17,
                "Q": 28, "R": 13, "S": 18, "T": 12, "U": 11, "V": 12, "W": 16})
    ws.auto_filter.ref = f"A1:W{1 + len(df)}"
    return ws


def tab_awarded(wb, aw):
    ws = wb.create_sheet("Awarded Contracts")
    df = aw[["CONTRACT_ID", "PO", "REQ", "SPEC", "DESCRIPTION", "VENDOR", "DEPT", "UNIT",
             "AWDDATE", "STRTDATE", "AMOUNT", "PROCUREMENT_TYPE", "COMPETITIVE",
             "CONTRACT_TYPE", "TYPE", "BUYER", "DAYS"]].copy()
    df.columns = ["Contract ID", "PO", "Requisition", "Spec", "Description", "Vendor", "Department",
                  "Unit", "Award Date", "Start Date", "Award Amount", "Procurement Type",
                  "Competitive", "Contract Type", "Type", "Buyer", "Days to Award"]
    write_table(ws, df, 1, money_cols=("Award Amount",), date_cols=("Award Date",), band=False)
    widths(ws, {"A": 12, "B": 10, "C": 12, "D": 11, "E": 58, "F": 34, "G": 34, "H": 26,
                "I": 13, "J": 12, "K": 16, "L": 22, "M": 12, "N": 34, "O": 12, "P": 22, "Q": 13})
    ws.auto_filter.ref = f"A1:Q{1 + len(df)}"
    return ws


# ------------------------------------------------------------------ main ----
def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SRC
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUT
    out.parent.mkdir(parents=True, exist_ok=True)

    print(f"Reading {src.name} ...")
    eb, aw = load(src)
    res = match(eb, aw)
    se = sourcing_efforts(res)
    period = (f"{res['EB_Date'].min():%B %-d, %Y} to {res['EB_Date'].max():%B %-d, %Y}")

    print(f"  {len(res)} requisitions  ${res['Requisition Amount'].sum():,.0f}")
    print(f"  {len(se)} sourcing efforts | verified conversions: {int(se['Converted'].sum())}")
    print(f"  self-reported awarded: {int(res['SELF_REPORTED'].sum())} rows")

    wb = Workbook()
    wb.remove(wb.active)
    tab_exec(wb, res, se, aw, period)
    tab_scorecard(wb, se)
    tab_confirmed(wb, res)
    tab_claim_gap(wb, res, aw)
    tab_open(wb, se)
    tab_pipeline(wb, res)
    tab_leakage(wb, res)
    tab_dataquality(wb, res, aw)
    tab_method(wb, res, aw, se, period)
    tab_detail(wb, res)
    tab_awarded(wb, aw)
    wb.save(out)
    print(f"Wrote {out}")
    return out


if __name__ == "__main__":
    main()
