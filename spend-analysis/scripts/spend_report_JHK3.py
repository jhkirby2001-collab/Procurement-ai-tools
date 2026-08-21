"""
Reusable spend-report engine for the NIGP-Sourced Procurement Category Mapper.

Takes a spend file, classifies its descriptions with the mapper's rules
(vectorized, no AI, no API key), then produces a deterministic, leadership-grade
spend report: an executive summary, summary tiles, spend-by-category, Pareto
80/20, top vendors, spend-by-department, and vendor consolidation — rendered as a
professionally formatted, brand-colored Excel workbook with charts.

Pure pandas/openpyxl. Same core is used by the Streamlit "Spend Report" page and
can be run standalone for testing:

    python spend_report_JHK3.py "data/raw/spen data1.csv" \
        --desc "Item Description" --amount "Amount Ordered" \
        --vendor "Vendor Name" --dept "Release Dept Code"
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Optional

import pandas as pd

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent))
from classifier_JHK3 import (  # noqa: E402
    load_keyword_rules,
    CONFIDENCE_BY_LEVEL,
    CONFIDENCE_LABEL,
    LOWINFO_RE,
)

# Every row lands in a real Business Category. Anything the keyword rules don't
# map to a specific commodity is placed in a single, honest catch-all so a
# leadership reader never sees an "Unclassified"/gap line. The dollars are still
# counted, and the Examples column shows exactly what fell into it.
CATCHALL = "General & Other Procurement"
# Grants is a real category, but on arbitrary uploads it must only fire on strong,
# generalizable grant signals (program-tag prefixes + grant keywords) — never on
# memorized full-description exact rules. See classify_series().
GRANTS_CATEGORY = "Grants & Pass-Through Funding"
# Internal Classification_Method tags (kept for audit; both display as CATCHALL)
METHOD_NO_RULE = "no_rule"
METHOD_NO_DESC = "no_description"
# Back-compat aliases (older callers referenced these names)
UNCLASSIFIED_NO_RULE = CATCHALL
UNCLASSIFIED_NO_DESC = CATCHALL

AMOUNT_HINTS = ["amount", "spend", "cost", "price", "total", "value", "paid",
                "billed", "ordered", "award", "expenditure", "payment"]
VENDOR_HINTS = ["vendor", "supplier", "payee", "merchant", "company", "contractor"]
DEPT_HINTS = ["department name", "dept name", "using area", "department", "dept",
              "agency", "division", "bureau", "office", "unit"]
DESC_HINTS = ["description", "item", "purpose", "commodity", "detail", "line",
              "service", "narrative"]
DATE_HINTS = ["date", "period", "fiscal", "year"]

# City of Chicago brand palette (hex, no '#')
NAVY = "002F6C"
BLUE = "41B6E6"
RED = "DA291C"
LT_BLUE = "D6EEF9"
GRAY = "595959"
WHITE = "FFFFFF"


# ---------------------------------------------------------------------------
# Column detection + currency cleaning
# ---------------------------------------------------------------------------
def _best_col(columns, hints) -> Optional[str]:
    cols = list(columns)
    low = {c: str(c).strip().lower() for c in cols}
    for h in hints:
        for c in cols:
            if low[c] == h:
                return c
    for h in hints:
        for c in cols:
            if h in low[c]:
                return c
    return None


def detect_columns(df: pd.DataFrame) -> dict:
    return {
        "desc": _best_col(df.columns, DESC_HINTS),
        "amount": _best_col(df.columns, AMOUNT_HINTS),
        "vendor": _best_col(df.columns, VENDOR_HINTS),
        "dept": _best_col(df.columns, DEPT_HINTS),
        "date": _best_col(df.columns, DATE_HINTS),
    }


def clean_amount(series: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(series):
        return series.astype(float)
    s = series.astype(str).str.strip()
    neg = s.str.startswith("(") & s.str.endswith(")")
    s = s.str.replace(r"[,$\s]", "", regex=True)
    s = s.str.replace(r"^\((.*)\)$", r"\1", regex=True)
    out = pd.to_numeric(s, errors="coerce")
    out = out.mask(neg, -out.abs())
    return out


# ---------------------------------------------------------------------------
# Vectorized classification (Tier 1 keyword rules; no AI, no account codes)
# ---------------------------------------------------------------------------
def classify_series(desc: pd.Series, rules_df=None) -> pd.DataFrame:
    if rules_df is None:
        rules_df = load_keyword_rules()
    desc = desc.fillna("").astype(str).str.strip()
    desc_u = desc.str.upper()
    n = len(desc)
    out = pd.DataFrame({
        "Business_Category": [""] * n,
        "NIGP_Class_3digit": [""] * n,
        "NIGP_Item_5digit": [""] * n,
        "NIGP_Code_Assigned": [""] * n,
        "Classification_Method": [""] * n,
        "Classification_Confidence": [""] * n,
        "Classification_Reason": [""] * n,
    }, index=desc.index)

    unclassified = pd.Series(True, index=desc.index)
    for _, rule in rules_df.iterrows():
        if not unclassified.any():
            break
        pat_u = rule["pattern_upper"]
        mt = rule["match_type"]
        # Upload-path scoping for Grants & Pass-Through Funding: only a genuine,
        # generalizable grant signal may route to Grants — a program-tag prefix
        # (starts_with) or an explicit grant keyword (contains). Full-description
        # `exact` rules memorized one agency's specific PO strings and misfire on
        # an arbitrary uploaded file (e.g. a sewer-improvement or JOC-closeout row
        # that merely shares a memorized string). Let those fall through to their
        # real commodity rule or the General & Other catch-all instead.
        if rule["business_category"] == GRANTS_CATEGORY and mt == "exact":
            continue
        if mt == "exact":
            mask = unclassified & (desc_u == pat_u)
        elif mt == "starts_with":
            mask = unclassified & desc_u.str.startswith(pat_u, na=False)
        else:
            mask = unclassified & desc_u.str.contains(re.escape(pat_u), regex=True, na=False)
        if not mask.any():
            continue
        level = rule["nigp_match_level"]
        score = CONFIDENCE_BY_LEVEL.get(level, 0.50)
        code = str(rule["nigp_item_5digit"]) or str(rule["nigp_class_3digit"])
        out.loc[mask, "Business_Category"] = rule["business_category"]
        out.loc[mask, "NIGP_Class_3digit"] = str(rule["nigp_class_3digit"])
        out.loc[mask, "NIGP_Item_5digit"] = str(rule["nigp_item_5digit"])
        out.loc[mask, "NIGP_Code_Assigned"] = code
        out.loc[mask, "Classification_Method"] = "keyword_rule"
        out.loc[mask, "Classification_Confidence"] = CONFIDENCE_LABEL(score)
        out.loc[mask, "Classification_Reason"] = f"matched '{rule['pattern']}' ({mt})"
        unclassified &= ~mask

    no_desc = unclassified & (desc == "")
    no_rule = unclassified & (desc != "")
    # Both fall into a single real catch-all category — never an "Unclassified" gap.
    out.loc[no_rule, "Business_Category"] = CATCHALL
    out.loc[no_rule, "Classification_Method"] = METHOD_NO_RULE
    out.loc[no_rule, "Classification_Reason"] = "no specific commodity rule matched — grouped as General & Other"
    out.loc[no_desc, "Business_Category"] = CATCHALL
    out.loc[no_desc, "Classification_Method"] = METHOD_NO_DESC
    out.loc[no_desc, "Classification_Reason"] = "no description text on this row — grouped as General & Other"
    return out


# ---------------------------------------------------------------------------
# Analyzers (pure pandas, column-parameterized)
# ---------------------------------------------------------------------------
def _is_classified(cat: pd.Series) -> pd.Series:
    """A row is 'classified' when it mapped to a specific commodity category.
    The General & Other catch-all is a real reporting category, but it is not a
    specific commodity, so it is excluded from consolidation / Pareto-of-specifics."""
    return cat != CATCHALL


def _truncate(text, max_len):
    text = str(text).strip()
    return text if len(text) <= max_len else text[: max_len - 1].rstrip() + "…"


def category_examples(df, category_col, desc_col, per_cat=3, max_len=42) -> dict:
    """For each Business Category, return a short string of representative example
    descriptions (most frequent first, de-duplicated, truncated). Gives leadership
    visibility into what actually sits inside every category — especially the
    General & Other catch-all."""
    if not desc_col or desc_col not in df:
        return {}
    d = df[[category_col, desc_col]].copy()
    d[desc_col] = d[desc_col].fillna("").astype(str).str.strip()
    d = d[d[desc_col] != ""]
    out = {}
    for cat, grp in d.groupby(category_col):
        vc = grp[desc_col].value_counts()
        picks = [_truncate(t, max_len) for t in vc.index[:per_cat]]
        out[cat] = "  ·  ".join(picks)
    return out


EXAMPLES_COL = "Examples of what's in this category"


def summary_tiles(df, category_col, amount_col=None, vendor_col=None,
                  date_col=None) -> dict:
    tiles = {"transactions": int(len(df)),
             "categories": int(df[category_col].nunique())}
    if amount_col and amount_col in df:
        amt = clean_amount(df[amount_col])
        tiles["total_spend"] = float(amt.sum(skipna=True))
        tiles["has_amount"] = True
    else:
        tiles["total_spend"] = None
        tiles["has_amount"] = False
    tiles["vendors"] = int(df[vendor_col].nunique()) if vendor_col and vendor_col in df else None
    if date_col and date_col in df:
        d = pd.to_datetime(df[date_col], errors="coerce")
        if d.notna().any():
            tiles["date_min"] = str(d.min().date())
            tiles["date_max"] = str(d.max().date())
    return tiles


def _rollup(df, group_col, amount_col):
    work = df.copy()
    if amount_col and amount_col in work:
        work["_amt"] = clean_amount(work[amount_col]).fillna(0.0)
        g = work.groupby(group_col)["_amt"].agg(["sum", "count"])
        g.columns = ["Spend", "Transactions"]
        g = g.sort_values("Spend", ascending=False)
        measure = "Spend"
    else:
        g = work.groupby(group_col).size().to_frame("Transactions")
        g = g.sort_values("Transactions", ascending=False)
        measure = "Transactions"
    total = g[measure].sum()
    g["% of Total"] = (g[measure] / total * 100).round(1) if total else 0.0
    return g.reset_index(), measure


def spend_by_category(df, category_col, amount_col=None, desc_col=None,
                      examples_per_cat=3):
    g, _ = _rollup(df, category_col, amount_col)
    g = g.rename(columns={category_col: "Business Category"})
    if desc_col and desc_col in df:
        ex = category_examples(df, category_col, desc_col, per_cat=examples_per_cat)
        g[EXAMPLES_COL] = g["Business Category"].map(ex).fillna("")
        g.loc[g[EXAMPLES_COL] == "", EXAMPLES_COL] = "(no description text on these rows)"
    return g


def spend_by_department(df, dept_col, amount_col=None, top_n=25):
    if not dept_col or dept_col not in df:
        return None
    g, _ = _rollup(df, dept_col, amount_col)
    g = g.rename(columns={dept_col: "Department"})
    g["Department"] = g["Department"].astype(str)
    return g.head(top_n) if top_n else g


def pareto(df, group_col, amount_col=None, top_n=None):
    g, measure = _rollup(df, group_col, amount_col)
    g["Cumulative %"] = g["% of Total"].cumsum().round(1)
    n80 = int((g["Cumulative %"] < 80).sum() + 1) if len(g) else 0
    n80 = min(n80, len(g))
    if top_n:
        g = g.head(top_n)
    return g.rename(columns={group_col: "Group"}), measure, n80


def top_vendors(df, vendor_col, amount_col=None, n=15):
    if not vendor_col or vendor_col not in df:
        return None
    g, _ = _rollup(df, vendor_col, amount_col)
    return g.rename(columns={vendor_col: "Vendor"}).head(n)


def consolidation_finder(df, category_col, vendor_col, amount_col=None,
                         dept_col=None, min_vendors=2):
    if not vendor_col or vendor_col not in df:
        return None
    work = df[_is_classified(df[category_col])].copy()
    if amount_col and amount_col in work:
        work["_amt"] = clean_amount(work[amount_col]).fillna(0.0)
        g = work.groupby(category_col).agg(**{
            "Spend": ("_amt", "sum"),
            "Vendors": (vendor_col, "nunique"),
            "Transactions": (vendor_col, "size"),
        })
    else:
        g = work.groupby(category_col).agg(**{
            "Vendors": (vendor_col, "nunique"),
            "Transactions": (vendor_col, "size"),
        })
    if dept_col and dept_col in work:
        g["Departments"] = work.groupby(category_col)[dept_col].nunique()
    g = g[g["Vendors"] >= min_vendors]
    sort_key = "Spend" if "Spend" in g.columns else "Transactions"
    g = g.sort_values(sort_key, ascending=False)
    return g.reset_index().rename(columns={category_col: "Business Category"})


ITEM_VENDORS_COL = "Vendors (examples)"
ITEM_DEPTS_COL = "Departments (examples)"


def _normalize_item(s: pd.Series) -> pd.Series:
    """Collapse a description into a comparable 'item' key: uppercase, trimmed,
    internal whitespace collapsed. Deliberately conservative — it does not stem
    or fuzzy-match, so two rows only group when the text truly matches."""
    x = s.fillna("").astype(str).str.upper().str.strip()
    return x.str.replace(r"\s+", " ", regex=True)


def item_consolidation(df, desc_col, amount_col=None, vendor_col=None, dept_col=None,
                       category_col="Business_Category", min_group=2, top_n=50,
                       max_examples=4):
    """Item-level fragmentation: the SAME item (by description) bought from more
    than one vendor and/or across more than one department — the strongest
    consolidation signal. Runs on EVERY item, whatever its category (the grouping
    is by item text, so it does not depend on classification). Needs a description
    plus at least a vendor or a department column to be meaningful."""
    if not desc_col or desc_col not in df:
        return None
    has_ven = bool(vendor_col and vendor_col in df)
    has_dep = bool(dept_col and dept_col in df)
    if not (has_ven or has_dep):
        return None
    work = df.copy()
    work["_item"] = _normalize_item(work[desc_col])
    work = work[work["_item"] != ""]
    if not len(work):
        return None
    has_amt = bool(amount_col and amount_col in work)
    work["_amt"] = clean_amount(work[amount_col]).fillna(0.0) if has_amt else 1.0

    parts = {"Transactions": ("_item", "size")}
    if has_amt:
        parts["Spend"] = ("_amt", "sum")
    g = work.groupby("_item").agg(**parts)
    if has_ven:
        g["Vendors"] = work.groupby("_item")[vendor_col].nunique()
        g[ITEM_VENDORS_COL] = work.groupby("_item")[vendor_col].agg(
            lambda s: "  ·  ".join(pd.Series(s.astype(str).str.strip()).replace("", pd.NA)
                                   .dropna().unique()[:max_examples]))
    if has_dep:
        g["Departments"] = work.groupby("_item")[dept_col].nunique()
        g[ITEM_DEPTS_COL] = work.groupby("_item")[dept_col].agg(
            lambda s: "  ·  ".join(pd.Series(s.astype(str).str.strip()).replace("", pd.NA)
                                   .dropna().unique()[:max_examples]))
    if category_col in work:
        g["Business Category"] = work.groupby("_item")[category_col].agg(
            lambda s: s.mode().iat[0] if len(s.mode()) else "")

    # Keep only genuinely fragmented items: >1 vendor OR >1 department.
    cond = pd.Series(False, index=g.index)
    if "Vendors" in g:
        cond = cond | (g["Vendors"] >= min_group)
    if "Departments" in g:
        cond = cond | (g["Departments"] >= min_group)
    g = g[cond]
    if not len(g):
        return None
    sort_key = "Spend" if has_amt else "Transactions"
    g = g.sort_values(sort_key, ascending=False).head(top_n).reset_index()
    g = g.rename(columns={"_item": "Item"})
    order = [c for c in ["Item", "Business Category", "Spend", "Transactions",
                         "Vendors", "Departments", ITEM_VENDORS_COL, ITEM_DEPTS_COL]
             if c in g.columns]
    return g[order]


DESC_VARIANTS_COL = "Descriptions (variants)"


def nigp_item_consolidation(df, desc_col, amount_col=None, vendor_col=None, dept_col=None,
                            item_col="NIGP_Item_5digit", class_col="NIGP_Class_3digit",
                            min_group=2, top_n=50, max_examples=4, dept_cap=15):
    """Like item_consolidation, but groups by the NIGP commodity CODE instead of the
    raw description — so two differently-worded descriptions of the same commodity
    roll together. Uses the most specific code present per row (5-digit Item if the
    rule assigned one, otherwise 3-digit Class). Rows with no NIGP code are omitted
    from this view (nothing to group them by). Needs a vendor or department column."""
    if not desc_col or desc_col not in df:
        return None
    has_ven = bool(vendor_col and vendor_col in df)
    has_dep = bool(dept_col and dept_col in df)
    if not (has_ven or has_dep):
        return None
    if item_col not in df and class_col not in df:
        return None
    work = df.copy()
    item5 = work[item_col].astype(str).str.strip() if item_col in work else pd.Series("", index=work.index)
    class3 = work[class_col].astype(str).str.strip() if class_col in work else pd.Series("", index=work.index)
    item5 = item5.replace({"nan": "", "None": ""})
    class3 = class3.replace({"nan": "", "None": ""})
    # most-specific code per row; prefix marks the level so codes don't collide
    key = pd.Series("", index=work.index)
    key = key.mask(item5 != "", "I:" + item5)
    key = key.mask((key == "") & (class3 != ""), "C:" + class3)
    work["_key"] = key
    work = work[work["_key"] != ""]
    if not len(work):
        return None
    has_amt = bool(amount_col and amount_col in work)
    work["_amt"] = clean_amount(work[amount_col]).fillna(0.0) if has_amt else 1.0
    work["_desc"] = _normalize_item(work[desc_col])

    parts = {"Transactions": ("_key", "size")}
    if has_amt:
        parts["Spend"] = ("_amt", "sum")
    g = work.groupby("_key").agg(**parts)
    g["NIGP Code"] = work.groupby("_key")["_key"].first().str[2:]
    g["NIGP Level"] = work.groupby("_key")["_key"].first().str[0].map(
        {"I": "5-digit item", "C": "3-digit class"})
    g[DESC_VARIANTS_COL] = work.groupby("_key")["_desc"].agg(
        lambda s: "  ·  ".join(pd.Series(s).replace("", pd.NA).dropna().unique()[:max_examples]))
    if has_ven:
        g["Vendors"] = work.groupby("_key")[vendor_col].nunique()
        g[ITEM_VENDORS_COL] = work.groupby("_key")[vendor_col].agg(
            lambda s: "  ·  ".join(pd.Series(s.astype(str).str.strip()).replace("", pd.NA)
                                   .dropna().unique()[:max_examples]))
    measure_col = "Total Spend" if has_amt else "Total Transactions"
    g = g.rename(columns={"Spend": "Total Spend", "Transactions": "Total Transactions"})
    if "Business_Category" in work:
        g["Business Category"] = work.groupby("_key")["Business_Category"].agg(
            lambda s: s.mode().iat[0] if len(s.mode()) else "")

    # Department pivot: one column per department (spend, or transactions if no amount).
    dept_cols = []
    if has_dep:
        dwork = work.copy()
        dwork["_dept"] = dwork[dept_col].fillna("").astype(str).str.strip().replace("", "(blank)")
        g["Departments (#)"] = dwork.groupby("_key")["_dept"].nunique()
        totals = (dwork.groupby("_dept")["_amt"].sum() if has_amt
                  else dwork.groupby("_dept").size())
        top_depts = list(totals.sort_values(ascending=False).head(dept_cap).index)
        dwork["_deptcol"] = dwork["_dept"].where(dwork["_dept"].isin(top_depts), "Other Depts")
        pv = pd.pivot_table(dwork, index="_key", columns="_deptcol",
                            values="_amt", aggfunc="sum", fill_value=0.0)
        dept_cols = [d for d in top_depts if d in pv.columns]
        if "Other Depts" in pv.columns:
            dept_cols.append("Other Depts")
        g = g.join(pv[dept_cols])

    # Keep only genuinely fragmented commodities: >1 vendor OR bought across >1 dept.
    cond = pd.Series(False, index=g.index)
    if "Vendors" in g:
        cond = cond | (g["Vendors"] >= min_group)
    if "Departments (#)" in g:
        cond = cond | (g["Departments (#)"] >= min_group)
    g = g[cond]
    if not len(g):
        return None
    g = g.sort_values(measure_col, ascending=False).head(top_n).reset_index(drop=True)
    lead = [c for c in ["NIGP Code", "NIGP Level", "Business Category", DESC_VARIANTS_COL,
                        "Vendors", "Departments (#)", measure_col] if c in g.columns]
    tail = [c for c in [ITEM_VENDORS_COL] if c in g.columns]
    return g[lead + dept_cols + tail]


# Fixed (non-department) columns of the NIGP pivot — everything else is a department column.
NIGP_META_COLS = ["NIGP Code", "NIGP Level", "Business Category", DESC_VARIANTS_COL,
                  "Vendors", "Departments (#)", "Total Spend", "Total Transactions",
                  ITEM_VENDORS_COL, ITEM_DEPTS_COL]


def nigp_dept_columns(df) -> list:
    """The department columns of a nigp_item_consolidation table (all non-meta columns)."""
    return [c for c in df.columns if c not in NIGP_META_COLS]


# ---------------------------------------------------------------------------
# Best-fit absorption — assign every catch-all row to a real Business Category
# ---------------------------------------------------------------------------
BESTFIT_DEFAULT = "Professional & Administrative Services"
# Ordered specific → general; first hit wins. Broad "services" fall through to default.
BESTFIT_HINTS = [
    ("Vehicles & Fleet", ["vehicle", "truck", "auto ", "autos", "fleet", "tire", "fuel", "towing",
                           "windshield", "forklift", "snow plow", "dump truck", "car wash"]),
    ("Janitorial, Sanitation & Waste", ["janitor", "cleaning", "sanitation", "refuse", "recycl",
                                        "garbage", "trash", "waste", "disposal", "sweeping", "floor mat"]),
    ("Chemicals & Water Treatment", ["chemical", "water treatment", "chlorine", "sulfate", "phosphate",
                                     "fluor", "corrosion", "rock salt"]),
    ("Medical & Health Services", ["medical", "health", "clinical", "laborator", "nursing", "pharma",
                                   "dental", "oxygen", "vaccine", "covid", "hospital"]),
    ("IT, Telecom & Audio/Visual", ["software", "hardware", "computer", "server", "network", "license",
                                    "cloud", "telecom", "audio", "video", "camera", "scada", "gis"]),
    ("Construction & Trades Services", ["construction", "concrete", "asphalt", "paving", "sewer",
                                        "water main", "excavat", "demolition", "guardrail", "restoration",
                                        "electrical work", "carpentry", "masonry", "roofing"]),
    ("Construction Materials", ["pipe", "steel", "lumber", "aggregate", "gravel", "rebar",
                                "street light pole", "traffic signal", "sign post", "fitting"]),
    ("Landscaping, Grounds & Irrigation", ["landscap", "tree", "stump", "lawn", "irrigation", "grounds",
                                           "mulch", "herbicide", "fence", "fencing"]),
    ("Public Safety, Uniforms & PPE", ["uniform", "police", "fire department", "ppe", "protective",
                                       "ammunition", "body armor", "badge"]),
    ("Furniture & Furnishings", ["furniture", "chair", "desk", "cabinet", "furnishing", "cubicle"]),
    ("Heavy Equipment & Machinery", ["generator", "compressor", "loader", "excavator", "crane",
                                     "machinery", "heavy equipment"]),
    ("Equipment Rental & Leasing", ["rental", "lease", " rent "]),
    ("Animal Care & Veterinary", ["animal", "veterinar", "canine", "k-9", "livestock"]),
    ("Office, Print & Marketing", ["office supply", "office supplies", "paper", "printing", "print ",
                                   "marketing", "advertis", "stationery", "book", "publication", "media"]),
    ("Facilities Operations & Maintenance", ["repair", "maintenance", "hvac", "plumb", "electric",
                                             "facility", "building", "paint", "roof", "door", "elevator",
                                             "boiler", "janitorial supplies", "graffiti"]),
    # Professional & Administrative is the default; explicit hints kept minimal on purpose.
    ("Professional & Administrative Services", ["consult", "training", "legal", "audit", "staffing",
                                                "engineering", "architect", "design", "study", "inspection",
                                                "security guard", "translation", "records", "management"]),
]


def assign_best_fit(desc_series: pd.Series) -> pd.Series:
    """Assign each description to its closest of the 17 Business Categories using the
    strongest keyword signal available; anything unmatched → BESTFIT_DEFAULT."""
    u = desc_series.fillna("").astype(str).str.upper()
    out = pd.Series(BESTFIT_DEFAULT, index=desc_series.index)
    assigned = pd.Series(False, index=desc_series.index)
    for cat, hints in BESTFIT_HINTS:
        pat = "|".join(re.escape(h.upper()) for h in hints)
        m = (~assigned) & u.str.contains(pat, regex=True, na=False)
        out[m] = cat
        assigned |= m
    return out


def absorb_catchall(df, desc_col):
    """Fold every 'General & Other Procurement' row into a real Business Category via
    best-fit assignment, so no visible catch-all remains. Marks Classification_Method
    'best_fit' for transparency/audit."""
    if "Business_Category" not in df:
        return df
    mask = (df["Business_Category"] == CATCHALL)
    if not mask.any():
        return df
    df = df.copy()
    src = df[desc_col] if (desc_col and desc_col in df) else pd.Series("", index=df.index)
    df.loc[mask, "Business_Category"] = assign_best_fit(src[mask]).values
    if "Classification_Method" in df.columns:
        df.loc[mask, "Classification_Method"] = "best_fit"
    return df


# ---------------------------------------------------------------------------
# Savings / cost-avoidance engine (addressable-spend methodology, benchmark rates)
# ---------------------------------------------------------------------------
# Transparent, adjustable planning rates. Estimated figures are COST AVOIDANCE
# (a negotiated/consolidation opportunity), not realized hard budget savings.
DEFAULT_SAVINGS_RATES = {
    "per_vendor": 0.03,     # +3% of addressable spend per additional vendor on a commodity
    "per_dept": 0.02,       # +2% of addressable spend per additional department buying it
    "cap": 0.15,            # maximum consolidation rate applied to any one opportunity
    "tail": 0.05,           # savings rate applied to tail-spend consolidation
    "hard_fraction": 0.40,  # share of identified savings that is HARD (cashable, budget-reducing);
                            # the remainder is COST AVOIDANCE (soft / non-cashable)
}


# Plain-English column labels for the opportunities table (leadership-friendly).
# The technical terms (addressable spend, cost avoidance, spend under management) live
# in the "What these terms mean" note and the Sources & Methodology tab.
COL_ITEM = "Item / service"
COL_CATEGORY = "Category"
COL_MATCHED = "Matched by"
COL_ISSUE = "Why it's an opportunity"
COL_VENDORS = "Vendors"
COL_DEPTS = "Departments"
COL_SPEND = "Spend we can combine"
COL_RATE = "Savings rate %"
COL_TOTAL = "Estimated savings"
COL_HARD = "Cash savings"
COL_AVOID = "Avoided costs"
COL_ACTION = "What to do"


def _commodity_group(df, desc_col, item_col="NIGP_Item_5digit", class_col="NIGP_Class_3digit"):
    """Return (key, level, normalized_desc): group by NIGP 5-digit item, else 3-digit
    class, else the normalized description — so the same commodity groups together."""
    item5 = (df[item_col].astype(str).str.strip().replace({"nan": "", "None": ""})
             if item_col in df else pd.Series("", index=df.index))
    class3 = (df[class_col].astype(str).str.strip().replace({"nan": "", "None": ""})
              if class_col in df else pd.Series("", index=df.index))
    desc = _normalize_item(df[desc_col]) if (desc_col and desc_col in df) else pd.Series("", index=df.index)
    key = pd.Series("", index=df.index)
    level = pd.Series("", index=df.index)
    key = key.mask(item5 != "", "I:" + item5)
    level = level.mask(item5 != "", "NIGP 5-digit item")
    key = key.mask((key == "") & (class3 != ""), "C:" + class3)
    level = level.mask((level == "") & (class3 != ""), "NIGP 3-digit class")
    key = key.mask(key == "", "D:" + desc)
    level = level.mask(level == "", "Description")
    return key, level, desc


def consolidation_opportunities(df, desc_col, amount_col, vendor_col=None, dept_col=None,
                                rates=None, top_n=40):
    """One row per commodity that is fragmented across vendors and/or departments,
    on an ADDRESSABLE-SPEND basis, with an estimated COST AVOIDANCE and a plain-English
    recommended action. Savings = addressable spend × benchmark rate (rate rewards
    fragmentation). Needs an amount column plus a vendor or department column."""
    if not amount_col or amount_col not in df:
        return None
    has_ven = bool(vendor_col and vendor_col in df)
    has_dep = bool(dept_col and dept_col in df)
    if not (has_ven or has_dep):
        return None
    r = {**DEFAULT_SAVINGS_RATES, **(rates or {})}
    work = df.copy()
    key, level, desc = _commodity_group(work, desc_col)
    work["_key"], work["_level"], work["_desc"] = key, level, desc
    work = work[work["_desc"] != ""]
    if not len(work):
        return None
    work["_amt"] = clean_amount(work[amount_col]).fillna(0.0)

    g = work.groupby("_key").agg(**{COL_SPEND: ("_amt", "sum"),
                                    "Transactions": ("_amt", "size")})
    g[COL_ITEM] = work.groupby("_key")["_desc"].agg(
        lambda s: s.mode().iat[0] if len(s.mode()) else (s.iloc[0] if len(s) else ""))
    g[COL_MATCHED] = work.groupby("_key")["_level"].first()
    if has_ven:
        g[COL_VENDORS] = work.groupby("_key")[vendor_col].nunique()
    if has_dep:
        g[COL_DEPTS] = work.groupby("_key")[dept_col].nunique()
    if "Business_Category" in work:
        g[COL_CATEGORY] = work.groupby("_key")["Business_Category"].agg(
            lambda s: s.mode().iat[0] if len(s.mode()) else "")

    frag = pd.Series(False, index=g.index)
    if COL_VENDORS in g:
        frag = frag | (g[COL_VENDORS] >= 2)
    if COL_DEPTS in g:
        frag = frag | (g[COL_DEPTS] >= 2)
    g = g[frag]
    if not len(g):
        return None

    vv = (g[COL_VENDORS] - 1).clip(lower=0) if COL_VENDORS in g else 0
    dd = (g[COL_DEPTS] - 1).clip(lower=0) if COL_DEPTS in g else 0
    rate = (r["per_vendor"] * vv + r["per_dept"] * dd).clip(upper=r["cap"])
    hf = r["hard_fraction"]
    total_sav = (g[COL_SPEND] * rate).round(0)
    g[COL_RATE] = (rate * 100).round(1)
    g[COL_TOTAL] = total_sav
    g[COL_HARD] = (total_sav * hf).round(0)
    g[COL_AVOID] = (total_sav * (1 - hf)).round(0)

    def _opp(row):
        mv = (COL_VENDORS in g) and row.get(COL_VENDORS, 0) >= 2
        md = (COL_DEPTS in g) and row.get(COL_DEPTS, 0) >= 2
        return ("Many vendors & departments" if (mv and md)
                else "Many vendors" if mv else "Many departments" if md else "")

    def _action(row):
        bits = []
        if (COL_VENDORS in g) and row.get(COL_VENDORS, 0) >= 2:
            bits.append(f"{int(row[COL_VENDORS])} vendors")
        if (COL_DEPTS in g) and row.get(COL_DEPTS, 0) >= 2:
            bits.append(f"{int(row[COL_DEPTS])} departments")
        frag_txt = " across ".join(bits) if bits else "multiple sources"
        return (f"Combine “{str(row[COL_ITEM])[:48]}” (bought from {frag_txt}) into one contract "
                "so the city buys it at a single, better price.")

    g[COL_ISSUE] = g.apply(_opp, axis=1)
    g[COL_ACTION] = g.apply(_action, axis=1)
    g = g.sort_values(COL_TOTAL, ascending=False).head(top_n).reset_index(drop=True)
    order = [c for c in [COL_ITEM, COL_CATEGORY, COL_MATCHED, COL_ISSUE, COL_VENDORS, COL_DEPTS,
                         COL_SPEND, COL_RATE, COL_TOTAL, COL_HARD, COL_AVOID, COL_ACTION]
             if c in g.columns]
    return g[order]


def savings_summary(opps, tiles, tail=None, rates=None, years=1):
    """Portfolio roll-up of the cost-avoidance opportunity (addressable-spend basis)."""
    r = {**DEFAULT_SAVINGS_RATES, **(rates or {})}
    hf = r["hard_fraction"]
    total_spend = float(tiles.get("total_spend") or 0.0)
    has_opps = opps is not None and len(opps)
    addressable = float(opps[COL_SPEND].sum()) if has_opps else 0.0
    identified = float(opps[COL_TOTAL].sum()) if has_opps else 0.0
    hard = float(opps[COL_HARD].sum()) if has_opps else 0.0
    avoidance = float(opps[COL_AVOID].sum()) if has_opps else 0.0
    by_type = (opps.groupby(COL_ISSUE)[COL_TOTAL].sum().round(0).to_dict()
               if has_opps else {})
    tail_addr = float(tail["tail_value"]) if tail else 0.0
    tail_total = round(tail_addr * r["tail"]) if tail else 0.0
    tail_hard = round(tail_total * hf)
    tail_avoid = tail_total - tail_hard
    yrs = max(1, int(years or 1))
    return {
        "total_spend": total_spend,
        "addressable": addressable,
        "addressable_pct": round(addressable / total_spend * 100, 1) if total_spend else 0.0,
        "identified": identified,           # total identified savings opportunity
        "hard": hard,                       # cashable / budget-reducing portion
        "avoidance": avoidance,             # cost-avoidance (soft) portion
        "identified_pct_addr": round(identified / addressable * 100, 1) if addressable else 0.0,
        "identified_pct_total": round(identified / total_spend * 100, 1) if total_spend else 0.0,
        "hard_fraction": hf,
        "by_type": by_type,
        "tail_addressable": tail_addr,
        "tail_total": tail_total,
        "tail_hard": tail_hard,
        "tail_avoidance": tail_avoid,
        "n_opportunities": (len(opps) if has_opps else 0),
        "years": yrs,
        "annual_identified": identified / yrs,
        "annual_hard": hard / yrs,
        "annual_avoidance": avoidance / yrs,
        "three_year": (identified / yrs) * 3,
        "rates": r,
    }


# ---------------------------------------------------------------------------
# Sources & Methodology content (single source of truth for Excel + the app)
# ---------------------------------------------------------------------------
def savings_rate_card(rates=None):
    r = {**DEFAULT_SAVINGS_RATES, **(rates or {})}
    hp = int(round(r["hard_fraction"] * 100))
    return [
        ("Each extra vendor buying the same item adds", f"+{int(r['per_vendor'] * 100)}% of that spend"),
        ("Each extra department buying it adds", f"+{int(r['per_dept'] * 100)}% of that spend"),
        ("Most we ever count for one item (cap)", f"{int(r['cap'] * 100)}%"),
        ("Combining many small vendors", f"{int(r['tail'] * 100)}% of that spend"),
        ("Share counted as cash savings", f"{hp}% (the other {100 - hp}% is avoided costs)"),
    ]


SOURCES_SECTIONS = [
    ("In plain words", [
        "This report finds where the city buys the same thing from more than one vendor or in more "
        "than one department, and estimates what combining that buying could save. The amount that "
        "could be combined is measured directly from your data; the savings percentage is an "
        "adjustable estimate you can change on the page — not a number pulled from your invoices.",
        "“Cash savings” = money that comes back to the budget. “Avoided costs” = money you keep from "
        "spending later. We show both, because finance treats them differently.",
    ]),
    ("What the technical terms mean (for the detail-minded)", [
        "Addressable spend — the part of total spend that procurement can influence, source, or "
        "negotiate. It means spend you CAN act on; it does NOT mean spend that is already optimized.",
        "Spend under management — the part of addressable spend that is already actively sourced or "
        "under contract (i.e., already being managed). This is the “already-optimized” idea.",
        "Non-addressable spend — costs procurement cannot control (payroll, taxes, regulated fees, "
        "grants/pass-through). These are excluded from savings targets.",
        "In this tool, to stay simple and accurate, we don’t assume a percentage for addressable "
        "spend. We measure the specific slice that can be combined (bought from more than one vendor "
        "or department) straight from your data, and call it “spend we can combine.”",
    ]),
    ("The methods used here, and who uses them", [
        "Spend analysis / the “spend cube” — grouping spend by commodity, supplier, and "
        "department to find opportunity. Foundational category-management practice taught by ISM "
        "and CIPS and used by every major sourcing consultancy.",
        "Pareto (80/20) & ABC analysis — the few categories and vendors that drive most spend. A "
        "classic tool in both public and private procurement.",
        "Demand aggregation / consolidation — combining fragmented buys into one competitively-bid "
        "agreement. This is the core mechanism of public cooperative purchasing (NASPO ValuePoint, "
        "NIGP, GSA, Sourcewell, OMNIA Partners, E&I) and of corporate strategic sourcing.",
        "Tail-spend management — consolidating the many small suppliers that make up the last ~20% "
        "of spend.",
        "Vendor rationalization — reducing overlapping suppliers within a commodity to concentrate "
        "volume and improve leverage.",
        "Supplier concentration (HHI) — the Herfindahl-Hirschman Index, from antitrust economics "
        "(US DOJ / FTC), applied here to measure reliance on a few suppliers.",
    ]),
    ("How we size the opportunity", [
        "Savings are applied ONLY to the spend that can be combined — items bought from more than "
        "one vendor and/or department — never to total spend. Anything bought from a single vendor "
        "in a single department is left out.",
    ]),
    ("Cash savings vs avoided costs", [
        "Cash savings reduce the budget against a baseline — real, measurable dollars (the "
        "procurement term is “hard savings”). Avoided costs prevent or defer future cost — "
        "negotiated reductions vs market, avoided increases, less admin work (the term is “cost "
        "avoidance”). Guidance is to report the two separately so finance sees what was reduced vs "
        "prevented.",
        "This report splits each opportunity into a cash portion and an avoided-cost portion "
        "(default 40% / 60%, adjustable). The split is a planning convention, not a realized result.",
    ]),
    ("Where the benchmark rates come from", [
        "The rates are conservative, commonly-cited strategic-sourcing / consolidation ranges. "
        "Public-sector cooperative-purchasing bodies (e.g., NIGP) cite savings of up to ~15% "
        "annually from aggregation; some analyses cite more. This tool caps any single opportunity "
        "at 15% and scales the rate with fragmentation, so estimates stay deliberately conservative.",
        "These are PLANNING ESTIMATES, not guarantees. Actual results depend on execution, market "
        "conditions, contract terms, and your true baseline. Calibrate the rates to your agency’s "
        "realized savings over time.",
    ]),
    ("Public and private — both do this", [
        "Public sector: cooperative purchasing and category management are built on exactly this "
        "demand-aggregation math (NASPO, NIGP, GSA, regional co-ops).",
        "Private sector: strategic sourcing and category management by corporate procurement teams "
        "and advisory firms use the same spend-analysis → consolidation → savings approach.",
    ]),
    ("Honest limits", [
        "Your file has total amounts, not unit price × quantity, so this is not an exact "
        "overpayment calculation — it is an estimate based on how spread-out the buying is and the "
        "adjustable rates. Adding unit price and quantity would allow exact price comparisons.",
        "Estimates are only as good as the categorization and the rates. Review the opportunities "
        "and set the rates before presenting any firm dollar commitment.",
    ]),
]

SOURCES_REFERENCES = [
    ("NIGP — Cooperative Purchasing Programs",
     "https://www.nigp.org/our-profession/cooperative-purchasing-programs",
     "Public-procurement standards body; demand aggregation and spend analysis via NIGP codes."),
    ("NASPO ValuePoint — Why Use Cooperative Purchasing",
     "https://naspovaluepoint.org/cooperative-contracts/",
     "“Cooperative procurement is a form of strategic sourcing that aggregates the spend of public bodies.”"),
    ("NASPO — An Introduction to Cooperative Purchasing (PDF)",
     "https://naspovaluepoint.org/wp-content/uploads/2020/08/Cooperative_Purchasing0410update.pdf",
     "Primer on demand aggregation and where savings come from."),
    ("SpendHQ — Tracking Cost Savings and Cost Avoidance",
     "https://www.spendhq.com/blog/tracking-cost-savings-and-cost-avoidance-to-measure-procurements-performance/",
     "Definitions; report hard savings and cost avoidance separately."),
    ("ProcureAbility — Hard and Soft Cost Savings in Procurement",
     "https://procureability.com/hard-soft-savings-procurement-guide/",
     "Hard (cashable) vs soft (cost-avoidance) savings, defined."),
    ("Institute for Supply Management (ISM)", "https://www.ismworld.org/",
     "Professional body; category management and spend-analysis standards."),
    ("Chartered Institute of Procurement & Supply (CIPS)", "https://www.cips.org/",
     "Global procurement body; Kraljic, category management, sourcing methodology."),
    ("GFOA — Government Finance Officers Association", "https://www.gfoa.org/",
     "Best practices for public procurement and strategic sourcing."),
]


def coverage(df, category_col, method_col="Classification_Method"):
    total = len(df)
    catchall = int((df[category_col] == CATCHALL).sum())
    classified = total - catchall
    no_rule = no_desc = 0
    if method_col and method_col in df:
        no_rule = int((df[method_col] == METHOD_NO_RULE).sum())
        no_desc = int((df[method_col] == METHOD_NO_DESC).sum())
    return {
        "total": total,
        "classified": classified,
        "classified_pct": round(classified / total * 100, 1) if total else 0.0,
        "catchall": catchall,
        "catchall_pct": round(catchall / total * 100, 1) if total else 0.0,
        "no_rule": no_rule,
        "no_description": no_desc,
    }


# ---------------------------------------------------------------------------
# Adaptive column profiler + analysis planner (schema/format agnostic)
# ---------------------------------------------------------------------------
CONTRACT_HINTS = ["contract awarded", "contract", "awarded", "bpa", "blanket", "on contract"]
DIVERSITY_HINTS = ["mbe", "wbe", "dbe", "bepd", "diversity", "minority", "m/wbe", "certif"]
ID_HINTS = ["number", "req #", "req#", " id", "id ", "invoice", "voucher", "requisition number", "po #"]


def _num_frac(s: pd.Series) -> float:
    return float(clean_amount(s).notna().mean()) if len(s) else 0.0


def _date_frac(s: pd.Series) -> float:
    if pd.api.types.is_datetime64_any_dtype(s):
        return 1.0
    try:
        return float(pd.to_datetime(s, errors="coerce", format="mixed").notna().mean())
    except Exception:
        try:
            return float(pd.to_datetime(s, errors="coerce").notna().mean())
        except Exception:
            return 0.0


def profile_columns(df: pd.DataFrame) -> dict:
    """Inspect every column and assign a probable role using name + content
    heuristics. Roles: amount, date, description, vendor, department, id,
    dimension (low-cardinality categorical, subtype contract/diversity), other.
    Returns per-column detail plus the chosen primary columns for each role."""
    def hint(low, hints):
        return any(h in low for h in hints)

    cols = []
    for c in df.columns:
        s = df[c]
        nn = int(s.notna().sum())
        low = str(c).strip().lower()
        nun = int(s.nunique(dropna=True))
        card = nun / nn if nn else 0.0
        strvals = s.dropna().astype(str)
        avglen = float(strvals.str.len().mean()) if len(strvals) else 0.0
        has_dollar = float(strvals.str.contains(r"\$", regex=True).mean()) if len(strvals) else 0.0
        numf = _num_frac(s)
        datef = _date_frac(s) if (s.dtype == object or "date" in low or pd.api.types.is_datetime64_any_dtype(s)) else 0.0

        role, sub = "other", None
        if datef >= 0.7 or (hint(low, DATE_HINTS) and datef >= 0.4):
            role = "date"
        elif has_dollar >= 0.3 or (numf >= 0.8 and hint(low, AMOUNT_HINTS)):
            role = "amount"
        elif card >= 0.9 and avglen < 22 and (numf >= 0.6 or hint(low, ID_HINTS)):
            role = "id"
        elif hint(low, DESC_HINTS) and avglen >= 18 and card >= 0.2:
            role = "description"
        elif avglen >= 30 and card >= 0.3 and numf < 0.3:
            role = "description"
        elif hint(low, DEPT_HINTS):
            role = "department"
        elif hint(low, VENDOR_HINTS):
            role = "vendor"
        elif nun <= 40 and card < 0.5 and numf < 0.8 and datef < 0.4:
            role = "dimension"
            if hint(low, DIVERSITY_HINTS):
                sub = "diversity"
            elif hint(low, CONTRACT_HINTS):
                sub = "contract"
        elif numf < 0.3 and 0.02 < card < 0.95:
            role = "vendor"  # fallback: free-text categorical → likely a name
        cols.append(dict(name=str(c).strip(), low=low, role=role, subtype=sub,
                         nun=nun, card=round(card, 3), avglen=round(avglen, 1),
                         numf=round(numf, 2), datef=round(datef, 2)))

    def pick(role, hints=None, exclude=()):
        cands = [c for c in cols if c["role"] == role and c["name"] not in exclude]
        if hints:
            for h in hints:
                for c in cands:
                    if h in c["low"]:
                        return c["name"]
        return cands[0]["name"] if cands else None

    desc = pick("description", DESC_HINTS)
    if not desc:
        texts = sorted([c for c in cols if c["numf"] < 0.5 and c["datef"] < 0.5],
                       key=lambda c: -c["avglen"])
        desc = texts[0]["name"] if texts else None
    amount = pick("amount", AMOUNT_HINTS)
    vendor = pick("vendor", VENDOR_HINTS, exclude={desc})
    department = pick("department", DEPT_HINTS)
    date = pick("date", DATE_HINTS)
    dims = [c for c in cols if c["role"] == "dimension"]
    roles = {
        "description": desc, "amount": amount, "vendor": vendor,
        "department": department, "date": date,
        "dimensions": [c["name"] for c in dims],
        "dimension_subtypes": {c["name"]: c["subtype"] for c in dims},
    }
    return {"columns": cols, "roles": roles}


def plan_analyses(profile: dict) -> list:
    r = profile["roles"]
    has = lambda k: bool(r.get(k))
    amt = has("amount")
    plan = []

    def add(key, label, ok, why=""):
        plan.append({"key": key, "label": label, "feasible": bool(ok), "reason": why})

    add("category", "Spend by Business Category", True)
    add("pareto", "Pareto 80/20", True)
    add("top_vendors", "Top Vendors", has("vendor"), "" if has("vendor") else "no vendor column")
    add("consolidation", "Vendor Consolidation / Fragmentation", has("vendor"),
        "" if has("vendor") else "no vendor column")
    add("item_consolidation", "Same Item — Multiple Vendors / Departments",
        has("vendor") or has("department"),
        "" if (has("vendor") or has("department")) else "needs a vendor or department column")
    add("nigp_consolidation", "Same Commodity (NIGP code) — Multiple Vendors / Departments",
        has("vendor") or has("department"),
        "" if (has("vendor") or has("department")) else "needs a vendor or department column")
    add("department", "Spend by Department", has("department"),
        "" if has("department") else "no department column")
    add("trend", "Spend Trend Over Time", has("date"),
        "" if has("date") else "no date column")
    add("tail", "Tail-Spend Analysis", has("vendor"), "" if has("vendor") else "no vendor column")
    add("concentration", "Vendor Concentration / Risk", has("vendor"),
        "" if has("vendor") else "no vendor column")
    add("single_multi", "Single- vs Multi-Source Categories", has("vendor"),
        "" if has("vendor") else "no vendor column")
    add("matrix", "Category × Department Matrix", has("department"),
        "" if has("department") else "no department column")
    for dim in r.get("dimensions", []):
        sub = r.get("dimension_subtypes", {}).get(dim)
        label = {"contract": f"Contract vs Non-Contract ({dim})",
                 "diversity": f"Supplier Diversity ({dim})"}.get(sub, f"Spend by {dim}")
        add(f"dim::{dim}", label, True)
    return plan


# ---------------------------------------------------------------------------
# Extended industry-standard analyses
# ---------------------------------------------------------------------------
def spend_trend(df, date_col, amount_col=None):
    if not date_col or date_col not in df:
        return None
    work = df.copy()
    work["_d"] = pd.to_datetime(work[date_col], errors="coerce", format="mixed")
    work = work[work["_d"].notna()]
    if not len(work):
        return None
    work["_amt"] = clean_amount(work[amount_col]).fillna(0.0) if amount_col else 1.0
    work["Year"] = work["_d"].dt.year.astype(int).astype(str)
    g = work.groupby("Year").agg(Spend=("_amt", "sum"), Transactions=("_amt", "size")).reset_index()
    g = g.sort_values("Year")
    if amount_col:
        g["YoY %"] = (g["Spend"].pct_change() * 100).round(1)
    return g


def tail_spend(df, vendor_col, amount_col=None):
    if not vendor_col or vendor_col not in df:
        return None
    g, measure = _rollup(df, vendor_col, amount_col)
    g["_cum"] = g["% of Total"].cumsum()
    total_v = len(g)
    top_v = int((g["_cum"] < 80).sum() + 1) if total_v else 0
    top_v = min(top_v, total_v)
    tail = g.iloc[top_v:]
    tail_val = float(tail[measure].sum())
    grand = float(g[measure].sum())
    return {"total_vendors": total_v, "vendors_to_80pct": top_v,
            "tail_vendors": total_v - top_v,
            "tail_pct_of_vendors": round((total_v - top_v) / total_v * 100, 1) if total_v else 0.0,
            "tail_value": tail_val,
            "tail_pct_of_value": round(tail_val / grand * 100, 1) if grand else 0.0,
            "measure": measure}


def vendor_concentration(df, vendor_col, amount_col=None):
    if not vendor_col or vendor_col not in df:
        return None
    g, measure = _rollup(df, vendor_col, amount_col)
    tot = float(g[measure].sum())
    share = lambda k: round(float(g.head(k)[measure].sum()) / tot * 100, 1) if tot else 0.0
    frac = g["% of Total"] / 100.0
    hhi = int(round(float((frac ** 2).sum()) * 10000))
    return {"total_vendors": len(g), "top1": share(1), "top5": share(5),
            "top10": share(10), "hhi": hhi, "measure": measure}


def single_vs_multi_source(df, category_col, vendor_col, amount_col=None):
    if not vendor_col or vendor_col not in df:
        return None
    work = df[_is_classified(df[category_col])].copy()
    if amount_col:
        work["_amt"] = clean_amount(work[amount_col]).fillna(0.0)
        grp = work.groupby(category_col).agg(Vendors=(vendor_col, "nunique"), Spend=("_amt", "sum"))
        col = "Spend"
    else:
        grp = work.groupby(category_col).agg(Vendors=(vendor_col, "nunique"), Spend=(vendor_col, "size"))
        col = "Spend"
    single = grp[grp["Vendors"] == 1]
    multi = grp[grp["Vendors"] > 1]
    return {"single_categories": int(len(single)), "multi_categories": int(len(multi)),
            "single_value": float(single[col].sum()), "multi_value": float(multi[col].sum()),
            "has_amount": bool(amount_col)}


def spend_by_dimension(df, dim_col, amount_col=None, top_n=20):
    if not dim_col or dim_col not in df:
        return None
    g, _ = _rollup(df, dim_col, amount_col)
    g = g.rename(columns={dim_col: str(dim_col)})
    return g.head(top_n)


def category_department_matrix(df, category_col, dept_col, amount_col=None, top_depts=8, top_cats=12):
    if not dept_col or dept_col not in df:
        return None
    work = df[_is_classified(df[category_col])].copy()
    work["_amt"] = clean_amount(work[amount_col]).fillna(0.0) if amount_col else 1.0
    top_d = work.groupby(dept_col)["_amt"].sum().sort_values(ascending=False).head(top_depts).index
    top_c = work.groupby(category_col)["_amt"].sum().sort_values(ascending=False).head(top_cats).index
    w = work[work[dept_col].isin(top_d) & work[category_col].isin(top_c)]
    pv = pd.pivot_table(w, index=category_col, columns=dept_col, values="_amt",
                        aggfunc="sum", fill_value=0.0)
    pv = pv.round(0).reset_index().rename(columns={category_col: "Business Category"})
    pv.columns = [str(c) for c in pv.columns]
    return pv


# ---------------------------------------------------------------------------
# Executive summary (deterministic, data-driven)
# ---------------------------------------------------------------------------
def _money_text(v):
    try:
        return f"${float(v):,.0f}"
    except (TypeError, ValueError):
        return "n/a"


def executive_summary(tiles, cat, n80, vend, cons, cov, dept_tbl=None,
                      trend=None, tail=None, concentration=None, item_consol=None,
                      savings=None, opps=None) -> dict:
    has_amt = tiles.get("has_amount")
    unit = "spend" if has_amt else "transactions"
    real = cat
    lines, steps = [], []

    # Opening headline — lead with the money: the cost-avoidance opportunity.
    headline_val = _money_text(tiles.get("total_spend")) if has_amt else f"{tiles['transactions']:,} transactions"
    if savings and savings.get("identified", 0) > 0:
        lines.append(
            f"Bottom line: of the {headline_val} reviewed, about {_money_text(savings['addressable'])} "
            f"is spent buying the same things from more than one vendor or department. Combining that "
            f"buying could save roughly {_money_text(savings['identified'])} — about "
            f"{_money_text(savings['hard'])} back in the budget (cash savings) and "
            f"{_money_text(savings['avoidance'])} in avoided future costs — across "
            f"{savings['n_opportunities']} opportunities. That is about "
            f"{_money_text(savings['annual_identified'])} a year, or "
            f"{_money_text(savings['three_year'])} over three years. These are estimates; you can "
            "change the savings rates on the page and see them update.")
    else:
        lines.append(
            f"Bottom line: of the {headline_val} reviewed, spend concentrates in a handful of "
            "categories" + (" and a short list of vendors" if (vend is not None and len(vend)) else "")
            + ". The fastest wins are combining the same buys made across many vendors and departments.")

    parts = []
    if has_amt:
        parts.append(f"{_money_text(tiles.get('total_spend'))} in spend")
    parts.append(f"{tiles['transactions']:,} transactions")
    if tiles.get("vendors"):
        parts.append(f"{tiles['vendors']:,} vendors")
    if dept_tbl is not None:
        parts.append(f"{len(dept_tbl):,} departments")
    span = f", covering {tiles['date_min']} to {tiles['date_max']}" if "date_min" in tiles else ""
    lines.append("This report analyzes " + ", ".join(parts) + span + ".")

    bestfit_pct = cov.get("catchall_pct", 0.0)
    if bestfit_pct:
        lines.append(
            f"{cov['classified_pct']}% of rows matched a specific commodity rule directly; the "
            f"remaining {bestfit_pct}% were assigned to their closest Business Category (best-fit), "
            "so every row lands in one of the 17 categories — no “unclassified” gap.")

    if len(real):
        top = real.iloc[0]
        val = _money_text(top["Spend"]) if has_amt else f"{int(top['Transactions']):,} transactions"
        lines.append(
            f"Classified {unit} is concentrated: the top {n80} categor"
            f"{'y' if n80 == 1 else 'ies'} drive about 80%. The largest is "
            f"“{top['Business Category']}” at {val} ({top['% of Total']}%).")

    # Clear, upfront, money-ranked recommendations — the first thing leadership reads.
    if opps is not None and len(opps):
        for _, o in opps.head(3).iterrows():
            steps.append(
                f"{o[COL_ACTION]} → about {_money_text(o[COL_TOTAL])} in estimated savings "
                f"({_money_text(o[COL_HARD])} cash + {_money_text(o[COL_AVOID])} avoided) on "
                f"{_money_text(o[COL_SPEND])} of combinable spend.")
    if savings and savings.get("tail_total", 0) > 0:
        steps.append(
            f"Combine the many small vendors that add up to little spend → about "
            f"{_money_text(savings['tail_total'])} in estimated savings "
            f"({_money_text(savings['tail_hard'])} cash + {_money_text(savings['tail_avoidance'])} "
            f"avoided) on {_money_text(savings['tail_addressable'])} of that spread-out spending "
            f"({int(savings['rates']['tail'] * 100)}%).")

    if cons is not None and len(cons):
        frag = cons.head(3)
        names = ", ".join(f"{r['Business Category']} ({int(r['Vendors'])} vendors)"
                          for _, r in frag.iterrows())
        lines.append(
            f"The most fragmented categories — bought from many vendors — are {names}. "
            "These are the clearest consolidation opportunities.")

    if vend is not None and len(vend):
        tv = vend.iloc[0]
        val = _money_text(tv["Spend"]) if has_amt else f"{int(tv['Transactions']):,} transactions"
        lines.append(
            f"The largest single vendor is {tv['Vendor']} at {val} ({tv['% of Total']}% of total).")
        steps.append("Review the top vendors for volume-based pricing, contract compliance, and overlap.")

    if item_consol is not None and len(item_consol):
        it = item_consol.iloc[0]
        detail = []
        if "Vendors" in item_consol.columns:
            detail.append(f"{int(it['Vendors'])} vendors")
        if "Departments" in item_consol.columns:
            detail.append(f"{int(it['Departments'])} departments")
        val = _money_text(it["Spend"]) if ("Spend" in item_consol.columns and has_amt) \
            else f"{int(it['Transactions']):,} transactions"
        lines.append(
            f"Same-item fragmentation exists at the line level: e.g., “{str(it['Item'])[:60]}” was "
            f"bought across {', '.join(detail)} ({val}). {len(item_consol)} such items are listed — "
            "the same thing sourced from more than one vendor or department.")
        steps.append(
            "Review the Same-Item report: standardize and consolidate the top line items bought "
            "from multiple vendors or across departments onto a single source/agreement.")

    if trend is not None and "Spend" in trend.columns and len(trend) >= 2:
        peak = trend.loc[trend["Spend"].idxmax()]
        lines.append(
            f"Annual spend spans {trend.iloc[0]['Year']}–{trend.iloc[-1]['Year']}, peaking in "
            f"{peak['Year']} at {_money_text(peak['Spend'])}.")

    if concentration:
        lvl = ("high" if concentration["top10"] >= 60 else
               "moderate" if concentration["top10"] >= 35 else "low")
        lines.append(
            f"Supplier concentration is {lvl}: the top 10 vendors represent "
            f"{concentration['top10']}% of {unit} (HHI {concentration['hhi']:,}).")

    if tail:
        lines.append(
            f"A long tail of {tail['tail_vendors']:,} vendors ({tail['tail_pct_of_vendors']}% of all "
            f"vendors) accounts for only {tail['tail_pct_of_value']}% of {unit} — prime consolidation "
            "territory.")

    if dept_tbl is not None and len(dept_tbl) > 1:
        steps.append(
            "Coordinate purchasing across departments buying the same commodity to capture "
            "cross-department leverage.")

    # Explicit, ranked list of the top cost-avoidance opportunities (each with the math).
    consolidation_items = []
    if opps is not None and len(opps):
        has_v = COL_VENDORS in opps.columns
        has_d = COL_DEPTS in opps.columns
        for _, r in opps.head(10).iterrows():
            bits = []
            if has_v and int(r.get(COL_VENDORS, 0)) >= 2:
                bits.append(f"{int(r[COL_VENDORS])} vendors")
            if has_d and int(r.get(COL_DEPTS, 0)) >= 2:
                bits.append(f"{int(r[COL_DEPTS])} depts")
            consolidation_items.append({
                "item": str(r.get(COL_ITEM, ""))[:70],
                "category": str(r.get(COL_CATEGORY, "")),
                "detail": ", ".join(bits),
                "addressable": _money_text(r[COL_SPEND]),
                "rate": f"{r[COL_RATE]}%",
                "total": _money_text(r[COL_TOTAL]),
                "hard": _money_text(r[COL_HARD]),
                "avoidance": _money_text(r[COL_AVOID]),
            })

    return {"lines": lines, "steps": steps, "consolidation_items": consolidation_items}


# ---------------------------------------------------------------------------
# Professional charts (matplotlib → PNG bytes; used by both Excel and the app)
# ---------------------------------------------------------------------------
CHART_NAVY = "#002F6C"
CHART_BLUE = "#41B6E6"
CHART_RED = "#DA291C"
CHART_GREEN = "#2E7D32"
CHART_GRAY = "#6B7280"


def _plt():
    """Lazy matplotlib import (Agg backend) so the module loads even without it."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    return plt


def _money_fmt(v, _pos=None):
    v = float(v)
    a = abs(v)
    if a >= 1e9:
        return f"${v/1e9:.1f}B"
    if a >= 1e6:
        return f"${v/1e6:.1f}M"
    if a >= 1e3:
        return f"${v/1e3:.0f}K"
    return f"${v:,.0f}"


def _fig_png(fig):
    import io
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
    _plt().close(fig)
    return buf.getvalue()


def _short(labels, n=32):
    return [(str(x) if len(str(x)) <= n else str(x)[:n - 1] + "…") for x in labels]


def barh_png(labels, values, title, value_label="Spend ($)", money=True,
             color=CHART_NAVY, highlight_idx=None):
    """Horizontal bar chart, largest at top, with data labels and a readable axis."""
    plt = _plt()
    from matplotlib.ticker import FuncFormatter
    labels = _short(labels)
    n = len(labels)
    fig, ax = plt.subplots(figsize=(9, max(2.4, 0.46 * n + 1.1)))
    ypos = range(n)
    colors = [color] * n
    if highlight_idx is not None:
        for i in highlight_idx:
            if 0 <= i < n:
                colors[i] = CHART_RED
    ax.barh(list(ypos), list(values), color=colors, edgecolor="#1a1a1a", linewidth=0.4, zorder=3)
    ax.set_yticks(list(ypos))
    ax.set_yticklabels(labels, fontsize=9)
    ax.invert_yaxis()
    ax.set_xlabel(value_label, fontsize=10, color=CHART_NAVY)
    ax.set_title(title, fontsize=13, fontweight="bold", color=CHART_NAVY, pad=10, loc="left")
    if money:
        ax.xaxis.set_major_formatter(FuncFormatter(_money_fmt))
    ax.grid(axis="x", color="#D9D9D9", linewidth=0.7, zorder=0)
    for s in ("top", "right"):
        ax.spines[s].set_visible(False)
    for s in ("left", "bottom"):
        ax.spines[s].set_color("#9AA0A6")
    vmax = max(values) if len(values) and max(values) else 1
    for i, v in enumerate(values):
        ax.text(v + vmax * 0.01, i, _money_fmt(v) if money else f"{v:,.0f}",
                va="center", ha="left", fontsize=8.5, color="#222")
    ax.set_xlim(0, vmax * 1.16)
    fig.tight_layout()
    return _fig_png(fig)


def pareto_png(labels, values, cum_pct, title, value_label="Spend ($)"):
    plt = _plt()
    from matplotlib.ticker import FuncFormatter
    labels = _short(labels, 22)
    n = len(labels)
    fig, ax = plt.subplots(figsize=(max(7, 0.7 * n + 2), 4.6))
    x = range(n)
    ax.bar(list(x), list(values), color=CHART_NAVY, edgecolor="#1a1a1a", linewidth=0.4, zorder=3)
    ax.set_ylabel(value_label, fontsize=10, color=CHART_NAVY)
    ax.yaxis.set_major_formatter(FuncFormatter(_money_fmt))
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, rotation=40, ha="right", fontsize=8.5)
    ax.set_title(title, fontsize=13, fontweight="bold", color=CHART_NAVY, pad=10, loc="left")
    ax.grid(axis="y", color="#D9D9D9", linewidth=0.7, zorder=0)
    ax2 = ax.twinx()
    ax2.plot(list(x), list(cum_pct), color=CHART_RED, marker="o", markersize=4, linewidth=2, zorder=4)
    ax2.axhline(80, color=CHART_GRAY, linestyle="--", linewidth=1)
    ax2.set_ylabel("Cumulative %", fontsize=10, color=CHART_RED)
    ax2.set_ylim(0, 105)
    ax2.yaxis.set_major_formatter(FuncFormatter(lambda v, p: f"{v:.0f}%"))
    for s in ("top",):
        ax.spines[s].set_visible(False)
        ax2.spines[s].set_visible(False)
    fig.tight_layout()
    return _fig_png(fig)


def line_png(x_labels, values, title, value_label="Spend ($)"):
    plt = _plt()
    from matplotlib.ticker import FuncFormatter
    fig, ax = plt.subplots(figsize=(8, 4.2))
    x = range(len(x_labels))
    ax.plot(list(x), list(values), color=CHART_NAVY, marker="o", markersize=6, linewidth=2.4, zorder=3)
    ax.fill_between(list(x), list(values), color=CHART_BLUE, alpha=0.18, zorder=2)
    ax.set_xticks(list(x))
    ax.set_xticklabels([str(l) for l in x_labels], fontsize=9)
    ax.set_ylabel(value_label, fontsize=10, color=CHART_NAVY)
    ax.yaxis.set_major_formatter(FuncFormatter(_money_fmt))
    ax.set_title(title, fontsize=13, fontweight="bold", color=CHART_NAVY, pad=10, loc="left")
    ax.grid(axis="y", color="#D9D9D9", linewidth=0.7, zorder=0)
    for s in ("top", "right"):
        ax.spines[s].set_visible(False)
    for i, v in enumerate(values):
        ax.text(i, v, "  " + _money_fmt(v), va="bottom", ha="center", fontsize=8.5, color="#222")
    fig.tight_layout()
    return _fig_png(fig)


def savings_split_png(hard, avoidance, title="Estimated Savings — Cash vs Avoided Costs"):
    """Single stacked bar showing cash savings vs avoided-costs split."""
    plt = _plt()
    from matplotlib.ticker import FuncFormatter
    fig, ax = plt.subplots(figsize=(8, 2.2))
    ax.barh([0], [hard], color=CHART_GREEN, edgecolor="#1a1a1a", linewidth=0.4, label="Cash savings", zorder=3)
    ax.barh([0], [avoidance], left=[hard], color=CHART_BLUE, edgecolor="#1a1a1a", linewidth=0.4,
            label="Avoided costs", zorder=3)
    total = hard + avoidance
    if hard:
        ax.text(hard / 2, 0, f"Cash savings\n{_money_fmt(hard)}", va="center", ha="center", fontsize=9,
                color="white", fontweight="bold")
    if avoidance:
        ax.text(hard + avoidance / 2, 0, f"Avoided costs\n{_money_fmt(avoidance)}", va="center",
                ha="center", fontsize=9, color="#0b3d5c", fontweight="bold")
    ax.set_xlim(0, total * 1.02 if total else 1)
    ax.set_yticks([])
    ax.xaxis.set_major_formatter(FuncFormatter(_money_fmt))
    ax.set_title(f"{title}   (total {_money_fmt(total)})", fontsize=12.5, fontweight="bold",
                 color=CHART_NAVY, pad=10, loc="left")
    for s in ("top", "right", "left"):
        ax.spines[s].set_visible(False)
    ax.grid(axis="x", color="#E3E3E3", linewidth=0.6, zorder=0)
    fig.tight_layout()
    return _fig_png(fig)


def _embed_png(ws, png_bytes, anchor_cell):
    """Place a matplotlib PNG onto a worksheet at anchor_cell (e.g. 'A20')."""
    if not png_bytes:
        return
    import io
    from openpyxl.drawing.image import Image as XLImage
    ws.add_image(XLImage(io.BytesIO(png_bytes)), anchor_cell)


# ---------------------------------------------------------------------------
# Excel report (professional, brand-colored, charts)
# ---------------------------------------------------------------------------
def _style_data_sheet(ws, title, df, *, money_cols=(), pct_cols=(), int_cols=(),
                      total_cols=(), add_total=True, wrap_cols=()):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ncols = df.shape[1]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tcell = ws.cell(row=1, column=1, value=title)
    tcell.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    tcell.fill = PatternFill("solid", fgColor=NAVY)
    tcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24

    header_row, data_first = 2, 3
    data_last = header_row + len(df)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border
    has_wrap = False
    for r in range(data_first, data_last + 1):
        band = (r - data_first) % 2 == 1
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if band:
                cell.fill = PatternFill("solid", fgColor=LT_BLUE)
            colname = df.columns[c - 1]
            if colname in money_cols:
                cell.number_format = '$#,##0'
            elif colname in pct_cols:
                cell.number_format = '0.0"%"'
            elif colname in int_cols:
                cell.number_format = '#,##0'
            elif colname in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                has_wrap = True
        if has_wrap:
            ws.row_dimensions[r].height = 42

    if add_total and len(df):
        tr = data_last + 1
        for c in range(1, ncols + 1):
            cell = ws.cell(row=tr, column=c)
            cell.font = Font(bold=True, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=RED)
            cell.border = border
            colname = df.columns[c - 1]
            if c == 1:
                cell.value = "TOTAL"
            elif colname in total_cols:
                col_letter = cell.column_letter
                cell.value = f"=SUM({col_letter}{data_first}:{col_letter}{data_last})"
                if colname in money_cols:
                    cell.number_format = '$#,##0'
                elif colname in int_cols:
                    cell.number_format = '#,##0'
            elif colname in pct_cols:
                cell.value = 100.0
                cell.number_format = '0.0"%"'

    for c in range(1, ncols + 1):
        colname = df.columns[c - 1]
        if colname in wrap_cols:
            # wrapped free-text column (examples): fixed generous width
            ws.column_dimensions[ws.cell(row=header_row, column=c).column_letter].width = 60
            continue
        maxlen = max([len(str(df.columns[c - 1]))] +
                     [len(str(df.iloc[r, c - 1])) for r in range(len(df))], default=12)
        # allow wide first columns (department/vendor/category names) to show in full
        cap = 60 if c == 1 else 46
        ws.column_dimensions[ws.cell(row=header_row, column=c).column_letter].width = min(max(maxlen + 3, 12), cap)
    ws.freeze_panes = ws.cell(row=data_first, column=1)
    ws.sheet_view.showGridLines = False
    return header_row, data_first, data_last


def _bar_chart(ws, title, cat_col, val_col, header_row, data_first, data_last, anchor,
               x_title="", y_title=""):
    from openpyxl.chart import BarChart, Reference
    ch = BarChart()
    ch.type = "col"
    ch.title = title
    ch.height = 8.5
    ch.width = 17
    ch.style = 10
    data = Reference(ws, min_col=val_col, min_row=header_row, max_row=data_last)
    cats = Reference(ws, min_col=cat_col, min_row=data_first, max_row=data_last)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ch.legend = None
    ch.x_axis.title = x_title
    ch.y_axis.title = y_title
    ch.x_axis.delete = False   # ensure axis + its title actually render
    ch.y_axis.delete = False
    ch.gapWidth = 60
    ws.add_chart(ch, anchor)


def _pareto_chart(ws, cat_col, spend_col, cum_col, header_row, data_first, data_last, anchor,
                  x_title="Business Category", y_title="Spend ($)"):
    from openpyxl.chart import BarChart, LineChart, Reference
    bar = BarChart()
    bar.type = "col"
    bar.title = "Pareto — spend and cumulative %"
    bar.height = 9
    bar.width = 18
    bar.style = 10
    d = Reference(ws, min_col=spend_col, min_row=header_row, max_row=data_last)
    cats = Reference(ws, min_col=cat_col, min_row=data_first, max_row=data_last)
    bar.add_data(d, titles_from_data=True)
    bar.set_categories(cats)
    bar.legend = None
    bar.x_axis.title = x_title
    bar.y_axis.title = y_title
    bar.x_axis.delete = False
    bar.y_axis.delete = False
    line = LineChart()
    dl = Reference(ws, min_col=cum_col, min_row=header_row, max_row=data_last)
    line.add_data(dl, titles_from_data=True)
    line.y_axis.axId = 200
    line.y_axis.title = "Cumulative %"
    line.y_axis.crosses = "max"
    line.y_axis.delete = False
    bar += line
    ws.add_chart(bar, anchor)


def _write_exec_summary(ws, title, tiles, cov, es):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2.5
    for col in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 19

    ws.merge_cells("B2:G2")
    t = ws["B2"]
    t.value = title
    t.font = Font(size=18, bold=True, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[2].height = 34
    ws.merge_cells("B3:G3")
    sub = ws["B3"]
    subtxt = "Procurement Spend Analysis"
    if "date_min" in tiles:
        subtxt += f"   |   {tiles['date_min']} to {tiles['date_max']}"
    sub.value = subtxt
    sub.font = Font(size=11, italic=True, color=NAVY)

    metrics = []
    if tiles.get("has_amount"):
        metrics.append(("Total Spend", _money_text(tiles.get("total_spend"))))
    metrics.append(("Transactions", f"{tiles['transactions']:,}"))
    metrics.append(("Categories", f"{tiles['categories']:,}"))
    if tiles.get("vendors"):
        metrics.append(("Vendors", f"{tiles['vendors']:,}"))
    metrics.append(("Classified", f"{cov['classified_pct']}%"))
    col = 2
    for label, val in metrics:
        v = ws.cell(row=5, column=col, value=val)
        v.font = Font(size=15, bold=True, color=RED)
        v.alignment = Alignment(horizontal="center")
        v.fill = PatternFill("solid", fgColor=LT_BLUE)
        d = ws.cell(row=6, column=col, value=label)
        d.font = Font(size=10, color=NAVY)
        d.alignment = Alignment(horizontal="center")
        d.fill = PatternFill("solid", fgColor=LT_BLUE)
        col += 1

    row = 8
    h = ws.cell(row=row, column=2, value="What the data shows")
    h.font = Font(size=13, bold=True, color=NAVY)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    row += 1
    for line in (es["lines"] if es else []):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        c = ws.cell(row=row, column=2, value="•  " + line)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = Font(size=11)
        ws.row_dimensions[row].height = 32
        row += 1

    row += 1
    h2 = ws.cell(row=row, column=2, value="Recommended first steps")
    h2.font = Font(size=13, bold=True, color=NAVY)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    row += 1
    for i, step in enumerate(es["steps"] if es else [], 1):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        c = ws.cell(row=row, column=2, value=f"{i}.  {step}")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = Font(size=11)
        ws.row_dimensions[row].height = 32
        row += 1

    items = (es or {}).get("consolidation_items") or []
    if items:
        row += 1
        h3 = ws.cell(row=row, column=2,
                     value="Top consolidation opportunities (estimated hard savings + cost avoidance)")
        h3.font = Font(size=13, bold=True, color=NAVY)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        row += 1
        for it in items:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
            c = ws.cell(row=row, column=2)
            detail = f" ({it['detail']})" if it.get("detail") else ""
            tail = (f"  —  {it['addressable']} addressable @ {it['rate']} → {it.get('total', '')} "
                    f"(hard {it.get('hard', '')} + avoidance {it.get('avoidance', '')})")
            c.value = "•  " + it["item"] + detail + tail
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.font = Font(size=11)
            ws.row_dimensions[row].height = 30
            row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    fn = ws.cell(row=row, column=2, value=(
        "Note: every row lands in a real Business Category. “General & Other Procurement” is the "
        "catch-all for buys that don’t match a specific commodity rule yet — the dollars are still "
        "counted and the Examples column shows what’s in it. Adding rules for those descriptions "
        "moves the spend into specific categories over time. All figures are deterministic — no AI "
        "is used in the analysis."))
    fn.font = Font(size=9, italic=True, color=GRAY)
    fn.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 46


def _write_savings_sheet(ws, savings, has_amt):
    from openpyxl.styles import Font, PatternFill, Alignment
    GREEN, TEAL = "2E7D32", "0B6FA4"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2.5
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 19
    ws.merge_cells("B2:G2")
    t = ws["B2"]
    t.value = "Savings Opportunity Summary"
    t.font = Font(size=18, bold=True, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[2].height = 34
    ws.merge_cells("B3:G3")
    ws["B3"].value = ("How much the city could save by buying the same things together — "
                      "estimates you can adjust (see Sources & Methodology)")
    ws["B3"].font = Font(size=10, italic=True, color=NAVY)
    if not savings or not has_amt or not savings.get("identified"):
        ws["B5"].value = ("Savings needs an amount column plus a vendor or department column, and at "
                          "least one item bought from more than one vendor or department.")
        ws["B5"].font = Font(size=12, color=RED)
        return
    s = savings
    tiles_data = [
        ("Total spend reviewed", _money_text(s["total_spend"]), NAVY),
        ("Spend we can combine", f"{_money_text(s['addressable'])} ({s['addressable_pct']}%)", NAVY),
        ("Estimated savings", _money_text(s["identified"]), RED),
        ("Cash savings", _money_text(s["hard"]), GREEN),
        ("Avoided costs", _money_text(s["avoidance"]), TEAL),
        ("Each year", _money_text(s["annual_identified"]), NAVY),
    ]
    col = 2
    for label, val, color in tiles_data:
        v = ws.cell(row=5, column=col, value=val)
        v.font = Font(size=13, bold=True, color=color)
        v.alignment = Alignment(horizontal="center")
        v.fill = PatternFill("solid", fgColor=LT_BLUE)
        d = ws.cell(row=6, column=col, value=label)
        d.font = Font(size=9.5, color=NAVY)
        d.alignment = Alignment(horizontal="center")
        d.fill = PatternFill("solid", fgColor=LT_BLUE)
        col += 1
    ws.merge_cells("B8:G8")
    c = ws["B8"]
    c.value = (f"Over three years: {_money_text(s['three_year'])}   ·   "
               f"{s['n_opportunities']} things we could combine   ·   {s['years']} year(s) of data")
    c.font = Font(size=12, bold=True, color=RED)
    _embed_png(ws, savings_split_png(s["hard"], s["avoidance"]), "B10")

    row = 24
    ws.cell(row=row, column=2, value="Estimated savings by reason").font = Font(size=12, bold=True, color=NAVY)
    row += 1
    for j, htxt in ((2, "Reason it's an opportunity"), (3, "Estimated savings")):
        hc = ws.cell(row=row, column=j, value=htxt)
        hc.font = Font(bold=True, color=WHITE)
        hc.fill = PatternFill("solid", fgColor=NAVY)
    row += 1
    for k, v in sorted(s["by_type"].items(), key=lambda kv: -kv[1]):
        ws.cell(row=row, column=2, value=k)
        mc = ws.cell(row=row, column=3, value=v)
        mc.number_format = '$#,##0'
        row += 1
    if s.get("tail_total"):
        ws.cell(row=row, column=2, value="Combining many small vendors (extra)")
        mc = ws.cell(row=row, column=3, value=s["tail_total"])
        mc.number_format = '$#,##0'
        row += 1
    row += 1
    ws.cell(row=row, column=2, value="Savings rates used (you can change these)").font = Font(size=11, bold=True, color=NAVY)
    row += 1
    for label, val in savings_rate_card(s.get("rates")):
        ws.cell(row=row, column=2, value=label).font = Font(size=10)
        ws.cell(row=row, column=4, value=val).font = Font(size=10, bold=True, color=NAVY)
        row += 1
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    n = ws.cell(row=row, column=2, value=(
        "These are estimates to help decide where to look first — part is money back in the budget "
        "(cash savings) and part is cost we avoid later (avoided costs). They are based on the amount "
        "spent, not on unit prices. The exact terms and sources are on the Sources & Methodology tab."))
    n.font = Font(size=9, italic=True, color=GRAY)
    n.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 42


def _write_sources_sheet(ws, rates=None):
    from openpyxl.styles import Font, PatternFill, Alignment
    TEAL = "0B6FA4"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2.5
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 74
    ws.merge_cells("B2:C2")
    t = ws["B2"]
    t.value = "Sources & Methodology"
    t.font = Font(size=18, bold=True, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[2].height = 34
    row = 4
    for heading, lines in SOURCES_SECTIONS:
        h = ws.cell(row=row, column=2, value=heading)
        h.font = Font(size=12, bold=True, color=NAVY)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        row += 1
        for ln in lines:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            c = ws.cell(row=row, column=2, value="•  " + ln)
            c.font = Font(size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 14 * (1 + len(ln) // 95)
            row += 1
        row += 1
    ws.cell(row=row, column=2, value="Rates used in this report (adjustable in the app)").font = Font(size=12, bold=True, color=NAVY)
    row += 1
    for label, val in savings_rate_card(rates):
        ws.cell(row=row, column=2, value=label).font = Font(size=10)
        ws.cell(row=row, column=3, value=val).font = Font(size=10, bold=True, color=NAVY)
        row += 1
    row += 1
    ws.cell(row=row, column=2, value="Reference organizations & sources").font = Font(size=12, bold=True, color=NAVY)
    row += 1
    for name, url, note in SOURCES_REFERENCES:
        nc = ws.cell(row=row, column=2, value=name)
        nc.font = Font(size=10, bold=True, color=TEAL, underline="single")
        nc.hyperlink = url
        d = ws.cell(row=row, column=3, value=note)
        d.font = Font(size=9, color=GRAY)
        d.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 28
        row += 1
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    n = ws.cell(row=row, column=2, value=(
        "Note: the specific savings percentages are this tool’s transparent planning assumptions, "
        "calibrated to conservative, commonly-cited ranges — not figures attributed to any single "
        "proprietary report. Adjust them to your agency’s realized savings."))
    n.font = Font(size=9, italic=True, color=GRAY)
    n.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 44


def _safe_sheet_name(name, used):
    base = re.sub(r'[\[\]\:\*\?\/\\]', ' ', str(name))[:28].strip() or "Sheet"
    nm, i = base, 2
    while nm in used:
        nm = f"{base[:25]} {i}"
        i += 1
    used.add(nm)
    return nm


def _write_analytics_sheet(ws, tail, conc, sm):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2.5
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 24
    ws.merge_cells("B2:C2")
    t = ws["B2"]
    t.value = "Vendor Analytics"
    t.font = Font(size=14, bold=True, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[2].height = 22
    state = {"row": 4}

    def section(title, rows):
        r = state["row"]
        h = ws.cell(row=r, column=2, value=title)
        h.font = Font(size=12, bold=True, color=NAVY)
        r += 1
        for label, val in rows:
            ws.cell(row=r, column=2, value=label).font = Font(size=11)
            c = ws.cell(row=r, column=3, value=val)
            c.font = Font(size=11, bold=True)
            c.alignment = Alignment(horizontal="right")
            r += 1
        state["row"] = r + 1

    if conc:
        section("Vendor concentration", [
            ("Total vendors", f"{conc['total_vendors']:,}"),
            ("Top 1 vendor share", f"{conc['top1']}%"),
            ("Top 5 vendor share", f"{conc['top5']}%"),
            ("Top 10 vendor share", f"{conc['top10']}%"),
            ("HHI (0–10,000; higher = more concentrated)", f"{conc['hhi']:,}")])
    if tail:
        m = tail["measure"].lower()
        fmt = _money_text if tail["measure"] == "Spend" else (lambda v: f"{int(v):,}")
        section("Tail spend", [
            (f"Vendors making up 80% of {m}", f"{tail['vendors_to_80pct']:,}"),
            ("Tail vendors (the rest)", f"{tail['tail_vendors']:,}  ({tail['tail_pct_of_vendors']}%)"),
            (f"Tail {m}", f"{fmt(tail['tail_value'])}  ({tail['tail_pct_of_value']}%)")])
    if sm:
        vfmt = _money_text if sm["has_amount"] else (lambda v: f"{int(v):,}")
        section("Sourcing", [
            ("Single-source categories", f"{sm['single_categories']}   ({vfmt(sm['single_value'])})"),
            ("Multi-source categories", f"{sm['multi_categories']}   ({vfmt(sm['multi_value'])})")])
    r = state["row"]
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    n = ws.cell(row=r, column=2, value="Tail vendors and multi-source categories are the primary "
                                       "consolidation targets.")
    n.font = Font(size=9, italic=True, color=GRAY)
    n.alignment = Alignment(wrap_text=True)


def _write_contents_sheet(ws, entries):
    """entries: list of (tab_name, why-it-matters)."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2.5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 82
    ws.merge_cells("B2:C2")
    t = ws["B2"]
    t.value = "Contents — what's in this report, and why it matters"
    t.font = Font(size=14, bold=True, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[2].height = 24
    for j, head in enumerate(["Tab", "Why this matters in a spend analysis"], start=2):
        c = ws.cell(row=4, column=j, value=head)
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.border = border
        c.alignment = Alignment(vertical="center")
    r = 5
    for name, why in entries:
        band = (r % 2 == 1)
        a = ws.cell(row=r, column=2, value=name)
        b = ws.cell(row=r, column=3, value=why)
        a.font = Font(bold=True, size=11, color=NAVY)
        b.font = Font(size=11)
        b.alignment = Alignment(wrap_text=True, vertical="top")
        a.alignment = Alignment(vertical="top")
        for cc in (a, b):
            cc.border = border
            if band:
                cc.fill = PatternFill("solid", fgColor=LT_BLUE)
        ws.row_dimensions[r].height = 30
        r += 1


_TAB_WHY = {
    "Executive Summary": "The headline findings in plain language, plus the recommended first steps — read this first.",
    "Vendor Analytics": "Supplier concentration (HHI), tail spend, and single- vs multi-source — where supply risk and consolidation opportunity live.",
    "Spend by Category": "Where the money goes by WHAT is bought — the core view for category management.",
    "Spend Trend": "Spend by year with year-over-year change — is spend rising or falling, and when did it peak?",
    "Pareto 80-20": "The few categories that drive ~80% of spend — where to focus sourcing effort first.",
    "Top Vendors": "The largest suppliers by spend — rationalization and negotiation targets.",
    "Spend by Department": "Which departments spend, and on what — surfaces cross-department demand.",
    "Category x Dept Matrix": "A cross-tab of category by department — the same commodity bought across many departments.",
    "Consolidation": "Categories bought from many vendors — the clearest consolidation (savings) opportunities.",
    "Same Item - Vendors_Depts": "The SAME item bought from more than one vendor and/or across more than one department — line-level maverick/fragmented buying and the sharpest consolidation targets.",
    "Same Commodity NIGP - Vendors_Depts": "The same NIGP commodity code (5-digit item where present, else 3-digit class) bought from more than one vendor and/or across more than one department — groups differently-worded descriptions of the same commodity.",
    "Savings Opportunity Summary": "The headline: addressable spend, identified savings split into hard (cashable) and cost avoidance, by opportunity type, annualized and projected three years.",
    "Top Consolidation Opportunities": "Each fragmented commodity with its addressable spend, savings rate, estimated hard + cost-avoidance savings, and a plain-English recommended action.",
    "Sources & Methodology": "Where the math and benchmark rates come from, who uses them (public and private), the addressable-spend and hard-vs-avoidance definitions, and reference sources.",
}


def build_excel_report(path, *, tiles, cat_tbl, pareto_tbl, vendors_tbl,
                       consolidation_tbl, cov, exec_summary=None, dept_tbl=None,
                       trend_tbl=None, tail=None, concentration=None, single_multi=None,
                       matrix_tbl=None, dimensions=None, item_consol_tbl=None,
                       nigp_consol_tbl=None, savings=None, opps_tbl=None,
                       report_title="Procurement Spend Analysis") -> str:
    has_amt = tiles.get("has_amount")
    measure = "Spend" if has_amt else "Transactions"
    money = ["Spend"] if has_amt else []
    used = set()

    def col_idx(df, name):
        return list(df.columns).index(name) + 1

    def anchor(ws, df):
        return ws.cell(row=2, column=df.shape[1] + 2).coordinate

    ylab = "Spend ($)" if has_amt else "Transactions"

    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        s_cat = _safe_sheet_name("Spend by Category", used)
        cat_tbl.to_excel(xl, sheet_name=s_cat, index=False, startrow=1)
        s_trend = None
        if trend_tbl is not None and len(trend_tbl):
            s_trend = _safe_sheet_name("Spend Trend", used)
            trend_tbl.to_excel(xl, sheet_name=s_trend, index=False, startrow=1)
        s_par = _safe_sheet_name("Pareto 80-20", used)
        pareto_tbl.to_excel(xl, sheet_name=s_par, index=False, startrow=1)
        s_ven = None
        if vendors_tbl is not None:
            s_ven = _safe_sheet_name("Top Vendors", used)
            vendors_tbl.to_excel(xl, sheet_name=s_ven, index=False, startrow=1)
        s_dep = None
        if dept_tbl is not None:
            s_dep = _safe_sheet_name("Spend by Department", used)
            dept_tbl.to_excel(xl, sheet_name=s_dep, index=False, startrow=1)
        s_mat = None
        if matrix_tbl is not None and len(matrix_tbl):
            s_mat = _safe_sheet_name("Category x Dept Matrix", used)
            matrix_tbl.to_excel(xl, sheet_name=s_mat, index=False, startrow=1)
        s_con = None
        if consolidation_tbl is not None:
            s_con = _safe_sheet_name("Consolidation", used)
            consolidation_tbl.to_excel(xl, sheet_name=s_con, index=False, startrow=1)
        s_item = None
        if item_consol_tbl is not None and len(item_consol_tbl):
            s_item = _safe_sheet_name("Same Item - Vendors_Depts", used)
            item_consol_tbl.to_excel(xl, sheet_name=s_item, index=False, startrow=1)
        s_nigp = None
        if nigp_consol_tbl is not None and len(nigp_consol_tbl):
            s_nigp = _safe_sheet_name("Same Commodity NIGP - Vendors_Depts", used)
            nigp_consol_tbl.to_excel(xl, sheet_name=s_nigp, index=False, startrow=1)
        s_opps = None
        if opps_tbl is not None and len(opps_tbl):
            s_opps = _safe_sheet_name("Top Consolidation Opportunities", used)
            opps_tbl.to_excel(xl, sheet_name=s_opps, index=False, startrow=1)
        dim_sheets = []
        for label, dfd in (dimensions or []):
            if dfd is None or not len(dfd):
                continue
            snm = _safe_sheet_name(label, used)
            dfd.to_excel(xl, sheet_name=snm, index=False, startrow=1)
            dim_sheets.append((snm, label, dfd))

        wb = xl.book
        es = wb.create_sheet("Executive Summary", 0)
        _write_exec_summary(es, report_title, tiles, cov, exec_summary)
        # Savings Opportunity Summary — the hero — right after the executive summary.
        s_save = None
        if savings and has_amt and savings.get("identified"):
            sv = wb.create_sheet("Savings Opportunity Summary", 1)
            _write_savings_sheet(sv, savings, has_amt)
            s_save = "Savings Opportunity Summary"
        contents_idx = 2 if s_save else 1
        analytics_idx = contents_idx + 1
        # Contents
        entries = [("Executive Summary", _TAB_WHY["Executive Summary"])]
        if s_save:
            entries.append(("Savings Opportunity Summary", _TAB_WHY["Savings Opportunity Summary"]))
        if s_opps:
            entries.append((s_opps, _TAB_WHY["Top Consolidation Opportunities"]))
        if tail or concentration or single_multi:
            entries.append(("Vendor Analytics", _TAB_WHY["Vendor Analytics"]))
        for snm, key in [(s_cat, "Spend by Category"), (s_trend, "Spend Trend"),
                         (s_par, "Pareto 80-20"), (s_ven, "Top Vendors"),
                         (s_dep, "Spend by Department"), (s_mat, "Category x Dept Matrix"),
                         (s_con, "Consolidation"), (s_item, "Same Item - Vendors_Depts"),
                         (s_nigp, "Same Commodity NIGP - Vendors_Depts")]:
            if snm:
                entries.append((snm, _TAB_WHY[key]))
        for snm, label, _dfd in dim_sheets:
            entries.append((snm, f"Spend broken down by “{label}” — e.g., on- vs off-contract, "
                                 "supplier diversity, or process status, depending on the column."))
        entries.append(("Sources & Methodology", _TAB_WHY["Sources & Methodology"]))
        contents = wb.create_sheet("Contents", contents_idx)
        _write_contents_sheet(contents, entries)
        if tail or concentration or single_multi:
            an = wb.create_sheet("Vendor Analytics", analytics_idx)
            _write_analytics_sheet(an, tail, concentration, single_multi)
        # Sources & Methodology — always last.
        src = wb.create_sheet("Sources & Methodology")
        _write_sources_sheet(src, savings.get("rates") if savings else None)

        def _below(dl):
            return f"A{dl + 3}"

        ws = wb[s_cat]
        hr, df1, dl = _style_data_sheet(ws, "Spend by Business Category", cat_tbl,
            money_cols=money, pct_cols=["% of Total"], int_cols=["Transactions"],
            total_cols=[measure] + (["Transactions"] if has_amt else []),
            wrap_cols=[EXAMPLES_COL])
        _embed_png(ws, barh_png(cat_tbl["Business Category"].tolist(), cat_tbl[measure].tolist(),
                                f"{measure} by Business Category", value_label=ylab, money=has_amt),
                   _below(dl))

        if s_trend:
            ws = wb[s_trend]
            hr, df1, dl = _style_data_sheet(ws, "Spend Trend Over Time", trend_tbl,
                money_cols=money, pct_cols=[c for c in ["YoY %"] if c in trend_tbl.columns],
                int_cols=["Transactions"], add_total=False)
            _embed_png(ws, line_png(trend_tbl["Year"].tolist(), trend_tbl[measure].tolist(),
                                    f"{measure} by Year", value_label=ylab), _below(dl))

        ws = wb[s_par]
        hr, df1, dl = _style_data_sheet(ws, "Pareto 80/20 (classified spend)", pareto_tbl,
            money_cols=money, pct_cols=["% of Total", "Cumulative %"], int_cols=["Transactions"], add_total=False)
        _pcat = "Group" if "Group" in pareto_tbl.columns else pareto_tbl.columns[0]
        _embed_png(ws, pareto_png(pareto_tbl[_pcat].tolist(), pareto_tbl[measure].tolist(),
                                  pareto_tbl["Cumulative %"].tolist(), "Pareto 80/20", value_label=ylab),
                   _below(dl))

        if s_ven:
            ws = wb[s_ven]
            hr, df1, dl = _style_data_sheet(ws, "Top Vendors by Spend", vendors_tbl,
                money_cols=money, pct_cols=["% of Total"], int_cols=["Transactions"],
                total_cols=[measure] + (["Transactions"] if has_amt else []))
            _embed_png(ws, barh_png(vendors_tbl["Vendor"].tolist(), vendors_tbl[measure].tolist(),
                                    f"Top Vendors by {measure}", value_label=ylab, money=has_amt),
                       _below(dl))

        if s_dep:
            ws = wb[s_dep]
            hr, df1, dl = _style_data_sheet(ws, "Spend by Department", dept_tbl,
                money_cols=money, pct_cols=["% of Total"], int_cols=["Transactions"],
                total_cols=[measure] + (["Transactions"] if has_amt else []))
            _embed_png(ws, barh_png(dept_tbl["Department"].tolist(), dept_tbl[measure].tolist(),
                                    f"{measure} by Department", value_label=ylab, money=has_amt),
                       _below(dl))

        if s_mat:
            ws = wb[s_mat]
            _style_data_sheet(ws, "Category × Department Matrix", matrix_tbl,
                money_cols=[c for c in matrix_tbl.columns if c != "Business Category"], add_total=False)

        if s_con:
            ws = wb[s_con]
            int_c = [c for c in ["Vendors", "Transactions", "Departments"] if c in consolidation_tbl.columns]
            _style_data_sheet(ws, "Vendor Consolidation / Fragmentation", consolidation_tbl,
                money_cols=money, int_cols=int_c,
                total_cols=(["Spend", "Transactions"] if has_amt else ["Transactions"]))

        if s_item:
            ws = wb[s_item]
            int_c = [c for c in ["Vendors", "Departments", "Transactions"] if c in item_consol_tbl.columns]
            wrap_c = [c for c in ["Item", ITEM_VENDORS_COL, ITEM_DEPTS_COL] if c in item_consol_tbl.columns]
            _style_data_sheet(ws, "Same Item — Multiple Vendors / Departments", item_consol_tbl,
                money_cols=money, int_cols=int_c, add_total=False, wrap_cols=wrap_c)

        if s_nigp:
            from openpyxl.styles import Alignment, Font, PatternFill
            ws = wb[s_nigp]
            cols_list = list(nigp_consol_tbl.columns)
            dept_cols = nigp_dept_columns(nigp_consol_tbl)
            total_c = [c for c in ["Total Spend", "Total Transactions"] if c in cols_list]
            int_c = [c for c in ["Vendors", "Departments (#)"] if c in cols_list]
            wrap_c = [c for c in [DESC_VARIANTS_COL, ITEM_VENDORS_COL] if c in cols_list]
            hr, dfirst, dlast = _style_data_sheet(
                ws, "Same Commodity (NIGP code) — Multiple Vendors / Departments, by department",
                nigp_consol_tbl, money_cols=(total_c if has_amt else []),
                int_cols=int_c + ([] if has_amt else total_c), add_total=False, wrap_cols=wrap_c)
            dept_idx = [cols_list.index(c) + 1 for c in dept_cols]
            var_idx = cols_list.index(DESC_VARIANTS_COL) + 1 if DESC_VARIANTS_COL in cols_list else None
            zero_fmt = '$#,##0;-$#,##0;""' if has_amt else '#,##0;;""'
            # compact, wrapped department headers so the matrix stays readable
            for ci in dept_idx:
                ws.column_dimensions[ws.cell(row=hr, column=ci).column_letter].width = 15
                hc = ws.cell(row=hr, column=ci)
                hc.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
            ws.row_dimensions[hr].height = 46
            # highlight which departments buy each commodity, and flag multi-dept rows
            for r in range(dfirst, dlast + 1):
                nonzero = 0
                for ci in dept_idx:
                    cell = ws.cell(row=r, column=ci)
                    cell.number_format = zero_fmt
                    cell.alignment = Alignment(horizontal="right")
                    try:
                        v = float(cell.value or 0)
                    except (TypeError, ValueError):
                        v = 0
                    if v > 0:
                        cell.fill = PatternFill("solid", fgColor="D6EEF9")  # light-blue: bought here
                        nonzero += 1
                if nonzero >= 2 and var_idx:
                    vc = ws.cell(row=r, column=var_idx)
                    vc.fill = PatternFill("solid", fgColor="FCE4E4")  # light-red: consolidation target
                    vc.font = Font(bold=True, color="9C1006")

        if s_opps:
            ws = wb[s_opps]
            ocols = list(opps_tbl.columns)
            money_c = [c for c in [COL_SPEND, COL_TOTAL, COL_HARD, COL_AVOID] if c in ocols]
            int_c = [c for c in [COL_VENDORS, COL_DEPTS] if c in ocols]
            pct_c = [c for c in [COL_RATE] if c in ocols]
            wrap_c = [c for c in [COL_ITEM, COL_ACTION] if c in ocols]
            _style_data_sheet(ws, "What to combine — estimated cash savings + avoided costs",
                opps_tbl, money_cols=money_c, int_cols=int_c, pct_cols=pct_c,
                total_cols=money_c, add_total=True, wrap_cols=wrap_c)

        for snm, label, dfd in dim_sheets:
            ws = wb[snm]
            hr, df1, dl = _style_data_sheet(ws, label, dfd,
                money_cols=money, pct_cols=["% of Total"], int_cols=["Transactions"],
                total_cols=[measure] + (["Transactions"] if has_amt else []))
            _embed_png(ws, barh_png([str(x) for x in dfd.iloc[:, 0].tolist()], dfd[measure].tolist(),
                                    label[:48], value_label=ylab, money=has_amt), _below(dl))

    return path


# ---------------------------------------------------------------------------
# Orchestrator — compute every feasible analysis given detected roles
# ---------------------------------------------------------------------------
def compute_all(df, roles, max_dims=6):
    """df must already carry a 'Business_Category' column (from classify_series).
    Returns a dict of every analysis the roles support; infeasible ones are None."""
    amt, ven = roles.get("amount"), roles.get("vendor")
    dep, dat = roles.get("department"), roles.get("date")
    desc = roles.get("description")
    # Absorb the catch-all so every row lands in a real Business Category.
    cov = coverage(df, "Business_Category")  # capture best-fit count before folding
    df = absorb_catchall(df, desc)
    tiles = summary_tiles(df, "Business_Category", amt, ven, dat)
    cat = spend_by_category(df, "Business_Category", amt, desc_col=desc)
    classified_df = df[_is_classified(df["Business_Category"])]
    par, measure, n80 = pareto(classified_df, "Business_Category", amt)
    vend = top_vendors(df, ven, amt)
    dept_tbl = spend_by_department(df, dep, amt)
    cons = consolidation_finder(df, "Business_Category", ven, amt, dep)
    item_consol = item_consolidation(df, desc, amt, ven, dep)
    nigp_consol = nigp_item_consolidation(df, desc, amt, ven, dep)
    trend = spend_trend(df, dat, amt)
    tail = tail_spend(df, ven, amt)
    conc = vendor_concentration(df, ven, amt)
    sm = single_vs_multi_source(df, "Business_Category", ven, amt) if ven else None
    matrix = category_department_matrix(df, "Business_Category", dep, amt)
    # Savings / cost-avoidance engine
    opps = consolidation_opportunities(df, desc, amt, ven, dep)
    years = len(trend) if (trend is not None and len(trend) >= 1) else 1
    savings = savings_summary(opps, tiles, tail=tail, years=years)

    subs = roles.get("dimension_subtypes", {})
    dim_names = sorted(roles.get("dimensions", []), key=lambda d: 0 if subs.get(d) else 1)
    dims = []
    for d in dim_names[:max_dims]:
        sub = subs.get(d)
        label = {"contract": f"Contract vs Non-Contract ({d})",
                 "diversity": f"Supplier Diversity ({d})"}.get(sub, f"Spend by {d}")
        tbl = spend_by_dimension(df, d, amt)
        if tbl is not None and len(tbl):
            dims.append((label, tbl))

    es = executive_summary(tiles, cat, n80, vend, cons, cov, dept_tbl,
                           trend=trend, tail=tail, concentration=conc,
                           item_consol=item_consol, savings=savings, opps=opps)
    return dict(tiles=tiles, cov=cov, cat=cat, par=par, measure=measure, n80=n80,
                vend=vend, dept_tbl=dept_tbl, cons=cons, item_consol=item_consol,
                nigp_consol=nigp_consol, trend=trend, tail=tail, concentration=conc,
                single_multi=sm, matrix=matrix, dimensions=dims,
                opps=opps, savings=savings, es=es, _work=df, _roles=dict(roles))


def recompute_savings(b, rates):
    """Re-run ONLY the savings layer (opportunities, summary, executive narrative)
    against the already-classified data in bundle `b`, using new benchmark `rates`.
    Returns (opps, savings, es). Classification is not re-run — instant for sliders."""
    work = b.get("_work")
    roles = b.get("_roles") or {}
    if work is None:
        return b.get("opps"), b.get("savings"), b.get("es")
    amt, ven = roles.get("amount"), roles.get("vendor")
    dep, desc = roles.get("department"), roles.get("description")
    opps = consolidation_opportunities(work, desc, amt, ven, dep, rates=rates)
    trend = b.get("trend")
    years = len(trend) if (trend is not None and len(trend) >= 1) else 1
    savings = savings_summary(opps, b["tiles"], tail=b.get("tail"), rates=rates, years=years)
    es = executive_summary(b["tiles"], b["cat"], b["n80"], b["vend"], b["cons"], b["cov"],
                           b["dept_tbl"], trend=trend, tail=b.get("tail"),
                           concentration=b.get("concentration"), item_consol=b.get("item_consol"),
                           savings=savings, opps=opps)
    return opps, savings, es


# ---------------------------------------------------------------------------
# Standalone runner (for testing)
# ---------------------------------------------------------------------------
def run(path, desc=None, amount=None, vendor=None, dept=None, out=None, sample=None):
    if str(path).lower().endswith(".csv"):
        try:
            df = pd.read_csv(path, low_memory=False)
        except UnicodeDecodeError:
            df = pd.read_csv(path, low_memory=False, encoding="latin-1")
    else:
        df = pd.read_excel(path)
    df.columns = [str(c).strip() for c in df.columns]
    if sample:
        df = df.head(int(sample))
    prof = profile_columns(df)
    roles = prof["roles"]
    # allow CLI overrides
    if desc: roles["description"] = desc
    if amount: roles["amount"] = amount
    if vendor: roles["vendor"] = vendor
    if dept: roles["department"] = dept
    print("Detected roles:", {k: roles[k] for k in ["description", "amount", "vendor", "department", "date"]})
    print("Dimensions:", roles["dimensions"])
    print("\nAnalysis plan:")
    for p in plan_analyses(prof):
        print(f"  [{'RUN ' if p['feasible'] else 'SKIP'}] {p['label']}"
              + (f"  ({p['reason']})" if not p["feasible"] else ""))
    print(f"\nClassifying {len(df):,} rows ...")
    cls = classify_series(df[roles["description"]])
    df = pd.concat([df.reset_index(drop=True), cls.reset_index(drop=True)], axis=1)

    b = compute_all(df, roles)
    print("\n=== EXECUTIVE SUMMARY ===")
    for ln in b["es"]["lines"]:
        print("  • " + ln)
    print("  Recommended steps:")
    for i, s in enumerate(b["es"]["steps"], 1):
        print(f"   {i}. {s}")
    print(f"\n=== COVERAGE === classified {b['cov']['classified']:,}/{b['cov']['total']:,} "
          f"({b['cov']['classified_pct']}%)")

    out = out or (Path(path).with_suffix("").as_posix() + "_Spend_Report_JHK3.xlsx")
    build_excel_report(out, tiles=b["tiles"], cat_tbl=b["cat"], pareto_tbl=b["par"],
                       vendors_tbl=b["vend"], consolidation_tbl=b["cons"], cov=b["cov"],
                       exec_summary=b["es"], dept_tbl=b["dept_tbl"], trend_tbl=b["trend"],
                       tail=b["tail"], concentration=b["concentration"],
                       single_multi=b["single_multi"], matrix_tbl=b["matrix"],
                       dimensions=b["dimensions"], item_consol_tbl=b.get("item_consol"),
                       nigp_consol_tbl=b.get("nigp_consol"), savings=b.get("savings"),
                       opps_tbl=b.get("opps"))
    print(f"\nWrote Excel report: {out}")
    return out


def main():
    p = argparse.ArgumentParser()
    p.add_argument("path")
    p.add_argument("--desc"); p.add_argument("--amount")
    p.add_argument("--vendor"); p.add_argument("--dept")
    p.add_argument("--out"); p.add_argument("--sample", type=int)
    a = p.parse_args()
    run(a.path, a.desc, a.amount, a.vendor, a.dept, a.out, a.sample)


if __name__ == "__main__":
    main()
