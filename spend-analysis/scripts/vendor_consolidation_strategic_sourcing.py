"""
City of Chicago - Landscaping Vendor Consolidation & Strategic Sourcing Analysis
Demonstrates cost savings, cost avoidance, and the need for category management
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from datetime import datetime
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

# ============================================================================
# CONFIGURATION & CONSTANTS
# ============================================================================

# City of Chicago colors
CHICAGO_BLUE = "1F4788"
CHICAGO_RED = "DC143C"
LIGHT_BLUE = "C5D9F1"
WHITE = "FFFFFF"
GRAY = "D3D3D3"

# File paths
input_path = "/workspaces/Spend_Analysis_and_Categorization/spend-analysis/data/processed/landscaping_structured.xlsx"
output_dir = Path("/workspaces/Spend_Analysis_and_Categorization/spend-analysis/outputs")
charts_dir = output_dir / "charts"
charts_dir.mkdir(parents=True, exist_ok=True)
output_file = output_dir / "VENDOR_CONSOLIDATION_STRATEGIC_SOURCING_ANALYSIS.xlsx"

# Accounting number format
accounting_format = '_($* #,##0_);_($* (#,##0));_($* "-"_);_(@_)'

# Industry benchmarks
ADMIN_COST_PER_VENDOR = 50000  # $50K per vendor eliminated
INFLATION_RATE = 0.045  # 4.5% annual (BLS landscaping CPI)

# Consolidation scenarios
SCENARIOS = [
    {'name': 'Conservative', 'reduction_pct': 26, 'savings_rate': 0.04},
    {'name': 'Moderate', 'reduction_pct': 45, 'savings_rate': 0.08},
    {'name': 'Aggressive', 'reduction_pct': 65, 'savings_rate': 0.13}
]

print("\n" + "="*100)
print("CITY OF CHICAGO - VENDOR CONSOLIDATION & STRATEGIC SOURCING ANALYSIS")
print("="*100)
print("\nGenerating comprehensive analysis with cost savings, cost avoidance, and visualizations...")

# ============================================================================
# SECTION 1: DATA LOADING & BASELINE ANALYSIS
# ============================================================================
print("\n1. Loading landscaping spend data...")

df = pd.read_excel(input_path)

# Calculate baseline metrics
total_spend = df['Award_Amount_Numeric'].sum()
total_vendors = df['Vendor name'].nunique()
total_contracts = len(df)
avg_contract_value = df['Award_Amount_Numeric'].mean()
median_contract_value = df['Award_Amount_Numeric'].median()

print(f"   Total Spend: ${total_spend:,.2f}")
print(f"   Total Vendors: {total_vendors}")
print(f"   Total Contracts: {total_contracts}")

# Vendor analysis
vendor_spend = df.groupby('Vendor name').agg({
    'Award_Amount_Numeric': 'sum',
    'Contract (PO) #': 'count',
    'Service_Category': lambda x: len(x.unique())
}).reset_index()
vendor_spend.columns = ['Vendor', 'Total Spend', 'Contract Count', 'Service Categories']
vendor_spend = vendor_spend.sort_values('Total Spend', ascending=False)
vendor_spend['% of Total'] = (vendor_spend['Total Spend'] / total_spend * 100)
vendor_spend['Cumulative %'] = vendor_spend['% of Total'].cumsum()

# Service category analysis
service_spend = df.groupby('Service_Category').agg({
    'Award_Amount_Numeric': 'sum',
    'Vendor name': 'nunique',
    'Contract (PO) #': 'count'
}).reset_index()
service_spend.columns = ['Service Category', 'Total Spend', 'Vendor Count', 'Contract Count']
service_spend = service_spend.sort_values('Total Spend', ascending=False)
service_spend['% of Total'] = (service_spend['Total Spend'] / total_spend * 100)
service_spend['Avg Contract Size'] = service_spend['Total Spend'] / service_spend['Contract Count']

# Department analysis
dept_spend = df.groupby('Department').agg({
    'Award_Amount_Numeric': 'sum',
    'Contract (PO) #': 'count',
    'Vendor name': 'nunique'
}).reset_index()
dept_spend.columns = ['Department', 'Total Spend', 'Contract Count', 'Vendor Count']
dept_spend = dept_spend.sort_values('Total Spend', ascending=False)

# ============================================================================
# SECTION 2: CONSOLIDATION SCENARIO CALCULATIONS
# ============================================================================
print("\n2. Calculating vendor consolidation scenarios...")

def calculate_consolidation_scenario(name, reduction_pct, savings_rate, total_spend, total_vendors):
    """Calculate savings for a consolidation scenario"""
    target_vendors = int(total_vendors * (1 - reduction_pct/100))
    vendors_eliminated = total_vendors - target_vendors

    volume_savings = total_spend * savings_rate
    admin_savings = vendors_eliminated * ADMIN_COST_PER_VENDOR
    total_savings = volume_savings + admin_savings

    return {
        'Scenario': name,
        'Current Vendors': total_vendors,
        'Target Vendors': target_vendors,
        'Vendors Eliminated': vendors_eliminated,
        'Reduction %': reduction_pct,
        'Volume Savings': volume_savings,
        'Admin Savings': admin_savings,
        'Total Annual Savings': total_savings
    }

scenarios_list = []
for scenario in SCENARIOS:
    result = calculate_consolidation_scenario(
        scenario['name'],
        scenario['reduction_pct'],
        scenario['savings_rate'],
        total_spend,
        total_vendors
    )
    scenarios_list.append(result)
    print(f"   {result['Scenario']}: ${result['Total Annual Savings']:,.0f}/year " +
          f"({result['Target Vendors']} vendors, {result['Reduction %']}% reduction)")

scenarios_df = pd.DataFrame(scenarios_list)

# ============================================================================
# SECTION 3: COST AVOIDANCE CALCULATIONS
# ============================================================================
print("\n3. Calculating 3-year cost avoidance...")

def calculate_cost_avoidance(years, inflation_rate, baseline_spend):
    """Calculate cost avoidance from strategic sourcing rate locks"""
    avoidance_by_year = []

    for year in range(1, years + 1):
        inflated_cost = baseline_spend * ((1 + inflation_rate) ** year)
        cost_with_lock = baseline_spend
        avoidance = inflated_cost - cost_with_lock

        avoidance_by_year.append({
            'Year': year,
            'Without Strategic Sourcing': inflated_cost,
            'With Strategic Sourcing (Rate Lock)': cost_with_lock,
            'Cost Avoidance': avoidance
        })

    return pd.DataFrame(avoidance_by_year)

cost_avoidance_df = calculate_cost_avoidance(3, INFLATION_RATE, total_spend)
total_3yr_avoidance = cost_avoidance_df['Cost Avoidance'].sum()

print(f"   Year 1 Avoidance: ${cost_avoidance_df.iloc[0]['Cost Avoidance']:,.0f}")
print(f"   Year 2 Cumulative: ${cost_avoidance_df.iloc[0:2]['Cost Avoidance'].sum():,.0f}")
print(f"   Year 3 Cumulative: ${total_3yr_avoidance:,.0f}")

# ============================================================================
# SECTION 4: VISUALIZATION GENERATION
# ============================================================================
print("\n4. Creating professional visualizations...")

# Set matplotlib style
sns.set_style("whitegrid")
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.sans-serif'] = ['Arial', 'DejaVu Sans']

# Chart 1: Vendor Concentration by Service Category
print("   Creating Chart 1: Vendor Concentration by Service Category...")
fig, ax = plt.subplots(figsize=(14, 10))

categories = service_spend['Service Category'].values
vendor_counts = service_spend['Vendor Count'].values

bars = ax.barh(categories, vendor_counts, color='#1F4788', alpha=0.7, edgecolor='black', linewidth=1.5)
ax.set_xlabel('Number of Vendors', fontsize=12, fontweight='bold')
ax.set_ylabel('Service Category', fontsize=12, fontweight='bold')
ax.set_title('Vendor Concentration by Service Category\nConsolidation Opportunity Analysis',
             fontsize=14, fontweight='bold', pad=20)

# Add data labels
for bar, count in zip(bars, vendor_counts):
    width = bar.get_width()
    ax.text(width + 0.5, bar.get_y() + bar.get_height()/2, f'{int(count)} vendors',
           ha='left', va='center', fontsize=10, fontweight='bold')

ax.grid(axis='x', alpha=0.3)
ax.set_xlim([0, max(vendor_counts) + 5])
plt.tight_layout()
chart1_path = charts_dir / '01_vendor_concentration_by_category.png'
plt.savefig(chart1_path, dpi=300, bbox_inches='tight')
plt.close()

# Chart 2: Current vs. Consolidated Vendor Counts
print("   Creating Chart 2: Before/After Vendor Consolidation...")
fig, ax = plt.subplots(figsize=(14, 8))

scenario_names = ['Current\nBaseline'] + [s['Scenario'] + '\nScenario' for s in scenarios_list]
vendor_counts_chart = [total_vendors] + [s['Target Vendors'] for s in scenarios_list]
reductions = [0] + [s['Reduction %'] for s in scenarios_list]

colors = ['#DC143C'] + ['#28a745', '#20803a', '#186830']  # Red for current, greens for scenarios
bars = ax.bar(scenario_names, vendor_counts_chart, color=colors, alpha=0.8, edgecolor='black', linewidth=2)

ax.set_ylabel('Number of Vendors', fontsize=12, fontweight='bold')
ax.set_title('Vendor Consolidation Impact\nCurrent State vs. Consolidation Scenarios',
             fontsize=14, fontweight='bold', pad=20)
ax.set_ylim([0, 100])
ax.grid(axis='y', alpha=0.3)

# Add data labels
for i, (bar, count, reduction) in enumerate(zip(bars, vendor_counts_chart, reductions)):
    height = bar.get_height()
    if i == 0:
        label_text = f'{int(count)}\nvendors'
    else:
        label_text = f'{int(count)}\nvendors\n(-{reduction}%)'
    ax.text(bar.get_x() + bar.get_width()/2, height + 2,
           label_text, ha='center', va='bottom', fontsize=11, fontweight='bold')

plt.tight_layout()
chart2_path = charts_dir / '02_before_after_consolidation.png'
plt.savefig(chart2_path, dpi=300, bbox_inches='tight')
plt.close()

# Chart 3: Cost Savings Waterfall (Moderate Scenario)
print("   Creating Chart 3: Cost Savings Waterfall...")
moderate_scenario = scenarios_list[1]  # Moderate scenario

fig, ax = plt.subplots(figsize=(16, 10))

categories_waterfall = [
    'Current\nAnnual\nSpend',
    'Volume\nLeverage\nSavings',
    'Administrative\nSavings',
    '3-Year\nCost\nAvoidance',
    'Net 3-Year\nPosition'
]

current_spend_millions = total_spend / 1_000_000
volume_savings_millions = moderate_scenario['Volume Savings'] / 1_000_000
admin_savings_millions = moderate_scenario['Admin Savings'] / 1_000_000
cost_avoidance_millions = total_3yr_avoidance / 1_000_000
net_position = current_spend_millions - volume_savings_millions - admin_savings_millions - cost_avoidance_millions

values = [current_spend_millions, -volume_savings_millions, -admin_savings_millions, -cost_avoidance_millions, net_position]

# Calculate bar positions for waterfall effect
cumulative = [0]
for i, val in enumerate(values[:-1]):
    cumulative.append(cumulative[-1] + val)

colors_waterfall = ['#1F4788', '#28a745', '#28a745', '#28a745', '#1F4788']

# Create bars
for i, (cat, val, cum, color) in enumerate(zip(categories_waterfall, values, cumulative, colors_waterfall)):
    if i == 0 or i == len(categories_waterfall) - 1:
        # Full bars for start and end
        ax.bar(i, abs(val), bottom=0, color=color, edgecolor='black', linewidth=2, alpha=0.8)
    else:
        # Floating bars for deltas
        bottom_val = cum + val if val < 0 else cum
        ax.bar(i, abs(val), bottom=bottom_val, color=color, edgecolor='black', linewidth=2, alpha=0.8)

# Add connecting lines
for i in range(len(cumulative) - 1):
    next_bottom = cumulative[i+1] + values[i+1] if values[i+1] < 0 else cumulative[i+1]
    curr_top = cumulative[i] + values[i]
    ax.plot([i + 0.4, i + 0.6], [curr_top, next_bottom], 'k--', linewidth=1, alpha=0.5)

# Data labels
for i, (val, cum) in enumerate(zip(values, cumulative)):
    if i == 0:
        label_y = val / 2
        label_text = f'${abs(val):.1f}M'
    elif i == len(values) - 1:
        label_y = val / 2
        label_text = f'${abs(val):.1f}M'
    else:
        label_y = cum + val / 2
        label_text = f'-${abs(val):.1f}M'

    ax.text(i, label_y, label_text, ha='center', va='center',
           fontsize=12, fontweight='bold', bbox=dict(boxstyle='round', facecolor='white', alpha=0.8))

ax.set_ylabel('Amount ($ Millions)', fontsize=12, fontweight='bold')
ax.set_title('Cost Savings Waterfall Analysis\nModerate Consolidation Scenario (3-Year View)',
            fontsize=14, fontweight='bold', pad=20)
ax.set_xticks(range(len(categories_waterfall)))
ax.set_xticklabels(categories_waterfall, fontsize=11)
ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x:.0f}M'))
ax.grid(axis='y', alpha=0.3)

# Total savings annotation
total_savings_3yr = volume_savings_millions + admin_savings_millions + cost_avoidance_millions
ax.text(0.5, 0.95, f'Total 3-Year Financial Impact: ${total_savings_3yr:.1f}M',
       transform=fig.transFigure, ha='center', fontsize=14,
       bbox=dict(boxstyle='round', facecolor='#FFFF00', alpha=0.9, edgecolor='black', linewidth=2),
       fontweight='bold')

plt.tight_layout()
chart3_path = charts_dir / '03_cost_savings_waterfall.png'
plt.savefig(chart3_path, dpi=300, bbox_inches='tight')
plt.close()

# Chart 4: Pareto Chart - Top 20 Vendors
print("   Creating Chart 4: Vendor Spend Pareto Analysis...")
fig, ax1 = plt.subplots(figsize=(16, 10))

top20_vendors = vendor_spend.head(20).copy()
vendor_names = top20_vendors['Vendor'].values
spend_millions = top20_vendors['Total Spend'].values / 1_000_000
cumulative_pct = top20_vendors['Cumulative %'].values

x_pos = np.arange(len(vendor_names))

# Bar chart
bars = ax1.bar(x_pos, spend_millions, color='#1F4788', alpha=0.7, edgecolor='black', linewidth=1)
ax1.set_xlabel('Vendor', fontsize=12, fontweight='bold')
ax1.set_ylabel('Total Spend ($ Millions)', fontsize=12, fontweight='bold', color='#1F4788')
ax1.tick_params(axis='y', labelcolor='#1F4788')
ax1.set_xticks(x_pos)
ax1.set_xticklabels(vendor_names, rotation=45, ha='right', fontsize=9)

# Cumulative line
ax2 = ax1.twinx()
line = ax2.plot(x_pos, cumulative_pct, color='#DC143C', marker='o', linewidth=3, markersize=6, label='Cumulative %')
ax2.set_ylabel('Cumulative Percentage (%)', fontsize=12, fontweight='bold', color='#DC143C')
ax2.tick_params(axis='y', labelcolor='#DC143C')
ax2.set_ylim([0, 105])
ax2.axhline(y=80, color='darkred', linestyle='--', linewidth=2, label='80% Threshold')

ax1.grid(axis='y', alpha=0.3)

plt.title('Pareto Analysis: Top 20 Vendors by Spend\n(80/20 Consolidation Opportunity)',
         fontsize=14, fontweight='bold', pad=20)

# Annotations
top5_pct = vendor_spend.head(5)['% of Total'].sum()
small_vendors = vendor_spend[vendor_spend['% of Total'] < 1]
small_vendors_count = len(small_vendors)
small_vendors_spend = small_vendors['Total Spend'].sum()

annotation_text = f'Top 5 vendors = {top5_pct:.1f}% of total spend\n' + \
                 f'{small_vendors_count} small vendors (<1%) = ${small_vendors_spend/1_000_000:.1f}M opportunity'
ax1.text(0.5, 0.95, annotation_text, transform=fig.transFigure,
        ha='center', fontsize=12, bbox=dict(boxstyle='round', facecolor='#FFFF00', alpha=0.8),
        fontweight='bold')

plt.tight_layout()
chart4_path = charts_dir / '04_pareto_vendor_spend.png'
plt.savefig(chart4_path, dpi=300, bbox_inches='tight')
plt.close()

# Chart 5: Service Category Bubble Chart
print("   Creating Chart 5: Service Category Bubble Chart...")
fig, ax = plt.subplots(figsize=(14, 10))

x = service_spend['Total Spend'].values / 1_000_000
y = service_spend['Contract Count'].values
sizes = service_spend['Vendor Count'].values * 100  # Scale bubble size
colors_palette = sns.color_palette("Set2", len(service_spend))

scatter = ax.scatter(x, y, s=sizes, c=colors_palette, alpha=0.6, edgecolors='black', linewidth=2)

# Add labels
for i, cat in enumerate(service_spend['Service Category'].values):
    ax.annotate(cat, (x[i], y[i]), fontsize=9, fontweight='bold',
               xytext=(5, 5), textcoords='offset points',
               bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.7))

ax.set_xlabel('Total Spend ($ Millions)', fontsize=12, fontweight='bold')
ax.set_ylabel('Number of Contracts', fontsize=12, fontweight='bold')
ax.set_title('Service Category Analysis\nSpend vs. Contracts vs. Vendor Count (Bubble Size)',
            fontsize=14, fontweight='bold', pad=20)
ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x:.0f}M'))
ax.grid(True, alpha=0.3)

plt.tight_layout()
chart5_path = charts_dir / '05_bubble_service_category.png'
plt.savefig(chart5_path, dpi=300, bbox_inches='tight')
plt.close()

# Chart 6: 3-Year Financial Projection
print("   Creating Chart 6: 3-Year Financial Projection...")
fig, ax = plt.subplots(figsize=(14, 8))

years = [0, 1, 2, 3]
volume_cumulative = [0] + [moderate_scenario['Volume Savings'] * i / 1_000_000 for i in range(1, 4)]
admin_cumulative = [0] + [moderate_scenario['Admin Savings'] * i / 1_000_000 for i in range(1, 4)]
avoidance_cumulative = [0] + [cost_avoidance_df.iloc[:i]['Cost Avoidance'].sum() / 1_000_000 for i in range(1, 4)]

# Stacked area chart
ax.fill_between(years, 0, volume_cumulative, label='Volume Leverage Savings', color='#186830', alpha=0.7)
ax.fill_between(years, volume_cumulative,
                [v + a for v, a in zip(volume_cumulative, admin_cumulative)],
                label='Administrative Savings', color='#20803a', alpha=0.7)
ax.fill_between(years, [v + a for v, a in zip(volume_cumulative, admin_cumulative)],
                [v + a + av for v, a, av in zip(volume_cumulative, admin_cumulative, avoidance_cumulative)],
                label='Cost Avoidance', color='#28a745', alpha=0.7)

# Baseline inflation scenario
baseline_costs = [total_spend/1_000_000 * ((1 + INFLATION_RATE) ** year) for year in years]
savings_scenario = [total_spend/1_000_000] * 4  # Locked rate
ax.plot(years, baseline_costs, 'r--', linewidth=3, label='Without Strategic Sourcing (Inflation)', marker='o', markersize=8)
ax.plot(years, savings_scenario, 'b-', linewidth=3, label='With Strategic Sourcing (Rate Lock)', marker='s', markersize=8)

# Data labels
for i, year in enumerate(years[1:], 1):
    total_benefit = volume_cumulative[i] + admin_cumulative[i] + avoidance_cumulative[i]
    ax.text(year, total_benefit + 5, f'${total_benefit:.1f}M', ha='center', fontsize=11, fontweight='bold')

ax.set_xlabel('Year', fontsize=12, fontweight='bold')
ax.set_ylabel('Cumulative Financial Impact ($ Millions)', fontsize=12, fontweight='bold')
ax.set_title('3-Year Financial Projection\nModerate Consolidation Scenario + Strategic Sourcing',
            fontsize=14, fontweight='bold', pad=20)
ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x:.0f}M'))
ax.legend(fontsize=10, loc='upper left')
ax.grid(True, alpha=0.3)
ax.set_xticks(years)

plt.tight_layout()
chart6_path = charts_dir / '06_three_year_projection.png'
plt.savefig(chart6_path, dpi=300, bbox_inches='tight')
plt.close()

# Chart 7: Category Management Heat Map
print("   Creating Chart 7: Category Management Priority Heat Map...")
fig, ax = plt.subplots(figsize=(12, 8))

# Calculate priority scores
service_spend['Spend Rank'] = service_spend['Total Spend'].rank(ascending=False)
service_spend['Vendor Rank'] = service_spend['Vendor Count'].rank(ascending=False)
service_spend['Priority Score'] = (service_spend['Vendor Rank'] / len(service_spend) * 0.5 +
                                   service_spend['Spend Rank'] / len(service_spend) * 0.5) * 10

# Create heat map data
heatmap_data = service_spend[['Service Category', 'Total Spend', 'Vendor Count', 'Priority Score']].copy()
heatmap_data['Total Spend ($M)'] = heatmap_data['Total Spend'] / 1_000_000
heatmap_values = heatmap_data[['Total Spend ($M)', 'Vendor Count', 'Priority Score']].values

# Create heat map
im = ax.imshow(heatmap_values, cmap='RdYlGn_r', aspect='auto')

# Set ticks
ax.set_xticks([0, 1, 2])
ax.set_xticklabels(['Total Spend ($M)', 'Vendor Count', 'Priority Score'], fontsize=11, fontweight='bold')
ax.set_yticks(range(len(heatmap_data)))
ax.set_yticklabels(heatmap_data['Service Category'].values, fontsize=10)

# Add values as text
for i in range(len(heatmap_data)):
    for j in range(3):
        value = heatmap_values[i, j]
        if j == 0:  # Spend
            text = f'${value:.1f}M'
        elif j == 1:  # Vendor count
            text = f'{int(value)}'
        else:  # Priority score
            text = f'{value:.1f}'
        ax.text(j, i, text, ha='center', va='center', fontsize=10, fontweight='bold')

ax.set_title('Category Management Priority Matrix\n(Red = High Priority for Consolidation)',
            fontsize=14, fontweight='bold', pad=20)

# Colorbar
cbar = plt.colorbar(im, ax=ax)
cbar.set_label('Priority Level', fontsize=11, fontweight='bold')

plt.tight_layout()
chart7_path = charts_dir / '07_category_priority_heatmap.png'
plt.savefig(chart7_path, dpi=300, bbox_inches='tight')
plt.close()

print(f"   All 7 charts saved to: {charts_dir}")

# ============================================================================
# SECTION 5: EXCEL WORKBOOK GENERATION
# ============================================================================
print("\n5. Creating Excel workbook with 8 professional sheets...")

wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Styling functions
border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_header_style():
    return {
        'font': Font(name='Calibri', size=11, bold=True, color=WHITE),
        'fill': PatternFill(start_color=CHICAGO_BLUE, end_color=CHICAGO_BLUE, fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': border_thin
    }

# Sheet 1: Executive Summary
print("   Building Sheet 1: Executive Summary...")
ws1 = wb.create_sheet("Executive Summary")
ws1['A1'] = "CITY OF CHICAGO"
ws1['A1'].font = Font(name='Calibri', size=24, bold=True, color=CHICAGO_BLUE)
ws1['A2'] = "Vendor Consolidation & Strategic Sourcing Analysis"
ws1['A2'].font = Font(name='Calibri', size=18, bold=True, color=CHICAGO_RED)
ws1['A3'] = f"Landscaping Services - Report Date: {datetime.now().strftime('%B %d, %Y')}"
ws1['A3'].font = Font(name='Calibri', size=12, italic=True)

row = 5
ws1[f'A{row}'] = "KEY FINDINGS"
ws1[f'A{row}'].font = Font(name='Calibri', size=14, bold=True, color=WHITE)
ws1[f'A{row}'].fill = PatternFill(start_color=CHICAGO_BLUE, end_color=CHICAGO_BLUE, fill_type='solid')
row += 2

findings = [
    f"• Current State: {total_vendors} vendors managing ${total_spend:,.0f} across {total_contracts} contracts",
    f"• Top 5 vendors control {vendor_spend.head(5)['% of Total'].sum():.1f}% of total spend",
    f"• {len(vendor_spend[vendor_spend['% of Total'] < 1])} small vendors (<1% each) represent consolidation opportunity",
    f"• Moderate consolidation scenario projects ${scenarios_list[1]['Total Annual Savings']:,.0f} annual savings",
    f"• 3-year strategic sourcing could avoid ${total_3yr_avoidance:,.0f} in inflation costs",
    f"• Total 3-year financial benefit: ${(scenarios_list[1]['Total Annual Savings'] * 3 + total_3yr_avoidance):,.0f}"
]

for finding in findings:
    ws1[f'A{row}'] = finding
    ws1[f'A{row}'].font = Font(name='Calibri', size=11)
    row += 1

row += 1
ws1[f'A{row}'] = "CONSOLIDATION SCENARIOS - ANNUAL SAVINGS"
ws1[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color=WHITE)
ws1[f'A{row}'].fill = PatternFill(start_color=CHICAGO_BLUE, end_color=CHICAGO_BLUE, fill_type='solid')
ws1.merge_cells(f'A{row}:E{row}')
row += 1

# Headers
headers = ['Scenario', 'Target Vendors', 'Reduction %', 'Volume Savings', 'Admin Savings', 'Total Annual Savings']
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=row, column=col, value=header)
    header_style = create_header_style()
    cell.font = header_style['font']
    cell.fill = header_style['fill']
    cell.alignment = header_style['alignment']
    cell.border = header_style['border']
row += 1

# Scenario data
for scenario in scenarios_list:
    ws1.cell(row=row, column=1, value=scenario['Scenario'])
    ws1.cell(row=row, column=2, value=scenario['Target Vendors'])
    ws1.cell(row=row, column=3, value=f"{scenario['Reduction %']}%")
    ws1.cell(row=row, column=4, value=scenario['Volume Savings']).number_format = accounting_format
    ws1.cell(row=row, column=5, value=scenario['Admin Savings']).number_format = accounting_format
    ws1.cell(row=row, column=6, value=scenario['Total Annual Savings']).number_format = accounting_format

    # Highlight recommended scenario
    if scenario['Scenario'] == 'Moderate':
        for col in range(1, 7):
            ws1.cell(row=row, column=col).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            ws1.cell(row=row, column=col).font = Font(name='Calibri', size=10, bold=True)

    for col in range(1, 7):
        ws1.cell(row=row, column=col).border = border_thin
    row += 1

row += 1
ws1[f'A{row}'] = "RECOMMENDATION"
ws1[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color=WHITE)
ws1[f'A{row}'].fill = PatternFill(start_color=CHICAGO_RED, end_color=CHICAGO_RED, fill_type='solid')
row += 1
ws1[f'A{row}'] = f"Implement MODERATE SCENARIO: Consolidate to {scenarios_list[1]['Target Vendors']} vendors for ${scenarios_list[1]['Total Annual Savings']:,.0f} annual savings"
ws1[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
ws1.merge_cells(f'A{row}:F{row}')

# Set column widths
ws1.column_dimensions['A'].width = 60
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 20
ws1.column_dimensions['E'].width = 20
ws1.column_dimensions['F'].width = 22

# Embed waterfall chart
row += 2
if chart3_path.exists():
    img = OpenpyxlImage(str(chart3_path))
    img.width = 800
    img.height = 500
    ws1.add_image(img, f'A{row}')

print("   ✓ Executive Summary complete")

# Sheet 2: Consolidation Scenarios
print("   Building Sheet 2: Consolidation Scenarios...")
ws2 = wb.create_sheet("Consolidation Scenarios")
ws2['A1'] = "CONSOLIDATION SCENARIOS - DETAILED COMPARISON"
ws2['A1'].font = Font(name='Calibri', size=16, bold=True, color=CHICAGO_BLUE)
ws2.merge_cells('A1:H1')

row = 3
headers = ['Scenario', 'Current Vendors', 'Target Vendors', 'Vendors Eliminated', 'Reduction %',
          'Volume Savings', 'Admin Savings', 'Total Annual Savings']
for col, header in enumerate(headers, 1):
    cell = ws2.cell(row=row, column=col, value=header)
    header_style = create_header_style()
    cell.font = header_style['font']
    cell.fill = header_style['fill']
    cell.alignment = header_style['alignment']
    cell.border = border_thin
row += 1

for scenario in scenarios_list:
    ws2.cell(row=row, column=1, value=scenario['Scenario'])
    ws2.cell(row=row, column=2, value=scenario['Current Vendors'])
    ws2.cell(row=row, column=3, value=scenario['Target Vendors'])
    ws2.cell(row=row, column=4, value=scenario['Vendors Eliminated'])
    ws2.cell(row=row, column=5, value=f"{scenario['Reduction %']}%")
    ws2.cell(row=row, column=6, value=scenario['Volume Savings']).number_format = accounting_format
    ws2.cell(row=row, column=7, value=scenario['Admin Savings']).number_format = accounting_format
    ws2.cell(row=row, column=8, value=scenario['Total Annual Savings']).number_format = accounting_format
    for col in range(1, 9):
        ws2.cell(row=row, column=col).border = border_thin
        ws2.cell(row=row, column=col).alignment = Alignment(horizontal='center' if col <= 5 else 'right')
    row += 1

for col in range(1, 9):
    ws2.column_dimensions[get_column_letter(col)].width = 18

# Embed charts
row += 2
if chart2_path.exists():
    img = OpenpyxlImage(str(chart2_path))
    img.width = 700
    img.height = 400
    ws2.add_image(img, f'A{row}')

print("   ✓ Consolidation Scenarios complete")

# Sheet 3: Cost Savings Calculations
print("   Building Sheet 3: Cost Savings Calculations...")
ws3 = wb.create_sheet("Cost Savings Calculations")
ws3['A1'] = "COST SAVINGS CALCULATION METHODOLOGY"
ws3['A1'].font = Font(name='Calibri', size=16, bold=True, color=CHICAGO_BLUE)

row = 3
ws3[f'A{row}'] = "VOLUME LEVERAGE SAVINGS"
ws3[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color=WHITE)
ws3[f'A{row}'].fill = PatternFill(start_color=CHICAGO_BLUE, end_color=CHICAGO_BLUE, fill_type='solid')
ws3.merge_cells(f'A{row}:D{row}')
row += 2

volume_data = [
    ['Scenario', 'Savings Rate', 'Calculation', 'Annual Volume Savings'],
    ['Conservative', '4%', f'${total_spend:,.0f} × 4%', scenarios_list[0]['Volume Savings']],
    ['Moderate', '8%', f'${total_spend:,.0f} × 8%', scenarios_list[1]['Volume Savings']],
    ['Aggressive', '13%', f'${total_spend:,.0f} × 13%', scenarios_list[2]['Volume Savings']]
]

for i, data_row in enumerate(volume_data):
    for col, value in enumerate(data_row, 1):
        cell = ws3.cell(row=row+i, column=col, value=value)
        if i == 0:
            cell.font = Font(bold=True, color=WHITE)
            cell.fill = PatternFill(start_color=CHICAGO_BLUE, end_color=CHICAGO_BLUE, fill_type='solid')
        elif col == 4 and i > 0:
            cell.number_format = accounting_format
        cell.border = border_thin

row += len(volume_data) + 2
ws3[f'A{row}'] = "ADMINISTRATIVE SAVINGS"
ws3[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color=WHITE)
ws3[f'A{row}'].fill = PatternFill(start_color=CHICAGO_BLUE, end_color=CHICAGO_BLUE, fill_type='solid')
ws3.merge_cells(f'A{row}:D{row}')
row += 2

admin_data = [
    ['Scenario', 'Vendors Eliminated', 'Cost per Vendor', 'Annual Admin Savings'],
    ['Conservative', scenarios_list[0]['Vendors Eliminated'], f'${ADMIN_COST_PER_VENDOR:,}', scenarios_list[0]['Admin Savings']],
    ['Moderate', scenarios_list[1]['Vendors Eliminated'], f'${ADMIN_COST_PER_VENDOR:,}', scenarios_list[1]['Admin Savings']],
    ['Aggressive', scenarios_list[2]['Vendors Eliminated'], f'${ADMIN_COST_PER_VENDOR:,}', scenarios_list[2]['Admin Savings']]
]

for i, data_row in enumerate(admin_data):
    for col, value in enumerate(data_row, 1):
        cell = ws3.cell(row=row+i, column=col, value=value)
        if i == 0:
            cell.font = Font(bold=True, color=WHITE)
            cell.fill = PatternFill(start_color=CHICAGO_BLUE, end_color=CHICAGO_BLUE, fill_type='solid')
        elif col == 4 and i > 0:
            cell.number_format = accounting_format
        cell.border = border_thin

for col in range(1, 5):
    ws3.column_dimensions[get_column_letter(col)].width = 25

print("   ✓ Cost Savings Calculations complete")

# Sheet 4: Cost Avoidance Analysis
print("   Building Sheet 4: Cost Avoidance Analysis...")
ws4 = wb.create_sheet("Cost Avoidance Analysis")
ws4['A1'] = "3-YEAR COST AVOIDANCE THROUGH STRATEGIC SOURCING"
ws4['A1'].font = Font(name='Calibri', size=16, bold=True, color=CHICAGO_BLUE)

row = 3
ws4[f'A{row}'] = f"Inflation Rate: {INFLATION_RATE*100:.1f}% annually (BLS Landscaping Services CPI)"
ws4[f'A{row}'].font = Font(name='Calibri', size=11, italic=True)
row += 2

headers = ['Year', 'Without Strategic Sourcing', 'With Strategic Sourcing (Rate Lock)', 'Annual Cost Avoidance', 'Cumulative Cost Avoidance']
for col, header in enumerate(headers, 1):
    cell = ws4.cell(row=row, column=col, value=header)
    header_style = create_header_style()
    cell.font = header_style['font']
    cell.fill = header_style['fill']
    cell.alignment = header_style['alignment']
    cell.border = border_thin
row += 1

cumulative_avoidance = 0
for i, year_data in cost_avoidance_df.iterrows():
    cumulative_avoidance += year_data['Cost Avoidance']
    ws4.cell(row=row, column=1, value=f"Year {year_data['Year']}")
    ws4.cell(row=row, column=2, value=year_data['Without Strategic Sourcing']).number_format = accounting_format
    ws4.cell(row=row, column=3, value=year_data['With Strategic Sourcing (Rate Lock)']).number_format = accounting_format
    ws4.cell(row=row, column=4, value=year_data['Cost Avoidance']).number_format = accounting_format
    ws4.cell(row=row, column=5, value=cumulative_avoidance).number_format = accounting_format
    for col in range(1, 6):
        ws4.cell(row=row, column=col).border = border_thin
    row += 1

# Total row
ws4.cell(row=row, column=1, value="TOTAL 3-YEAR AVOIDANCE").font = Font(bold=True)
ws4.cell(row=row, column=5, value=total_3yr_avoidance).number_format = accounting_format
ws4.cell(row=row, column=5).font = Font(bold=True)
ws4.cell(row=row, column=5).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
for col in range(1, 6):
    ws4.cell(row=row, column=col).border = border_thin

for col in range(1, 6):
    ws4.column_dimensions[get_column_letter(col)].width = 28

# Embed chart
row += 2
if chart6_path.exists():
    img = OpenpyxlImage(str(chart6_path))
    img.width = 700
    img.height = 400
    ws4.add_image(img, f'A{row}')

print("   ✓ Cost Avoidance Analysis complete")

# Sheet 5: Service Category Analysis
print("   Building Sheet 5: Service Category Analysis...")
ws5 = wb.create_sheet("Service Category Analysis")
ws5['A1'] = "SERVICE CATEGORY ANALYSIS"
ws5['A1'].font = Font(name='Calibri', size=16, bold=True, color=CHICAGO_BLUE)

row = 3
headers = ['Service Category', 'Total Spend', '% of Total', 'Vendor Count', 'Contract Count', 'Avg Contract Size']
for col, header in enumerate(headers, 1):
    cell = ws5.cell(row=row, column=col, value=header)
    header_style = create_header_style()
    cell.font = header_style['font']
    cell.fill = header_style['fill']
    cell.alignment = header_style['alignment']
    cell.border = border_thin
row += 1

for _, cat_data in service_spend.iterrows():
    ws5.cell(row=row, column=1, value=cat_data['Service Category'])
    ws5.cell(row=row, column=2, value=cat_data['Total Spend']).number_format = accounting_format
    ws5.cell(row=row, column=3, value=f"{cat_data['% of Total']:.2f}%")
    ws5.cell(row=row, column=4, value=cat_data['Vendor Count'])
    ws5.cell(row=row, column=5, value=cat_data['Contract Count'])
    ws5.cell(row=row, column=6, value=cat_data['Avg Contract Size']).number_format = accounting_format
    for col in range(1, 7):
        ws5.cell(row=row, column=col).border = border_thin
    row += 1

ws5.column_dimensions['A'].width = 35
for col in range(2, 7):
    ws5.column_dimensions[get_column_letter(col)].width = 18

# Embed charts
row += 2
if chart1_path.exists():
    img = OpenpyxlImage(str(chart1_path))
    img.width = 700
    img.height = 500
    ws5.add_image(img, f'A{row}')

print("   ✓ Service Category Analysis complete")

# Sheet 6: Vendor Performance Matrix
print("   Building Sheet 6: Vendor Performance Matrix...")
ws6 = wb.create_sheet("Vendor Performance Matrix")
ws6['A1'] = "TOP 20 VENDORS BY SPEND"
ws6['A1'].font = Font(name='Calibri', size=16, bold=True, color=CHICAGO_BLUE)

row = 3
headers = ['Rank', 'Vendor Name', 'Total Spend', '% of Total', 'Cumulative %', 'Contract Count', 'Service Categories']
for col, header in enumerate(headers, 1):
    cell = ws6.cell(row=row, column=col, value=header)
    header_style = create_header_style()
    cell.font = header_style['font']
    cell.fill = header_style['fill']
    cell.alignment = header_style['alignment']
    cell.border = border_thin
row += 1

for i, (_, vendor_data) in enumerate(vendor_spend.head(20).iterrows(), 1):
    ws6.cell(row=row, column=1, value=i)
    ws6.cell(row=row, column=2, value=vendor_data['Vendor'])
    ws6.cell(row=row, column=3, value=vendor_data['Total Spend']).number_format = accounting_format
    ws6.cell(row=row, column=4, value=f"{vendor_data['% of Total']:.2f}%")
    ws6.cell(row=row, column=5, value=f"{vendor_data['Cumulative %']:.2f}%")
    ws6.cell(row=row, column=6, value=vendor_data['Contract Count'])
    ws6.cell(row=row, column=7, value=vendor_data['Service Categories'])
    for col in range(1, 8):
        ws6.cell(row=row, column=col).border = border_thin
    row += 1

ws6.column_dimensions['A'].width = 8
ws6.column_dimensions['B'].width = 45
for col in range(3, 8):
    ws6.column_dimensions[get_column_letter(col)].width = 18

# Embed chart
row += 2
if chart4_path.exists():
    img = OpenpyxlImage(str(chart4_path))
    img.width = 800
    img.height = 500
    ws6.add_image(img, f'A{row}')

print("   ✓ Vendor Performance Matrix complete")

# Sheet 7: Implementation Roadmap
print("   Building Sheet 7: Implementation Roadmap...")
ws7 = wb.create_sheet("Implementation Roadmap")
ws7['A1'] = "18-MONTH IMPLEMENTATION ROADMAP"
ws7['A1'].font = Font(name='Calibri', size=16, bold=True, color=CHICAGO_BLUE)

row = 3
roadmap_data = [
    ['Phase', 'Timeline', 'Key Activities', 'Deliverables'],
    ['Phase 1: Strategy Development', 'Months 1-3',
     'Spend analysis validation, Market research, Vendor capability assessment, Consolidation strategy approval',
     'Category strategy document, Vendor assessment report, Executive approval'],
    ['Phase 2: RFP Process', 'Months 4-9',
     'RFP development and release, Vendor proposals and evaluation, Contract negotiations, Award decisions',
     'RFP documentation, Evaluation scorecards, Negotiated contracts'],
    ['Phase 3: Transition', 'Months 10-12',
     'Vendor onboarding, Service transition management, Performance baseline establishment, Training',
     'Transition plan, Performance baselines, Training materials'],
    ['Phase 4: Monitoring', 'Months 13-18',
     'Savings tracking, Service quality monitoring, Continuous improvement, Vendor performance reviews',
     'Monthly savings reports, Quality dashboards, Performance reviews']
]

for i, data_row in enumerate(roadmap_data):
    for col, value in enumerate(data_row, 1):
        cell = ws7.cell(row=row+i, column=col, value=value)
        if i == 0:
            cell.font = Font(bold=True, color=WHITE)
            cell.fill = PatternFill(start_color=CHICAGO_BLUE, end_color=CHICAGO_BLUE, fill_type='solid')
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

ws7.column_dimensions['A'].width = 25
ws7.column_dimensions['B'].width = 18
ws7.column_dimensions['C'].width = 50
ws7.column_dimensions['D'].width = 40

print("   ✓ Implementation Roadmap complete")

# Sheet 8: Assumptions & Methodology
print("   Building Sheet 8: Assumptions & Methodology...")
ws8 = wb.create_sheet("Assumptions & Methodology")
ws8['A1'] = "ASSUMPTIONS & METHODOLOGY"
ws8['A1'].font = Font(name='Calibri', size=16, bold=True, color=CHICAGO_BLUE)

row = 3
ws8[f'A{row}'] = "DATA SOURCES"
ws8[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color=WHITE)
ws8[f'A{row}'].fill = PatternFill(start_color=CHICAGO_BLUE, end_color=CHICAGO_BLUE, fill_type='solid')
ws8.merge_cells(f'A{row}:B{row}')
row += 2

data_sources = [
    ['Source File:', 'landscaping_structured.xlsx'],
    ['Total Contracts:', f'{total_contracts}'],
    ['Total Vendors:', f'{total_vendors}'],
    ['Total Spend:', f'${total_spend:,.2f}'],
    ['Date Range:', 'July 1993 - October 2025'],
    ['', '']
]

for item in data_sources:
    ws8[f'A{row}'] = item[0]
    ws8[f'A{row}'].font = Font(bold=True)
    ws8[f'B{row}'] = item[1]
    row += 1

row += 1
ws8[f'A{row}'] = "INDUSTRY BENCHMARKS"
ws8[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color=WHITE)
ws8[f'A{row}'].fill = PatternFill(start_color=CHICAGO_BLUE, end_color=CHICAGO_BLUE, fill_type='solid')
ws8.merge_cells(f'A{row}:B{row}')
row += 2

benchmarks = [
    ['Volume Leverage Savings:', 'NIGP, NASPO, ISM Research'],
    ['  Conservative (20-30% reduction):', '4% savings rate'],
    ['  Moderate (40-50% reduction):', '8% savings rate'],
    ['  Aggressive (60-70% reduction):', '13% savings rate'],
    ['Administrative Cost Savings:', f'${ADMIN_COST_PER_VENDOR:,} per eliminated vendor'],
    ['Inflation Rate:', f'{INFLATION_RATE*100:.1f}% annually (BLS Landscaping CPI)'],
    ['', '']
]

for item in benchmarks:
    ws8[f'A{row}'] = item[0]
    if item[0].startswith('  '):
        ws8[f'A{row}'].font = Font(italic=True)
    else:
        ws8[f'A{row}'].font = Font(bold=True)
    ws8[f'B{row}'] = item[1]
    row += 1

row += 1
ws8[f'A{row}'] = "LIMITATIONS"
ws8[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color=WHITE)
ws8[f'A{row}'].fill = PatternFill(start_color=CHICAGO_BLUE, end_color=CHICAGO_BLUE, fill_type='solid')
ws8.merge_cells(f'A{row}:B{row}')
row += 2

limitations = [
    '• Vendor performance data not assessed',
    '• Transition costs not included in calculations',
    '• Geographic coverage requirements not analyzed',
    '• Service quality metrics not available',
    '• Vendor capacity constraints not evaluated'
]

for limitation in limitations:
    ws8[f'A{row}'] = limitation
    ws8.merge_cells(f'A{row}:B{row}')
    row += 1

ws8.column_dimensions['A'].width = 40
ws8.column_dimensions['B'].width = 50

print("   ✓ Assumptions & Methodology complete")

# Save workbook
print(f"\n6. Saving workbook to: {output_file}")
wb.save(output_file)

print("\n" + "="*100)
print("VENDOR CONSOLIDATION & STRATEGIC SOURCING ANALYSIS COMPLETE")
print("="*100)
print(f"\nOutput File: {output_file}")
print(f"\nAnalysis Summary:")
print(f"  Current State:")
print(f"    • Total Spend: ${total_spend:,.2f}")
print(f"    • Vendors: {total_vendors}")
print(f"    • Contracts: {total_contracts}")
print(f"\n  Consolidation Savings (Annual):")
print(f"    • Conservative: ${scenarios_list[0]['Total Annual Savings']:,.0f}")
print(f"    • Moderate (RECOMMENDED): ${scenarios_list[1]['Total Annual Savings']:,.0f}")
print(f"    • Aggressive: ${scenarios_list[2]['Total Annual Savings']:,.0f}")
print(f"\n  Cost Avoidance (3-Year):")
print(f"    • Strategic Sourcing: ${total_3yr_avoidance:,.0f}")
print(f"\n  Total 3-Year Benefit (Moderate + Avoidance):")
print(f"    • ${(scenarios_list[1]['Total Annual Savings'] * 3 + total_3yr_avoidance):,.0f}")
print("\n" + "="*100)
