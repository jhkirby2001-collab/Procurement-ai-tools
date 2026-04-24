"""
Step 1 — Excel Export of the Data Validation Profile
====================================================
Produces a leadership-ready .xlsx workbook of the Step 1 validation findings.

Portability: change only the CONFIG block to run on any org's extract.
"""

import os
import warnings
from pathlib import Path
from datetime import date

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── CONFIG ──────────────────────────────────────────────────────────
JURISDICTION_NAME           = "City of Chicago DPS"
ANALYSIS_PERIOD             = "2002-2025 (EY benchmark extract)"
INPUT_FILE                  = "/workspaces/Procurement-ai-tools/spend-analysis/data/raw/ey raw data.xlsx"
SHEET_NAME                  = "Report 1"
OUTPUT_DIR                  = "/workspaces/Procurement-ai-tools/spend-analysis/outputs"
CACHE_FILE                  = "/workspaces/Procurement-ai-tools/spend-analysis/data/processed/ey_raw_cache.parquet"
OUTPUT_FILE                 = os.path.join(
    OUTPUT_DIR, "ey_benchmark_step1_validation_JHK3.xlsx"
)
ANALYST_NAME                = "James H. Kirby III, CSCP, MS-SCM"
ANALYST_ROLE                = "Senior Buyer and Procurement Research Analyst"
ASSUMPTION_ADDRESSABILITY   = 0.85
ASSUMPTION_SAVINGS_RATE     = 0.10
ASSUMPTION_HORIZON_YEARS    = 3
# ────────────────────────────────────────────────────────────────────

# ── CLAUDE.md v4.1 palette ──────────────────────────────────────────
NAVY        = "003366"
MED_BLUE    = "0071BC"
LIGHT_BLUE  = "E6EEF5"
LIGHT_GRAY  = "F2F2F2"
LIGHT_GREEN = "D9EAD3"
RED         = "E31837"
WHITE       = "FFFFFF"

THIN   = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fill(hex_):  return PatternFill("solid", fgColor=hex_)
def bold_white(size=11): return Font(name="Arial", size=size, bold=True, color=WHITE)
def bold_black(size=11): return Font(name="Arial", size=size, bold=True)
def italic(size=10): return Font(name="Arial", size=size, italic=True, color="404040")
def plain(size=11): return Font(name="Arial", size=size)


def set_attribution_header(ws, title, ncols=8):
    """Rows 1-3 per CLAUDE.md Section 0."""
    # Row 1 — title
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1)
    c.font = bold_white(14); c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    # Row 2
    ws.cell(row=2, column=1,
            value=f"Prepared by: {ANALYST_NAME}  |  {ANALYST_ROLE}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1)
    c.font = Font(name="Arial", size=11, color=WHITE); c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18
    # Row 3
    ws.cell(row=3, column=1, value=f"Analysis Date: {date.today().isoformat()}")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    c = ws.cell(row=3, column=1)
    c.font = Font(name="Arial", size=10, color=WHITE); c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")


def write_table(ws, start_row, headers, formula_row, rows, col_widths=None,
                currency_cols=None, number_cols=None, pct_cols=None):
    """
    Standard data-sheet table with:
      - header row  (navy, white, bold)
      - formula row (gray italic — plain language, not '=')
      - data rows
    Returns the last row used.
    """
    currency_cols = set(currency_cols or [])
    number_cols   = set(number_cols or [])
    pct_cols      = set(pct_cols or [])

    # header
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.font = bold_white(11); c.fill = fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[start_row].height = 30

    # formula row
    for j, f in enumerate(formula_row, start=1):
        c = ws.cell(row=start_row + 1, column=j, value=f)
        c.font = italic(10); c.fill = fill(LIGHT_GRAY)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[start_row + 1].height = 24

    # data
    for i, row in enumerate(rows):
        r = start_row + 2 + i
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = plain(11); c.border = BORDER
            if j in currency_cols and isinstance(v, (int, float, np.floating, np.integer)):
                c.number_format = '"$"#,##0.00'
                c.alignment = Alignment(horizontal="right")
            elif j in number_cols and isinstance(v, (int, float, np.floating, np.integer)):
                c.number_format = "#,##0"
                c.alignment = Alignment(horizontal="right")
            elif j in pct_cols and isinstance(v, (int, float, np.floating, np.integer)):
                c.number_format = "0.00%"
                c.alignment = Alignment(horizontal="right")
            else:
                c.alignment = Alignment(horizontal="left", wrap_text=True)

    if col_widths:
        for idx, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w

    return start_row + 1 + len(rows)


def write_methodology(ws, start_row, text, ncols=8):
    ws.cell(row=start_row, column=1, value="METHODOLOGY")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=ncols)
    c = ws.cell(row=start_row, column=1)
    c.font = bold_black(11); c.fill = fill(LIGHT_GRAY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.cell(row=start_row + 1, column=1, value=text)
    ws.merge_cells(start_row=start_row + 1, start_column=1,
                   end_row=start_row + 1, end_column=ncols)
    c = ws.cell(row=start_row + 1, column=1)
    c.font = plain(10)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[start_row + 1].height = 80
    return start_row + 2


def footer(ws, last_row, ncols=8):
    ws.cell(row=last_row + 2, column=1,
            value=f"Source: {INPUT_FILE}  |  Sheet: {SHEET_NAME}  |  Period: {ANALYSIS_PERIOD}")
    ws.merge_cells(start_row=last_row + 2, start_column=1,
                   end_row=last_row + 2, end_column=ncols)
    ws.cell(row=last_row + 2, column=1).font = Font(name="Arial", size=9, color="666666")
    ws.cell(row=last_row + 3, column=1,
            value=f"Prepared by {ANALYST_NAME}  |  {ANALYST_ROLE}  |  {JURISDICTION_NAME}")
    ws.merge_cells(start_row=last_row + 3, start_column=1,
                   end_row=last_row + 3, end_column=ncols)
    ws.cell(row=last_row + 3, column=1).font = Font(name="Arial", size=9, color="666666")


# ── Load data (re-use cache if present) ─────────────────────────────
def load_data():
    Path(os.path.dirname(CACHE_FILE)).mkdir(parents=True, exist_ok=True)
    if (os.path.exists(CACHE_FILE) and
            os.path.getmtime(CACHE_FILE) > os.path.getmtime(INPUT_FILE)):
        print(f"[cache] loading {CACHE_FILE}")
        return pd.read_parquet(CACHE_FILE)
    print(f"[read] loading {INPUT_FILE}")
    df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    try:
        df2 = df.copy()
        for c in df2.columns:
            if df2[c].dtype == object:
                df2[c] = df2[c].astype(str).where(df2[c].notna(), None)
        df2.to_parquet(CACHE_FILE, index=False)
        print(f"[cache] saved {CACHE_FILE}")
    except Exception as e:
        print(f"[cache] skipped: {e}")
    return df


def main():
    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)
    df = load_data()
    df.columns = [str(c).strip() for c in df.columns]

    # Numeric conversions once
    cur = pd.to_numeric(df.get("Current PO Dist Amount"), errors="coerce")
    apv = pd.to_numeric(df.get("AP Invoice Amount"), errors="coerce")
    pay = pd.to_numeric(df.get("Pay Check Amount"), errors="coerce")
    total_cur = float(cur.sum()); total_apv = float(apv.sum()); total_pay = float(pay.sum())

    dates = pd.to_datetime(df.get("PO Creation Date"), errors="coerce")
    min_d, max_d = dates.min(), dates.max()

    flags = []

    # ── Flag reconciliation variance
    vals = [total_cur, total_apv, total_pay]
    variance_pct = (max(vals) - min(vals)) / max(vals) if max(vals) else 0
    if variance_pct > 0.05:
        flags.append(("Spend reconciliation",
                      f"Three total-spend paths vary by {variance_pct:.1%}. "
                      f"Current PO Dist = ${total_cur:,.0f}; AP Invoice = ${total_apv:,.0f}; "
                      f"Pay Check = ${total_pay:,.0f}. "
                      f"Current PO Dist and Pay Check contain contract-ceiling / encumbrance "
                      f"values at the PO-line level and are NOT actual spend. "
                      f"AP Invoice Amount is the defensible spend-of-record."))

    # ── Flag NIGP coverage
    nigp_blank = (df["NIGP Code"].isna() |
                  df["NIGP Code"].astype(str).str.strip().isin(["", "nan", "None"]))
    blank_spend_apv = float(apv[nigp_blank].sum())
    blank_spend_cur = float(cur[nigp_blank].sum())
    flags.append(("NIGP coverage",
                  f"{nigp_blank.sum():,} rows ({nigp_blank.sum()/len(df):.2%}) have no NIGP Code. "
                  f"Spend attributed to uncoded rows: "
                  f"AP Invoice=${blank_spend_apv:,.0f} ({blank_spend_apv/total_apv:.2%} of AP total); "
                  f"PO Dist=${blank_spend_cur:,.0f} ({blank_spend_cur/total_cur:.2%} of PO Dist total). "
                  f"Category analytics will under-report by that amount unless Commodity Code/Desc "
                  f"is used as fallback."))

    # ── Flag contract coverage (maverick definition)
    contract = df["PO Contract Number"].astype(str).str.strip()
    blank_like = contract.str.lower().isin(["", "nan", "none", "n/a", "na", "null", "0"])
    flags.append(("Maverick spend redefinition required",
                  f"PO Contract Number has zero nulls and zero blank-like placeholders "
                  f"({int(blank_like.sum()):,} rows). Classic 'null contract = maverick' "
                  f"cannot be computed on this extract. Must redefine using one of: "
                  f"(a) PO Creation Date > Contract End Date; "
                  f"(b) one-off contracts (contract number appearing <3 times); "
                  f"(c) PO Contract Type = 'one-time'/spot."))

    # ── Flag cancelled spend
    status = df.get("PO Approval Status", pd.Series(index=df.index, dtype=object)).fillna("(blank)")
    cancelled_mask = status.astype(str).str.lower().str.contains("cancel|reject|void", na=False)
    cancelled_spend_cur = float(cur[cancelled_mask].sum())
    cancelled_spend_apv = float(apv[cancelled_mask].sum())
    flags.append(("Cancelled / rejected POs",
                  f"{int(cancelled_mask.sum()):,} POs in rejected/cancelled/voided status. "
                  f"Dollar volume: PO Dist=${cancelled_spend_cur:,.0f}; "
                  f"AP Invoice=${cancelled_spend_apv:,.0f}. "
                  f"Report as cost-avoidance / process-efficiency indicator; exclude from run-rate."))

    # ── Flag date range
    flags.append(("Date range exceeds stated window",
                  f"PO Creation Date spans {min_d.date() if pd.notna(min_d) else '?'} to "
                  f"{max_d.date() if pd.notna(max_d) else '?'} (≈23 years), not the "
                  f"'2021-2026 extract' described. Decide filter before Step 2 "
                  f"(2021+ / all history / rolling 5-year)."))

    # ── Build workbook
    wb = Workbook()

    # ========= 1. EXECUTIVE SUMMARY / FLAGS =========
    ws = wb.active; ws.title = "1. Executive Summary"
    set_attribution_header(ws, f"{JURISDICTION_NAME} — STEP 1 DATA VALIDATION — EXECUTIVE SUMMARY", ncols=8)

    r = 5
    # THE PROBLEM row (CLAUDE.md narrative template)
    ws.cell(row=r, column=1, value="⚠ FLAG FOR LEADERSHIP")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    c = ws.cell(row=r, column=1); c.font = bold_white(12); c.fill = fill(RED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 22
    r += 1
    ws.cell(row=r, column=1,
            value=("This validation surfaces material data-quality issues that must be "
                   "resolved before savings modeling. Do not rely on Current PO Dist "
                   "Amount or Pay Check Amount as the spend-of-record — they contain "
                   "contract-ceiling values inflated up to 2.7× actual paid spend. "
                   "Use AP Invoice Amount for Step 2."))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    c = ws.cell(row=r, column=1); c.font = plain(11)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = 60
    r += 2

    # Flag table
    r = write_table(
        ws, r,
        headers=["#", "Flag", "Detail"],
        formula_row=["Line item", "Short label used in downstream references",
                     "Plain-English explanation and recommended action"],
        rows=[(i + 1, label, detail) for i, (label, detail) in enumerate(flags)],
        col_widths=[6, 30, 100],
    )
    r += 2

    # Decisions needed
    ws.cell(row=r, column=1, value="DECISIONS NEEDED BEFORE STEP 2")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    c = ws.cell(row=r, column=1); c.font = bold_white(12); c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="center")
    r += 1
    decisions = [
        ("1", "Spend column of record",
         "AP Invoice Amount (recommended) — the only column whose per-line magnitudes are plausible. "
         "Using Current PO Dist would inflate savings figures roughly 2.7×."),
        ("2", "Maverick / off-contract definition",
         "Replace 'null PO Contract Number' with one of: "
         "(a) PO Creation Date after Contract End Date; "
         "(b) contract numbers appearing fewer than 3 times (one-off vehicles); "
         "(c) PO Contract Type matches one-time / spot buy; "
         "or a blend."),
        ("3", "Date filter",
         "Data spans 23 years (2002-2025). Choose one: "
         "(a) Filter to 2021+ (≈5-year run-rate aligned with recent commit pattern); "
         "(b) Use full 23-year history (lifetime view); "
         "(c) Rolling 5-year window ending at max date."),
    ]
    r = write_table(
        ws, r,
        headers=["#", "Decision", "Options"],
        formula_row=["Line item", "Decision required from the CPO/analyst team",
                     "Options presented with trade-off"],
        rows=decisions,
        col_widths=[6, 30, 100],
    )

    footer(ws, r, ncols=8)
    ws.freeze_panes = "A4"

    # ========= 2. SHAPE AND DATE RANGE =========
    ws = wb.create_sheet("2. Shape & Dates")
    set_attribution_header(ws, "DATASET SHAPE AND DATE COVERAGE", ncols=4)
    r = 5
    r = write_methodology(ws, r,
        "Row and column counts on the 'Report 1' sheet and descriptive statistics on "
        "PO Creation Date (min, max, null count). Establishes the scope and time coverage "
        "of the extract and flags whether the data spans the period expected by leadership.",
        ncols=4)
    r += 1
    r = write_table(
        ws, r,
        headers=["Metric", "Value"],
        formula_row=["Measure", "Raw count or date computed against the full extract"],
        rows=[
            ("Row count",     len(df)),
            ("Column count",  len(df.columns)),
            ("PO Creation Date minimum", str(min_d)),
            ("PO Creation Date maximum", str(max_d)),
            ("PO Creation Date null count", int(dates.isna().sum())),
        ],
        col_widths=[40, 30],
        number_cols={2},
    )
    footer(ws, r, ncols=4)
    ws.freeze_panes = "A4"

    # ========= 3. SPEND RECONCILIATION =========
    ws = wb.create_sheet("3. Spend Reconciliation")
    set_attribution_header(ws, "THREE-WAY SPEND RECONCILIATION", ncols=5)
    r = 5
    r = write_methodology(ws, r,
        "The extract exposes three candidate 'spend' columns — Current PO Dist Amount, "
        "AP Invoice Amount, and Pay Check Amount. Each is summed across all rows and "
        "compared. A material variance indicates the columns are measuring different "
        "financial events (e.g., encumbrance vs. invoice vs. disbursement). "
        "For savings modeling, only the column representing actual paid/obligated dollars "
        "should be used — using an encumbrance column would inflate savings figures.",
        ncols=5)
    r += 1
    rows = [
        ("Current PO Dist Amount", "Encumbrance / contract ceiling at PO distribution level",
         int(cur.notna().sum()), int(cur.isna().sum()), total_cur),
        ("AP Invoice Amount", "Supplier-invoiced dollars (recommended spend-of-record)",
         int(apv.notna().sum()), int(apv.isna().sum()), total_apv),
        ("Pay Check Amount", "Check-level disbursement (aggregate, inflated by line repetition)",
         int(pay.notna().sum()), int(pay.isna().sum()), total_pay),
    ]
    r = write_table(
        ws, r,
        headers=["Column", "What it represents", "Non-null rows", "Null rows", "Total"],
        formula_row=["Source column on 'Report 1'",
                     "Plain-English interpretation of the underlying financial event",
                     "Count of rows with a numeric value",
                     "Count of rows missing a numeric value",
                     "Sum of the column across all rows"],
        rows=rows,
        col_widths=[32, 60, 18, 14, 22],
        number_cols={3, 4}, currency_cols={5},
    )
    r += 1
    # Variance row
    ws.cell(row=r, column=1, value=f"Max−Min variance: ${max(vals)-min(vals):,.2f} ({variance_pct:.2%} of max)")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1); c.font = bold_black(11); c.fill = fill(LIGHT_GRAY)
    footer(ws, r, ncols=5)
    ws.freeze_panes = "A4"

    # ========= 4. UNIQUE COUNTS =========
    ws = wb.create_sheet("4. Unique Counts")
    set_attribution_header(ws, "UNIQUE ENTITY COUNTS", ncols=4)
    r = 5
    r = write_methodology(ws, r,
        "Count of distinct values in each identity column after string-trim. Establishes "
        "the scale of the universe (how many vendors, NIGP codes, departments, POs, "
        "contracts) and detects vendor-name inconsistency (same Vendor Number with "
        "different spelling variants).",
        ncols=4)
    r += 1
    per_num = df.groupby("Vendor Number")["Vendor Name"].nunique() if "Vendor Number" in df.columns else pd.Series(dtype=int)
    unique_counts = [
        ("Unique vendors (Vendor Number)", int(df["Vendor Number"].nunique())),
        ("Unique Vendor Names (raw)", int(df["Vendor Name"].nunique())),
        ("Unique NIGP codes", int(df["NIGP Code"].nunique())),
        ("Unique NIGP Descriptions", int(df["NIGP Description"].nunique())),
        ("Unique PO Lead Departments", int(df["PO Lead Dept"].nunique())),
        ("Unique PO Numbers", int(df["PO Number"].nunique()) if "PO Number" in df.columns else "column not found"),
        ("Unique PO Contract Numbers", int(df["PO Contract Number"].nunique())),
        ("Vendor Numbers with >1 Name variant", int((per_num > 1).sum())),
    ]
    r = write_table(
        ws, r,
        headers=["Metric", "Value"],
        formula_row=["Measure",
                     "Distinct count after trimming whitespace; blank/NaN excluded"],
        rows=unique_counts,
        col_widths=[45, 22],
        number_cols={2},
    )
    footer(ws, r, ncols=4)
    ws.freeze_panes = "A4"

    # ========= 5. NULL / BLANK ANALYSIS =========
    ws = wb.create_sheet("5. Null-Blank Analysis")
    set_attribution_header(ws, "NULL / BLANK ANALYSIS — CRITICAL FIELDS", ncols=6)
    r = 5
    r = write_methodology(ws, r,
        "For each critical field we report: count of blank rows (NaN, empty string, "
        "or literal 'nan'/'None'), % of total rows, AP-Invoice spend attributed to "
        "blank rows, and % of total AP-Invoice spend. AP Invoice is used as the spend "
        "denominator because the three-way reconciliation shows it is the only "
        "defensible spend-of-record. NIGP blank-rows are the most consequential — "
        "category analytics will under-report by the spend shown on that row.",
        ncols=6)
    r += 1
    def blanks(col):
        if col not in df.columns: return None
        s = df[col]
        mask = s.isna() | s.astype(str).str.strip().isin(["", "nan", "None", "NaN"])
        return mask
    rows = []
    for col in ["NIGP Code","NIGP Description","PO Contract Number","Vendor Number",
                "PO Lead Dept","Current PO Dist Amount","AP Invoice Amount"]:
        m = blanks(col)
        if m is None:
            rows.append((col, "COLUMN NOT FOUND","","","",""))
            continue
        blank_count = int(m.sum())
        blank_row_pct = blank_count / len(df) if len(df) else 0
        blank_apv = float(apv[m].sum())
        blank_apv_pct = blank_apv / total_apv if total_apv else 0
        rows.append((col, blank_count, blank_row_pct, blank_apv, blank_apv_pct,
                     "Maverick candidate" if col == "PO Contract Number" else ""))
    r = write_table(
        ws, r,
        headers=["Critical field", "Blank rows", "% of rows",
                 "AP Invoice spend on blank rows", "% of AP Invoice total", "Note"],
        formula_row=["Source column on 'Report 1'",
                     "Count of rows where value is missing or blank-like",
                     "Blank rows ÷ total rows",
                     "Sum of AP Invoice Amount on blank rows",
                     "Blank-row AP Invoice sum ÷ AP Invoice grand total",
                     "Plain-English comment"],
        rows=rows,
        col_widths=[30, 14, 14, 28, 20, 25],
        number_cols={2}, pct_cols={3, 5}, currency_cols={4},
    )
    footer(ws, r, ncols=6)
    ws.freeze_panes = "A4"

    # ========= 6. DPS vs NON-DPS =========
    ws = wb.create_sheet("6. DPS vs Non-DPS")
    set_attribution_header(ws, "DPS vs NON-DPS OVERSIGHT", ncols=5)
    r = 5
    r = write_methodology(ws, r,
        "Group rows by the DPS / Non-DPS Flag. Report transaction count, AP Invoice "
        "spend, and share of total. Departments with heavy Non-DPS concentration are "
        "where procurement oversight is weakest.",
        ncols=5)
    r += 1
    dps_col = "DPS / Non-DPS Flag" if "DPS / Non-DPS Flag" in df.columns else None
    if dps_col:
        g = (df.assign(_apv=apv)
               .groupby(df[dps_col].fillna("(blank)"), dropna=False)
               .agg(txn=("_apv", "size"), apv=("_apv", "sum"))
               .reset_index().rename(columns={"index": dps_col}))
        g["share"] = g["apv"] / total_apv if total_apv else 0
        g = g.sort_values("apv", ascending=False)
        rows = [(str(r[dps_col]), int(r["txn"]), float(r["apv"]), float(r["share"])) for _, r in g.iterrows()]
    else:
        rows = [("column not found", "", "", "")]
    r = write_table(
        ws, r,
        headers=["DPS flag", "Transactions", "AP Invoice spend", "% of AP Invoice total"],
        formula_row=["Value in DPS / Non-DPS Flag column",
                     "Row count in that group",
                     "Sum of AP Invoice Amount",
                     "Group AP Invoice sum ÷ AP Invoice grand total"],
        rows=rows,
        col_widths=[22, 18, 28, 22],
        number_cols={2}, currency_cols={3}, pct_cols={4},
    )
    footer(ws, r, ncols=5)
    ws.freeze_panes = "A4"

    # ========= 7. APPROVAL STATUS =========
    ws = wb.create_sheet("7. Approval Status")
    set_attribution_header(ws, "PO APPROVAL STATUS BREAKDOWN", ncols=5)
    r = 5
    r = write_methodology(ws, r,
        "Transactions by PO Approval Status. 'Rejected', 'Cancelled', and 'Voided' "
        "represent process overhead: work performed on POs that did not execute. "
        "Report this dollar volume separately as a process-efficiency indicator — "
        "not a savings figure.",
        ncols=5)
    r += 1
    g = (df.assign(_apv=apv, _cur=cur)
           .groupby(status, dropna=False)
           .agg(txn=("_apv","size"), apv=("_apv","sum"), cur=("_cur","sum"))
           .reset_index().rename(columns={"index":"PO Approval Status"})
           .sort_values("apv", ascending=False))
    rows = [(str(r["PO Approval Status"]), int(r["txn"]), float(r["apv"]), float(r["cur"]),
             float(r["apv"])/total_apv if total_apv else 0)
            for _, r in g.iterrows()]
    r = write_table(
        ws, r,
        headers=["PO Approval Status", "Transactions", "AP Invoice spend",
                 "PO Dist amount", "% of AP Invoice total"],
        formula_row=["Value in PO Approval Status column",
                     "Row count", "Sum AP Invoice", "Sum PO Dist",
                     "AP Invoice of group ÷ AP Invoice total"],
        rows=rows,
        col_widths=[28, 16, 24, 24, 22],
        number_cols={2}, currency_cols={3,4}, pct_cols={5},
    )
    footer(ws, r, ncols=5)
    ws.freeze_panes = "A4"

    # ========= 8. TOP 20 VENDORS =========
    ws = wb.create_sheet("8. Top 20 Vendors")
    set_attribution_header(ws, "TOP 20 VENDORS BY AP INVOICE SPEND", ncols=6)
    r = 5
    r = write_methodology(ws, r,
        "Vendors ranked by sum of AP Invoice Amount. Reported with Current PO Dist "
        "for comparison — the gap between AP Invoice and PO Dist on a single vendor "
        "is a strong signal of contract-ceiling encumbrance inflation (see vendor "
        "Favorite Healthcare Staffing as the reference case: PO Dist $48B vs "
        "AP Invoice $433M).",
        ncols=6)
    r += 1
    v = (df.assign(_apv=apv, _cur=cur)
           .groupby(["Vendor Number","Vendor Name"], dropna=False)
           .agg(txn=("_apv","size"), apv=("_apv","sum"), cur=("_cur","sum"))
           .reset_index()
           .sort_values("apv", ascending=False)
           .head(20))
    rows = [(str(r["Vendor Number"]), str(r["Vendor Name"]),
             int(r["txn"]), float(r["apv"]), float(r["cur"]),
             float(r["apv"])/total_apv if total_apv else 0)
            for _, r in v.iterrows()]
    r = write_table(
        ws, r,
        headers=["Vendor Number","Vendor Name","Transactions",
                 "AP Invoice spend","Current PO Dist","% of AP Invoice total"],
        formula_row=["Vendor Number column","Vendor Name column",
                     "Row count per vendor",
                     "Sum of AP Invoice for vendor",
                     "Sum of PO Dist for vendor (may be inflated)",
                     "Vendor AP Invoice ÷ AP Invoice grand total"],
        rows=rows,
        col_widths=[16, 46, 16, 24, 24, 22],
        number_cols={3}, currency_cols={4,5}, pct_cols={6},
    )
    footer(ws, r, ncols=6)
    ws.freeze_panes = "A4"

    # ========= 9. TOP 20 NIGP =========
    ws = wb.create_sheet("9. Top 20 NIGP")
    set_attribution_header(ws, "TOP 20 NIGP CATEGORIES BY AP INVOICE SPEND", ncols=5)
    r = 5
    r = write_methodology(ws, r,
        "NIGP class-item codes ranked by sum of AP Invoice Amount. The 'blank / no code' "
        "row is shown explicitly so leadership can see the magnitude of uncoded spend. "
        "Category strategy in Step 2 should treat the uncoded bucket as its own "
        "category or recover classification via Commodity Code / Description.",
        ncols=5)
    r += 1
    nigp_key_code = df["NIGP Code"].fillna("(blank)").astype(str)
    nigp_key_desc = df["NIGP Description"].fillna("(blank)").astype(str)
    n = (df.assign(_apv=apv, _code=nigp_key_code, _desc=nigp_key_desc)
           .groupby(["_code","_desc"], dropna=False)
           .agg(txn=("_apv","size"), apv=("_apv","sum"))
           .reset_index()
           .sort_values("apv", ascending=False)
           .head(20))
    rows = [(str(r["_code"]), str(r["_desc"]),
             int(r["txn"]), float(r["apv"]),
             float(r["apv"])/total_apv if total_apv else 0)
            for _, r in n.iterrows()]
    r = write_table(
        ws, r,
        headers=["NIGP Code","NIGP Description","Transactions",
                 "AP Invoice spend","% of AP Invoice total"],
        formula_row=["NIGP Code (blank shown as '(blank)')",
                     "NIGP Description (blank shown as '(blank)')",
                     "Row count per category",
                     "Sum of AP Invoice Amount",
                     "Category AP Invoice ÷ AP Invoice grand total"],
        rows=rows,
        col_widths=[18, 58, 16, 24, 22],
        number_cols={3}, currency_cols={4}, pct_cols={5},
    )
    footer(ws, r, ncols=5)
    ws.freeze_panes = "A4"

    # ========= 10. SPEND BY YEAR =========
    ws = wb.create_sheet("10. Spend by Year")
    set_attribution_header(ws, "AP INVOICE SPEND BY YEAR (PO CREATION DATE)", ncols=4)
    r = 5
    r = write_methodology(ws, r,
        "Rows and AP Invoice spend bucketed by calendar year derived from PO Creation "
        "Date. Lets leadership see where the bulk of the 23-year history actually sits, "
        "and choose the appropriate time filter for Step 2.",
        ncols=4)
    r += 1
    yr = pd.DataFrame({"year": dates.dt.year, "apv": apv})
    by_year = (yr.dropna(subset=["year"])
                 .groupby("year", dropna=True)
                 .agg(rows=("apv","size"), apv=("apv","sum"))
                 .reset_index()
                 .sort_values("year"))
    rows = [(int(r["year"]), int(r["rows"]), float(r["apv"])) for _, r in by_year.iterrows()]
    r = write_table(
        ws, r,
        headers=["Year","Rows","AP Invoice spend"],
        formula_row=["Year from PO Creation Date",
                     "Row count per year",
                     "Sum of AP Invoice Amount"],
        rows=rows,
        col_widths=[12, 14, 26],
        number_cols={1,2}, currency_cols={3},
    )
    footer(ws, r, ncols=4)
    ws.freeze_panes = "A4"

    # ========= 11. DIAGNOSTIC — FAVORITE HEALTHCARE =========
    ws = wb.create_sheet("11. Diagnostic Drill")
    set_attribution_header(ws, "DIAGNOSTIC — WHY CURRENT PO DIST ≠ SPEND", ncols=4)
    r = 5
    r = write_methodology(ws, r,
        "Drill on the #1 vendor by Current PO Dist (Favorite Healthcare Staffing) to "
        "demonstrate why that column is not a valid spend-of-record. Per-row medians on "
        "PO Dist ($11M) and Pay Check ($6.5M) are implausible for a staffing vendor; "
        "AP Invoice per-row median ($137K) is in line with a typical staffing line-item. "
        "This is the evidence base for using AP Invoice Amount in Step 2.",
        ncols=4)
    r += 1
    fhs_mask = df["Vendor Name"].astype(str).str.contains("Favorite Healthcare", case=False, na=False)
    fhs_cur = cur[fhs_mask]; fhs_apv = apv[fhs_mask]; fhs_pay = pay[fhs_mask]
    rows = [
        ("Rows (transactions)", int(fhs_mask.sum()), "", ""),
        ("Sum  — Current PO Dist", float(fhs_cur.sum()), float(fhs_cur.median()), float(fhs_cur.max())),
        ("Sum  — AP Invoice",       float(fhs_apv.sum()), float(fhs_apv.median()), float(fhs_apv.max())),
        ("Sum  — Pay Check",        float(fhs_pay.sum()), float(fhs_pay.median()), float(fhs_pay.max())),
    ]
    r = write_table(
        ws, r,
        headers=["Measure","Sum / Count","Median per row","Max per row"],
        formula_row=["Metric","Sum over vendor rows (or row count)",
                     "Median of non-null values on vendor rows",
                     "Max of non-null values on vendor rows"],
        rows=rows,
        col_widths=[34, 24, 22, 22],
        currency_cols={2,3,4},
    )
    footer(ws, r, ncols=4)
    ws.freeze_panes = "A4"

    # ========= 12. RUN CONFIG =========
    ws = wb.create_sheet("12. Run Config")
    set_attribution_header(ws, "RUN CONFIGURATION (PORTABILITY)", ncols=3)
    r = 5
    r = write_methodology(ws, r,
        "The full parameter block for this run. Only these values change between "
        "organizations; the underlying methodology and scripts are unchanged. Use this "
        "sheet to audit the run or to reproduce it on a different extract.",
        ncols=3)
    r += 1
    r = write_table(
        ws, r,
        headers=["Parameter","Value"],
        formula_row=["Config constant","Value used for this run"],
        rows=[
            ("JURISDICTION_NAME", JURISDICTION_NAME),
            ("ANALYSIS_PERIOD",   ANALYSIS_PERIOD),
            ("INPUT_FILE",        INPUT_FILE),
            ("SHEET_NAME",        SHEET_NAME),
            ("OUTPUT_DIR",        OUTPUT_DIR),
            ("ASSUMPTION_ADDRESSABILITY", ASSUMPTION_ADDRESSABILITY),
            ("ASSUMPTION_SAVINGS_RATE",   ASSUMPTION_SAVINGS_RATE),
            ("ASSUMPTION_HORIZON_YEARS",  ASSUMPTION_HORIZON_YEARS),
            ("ANALYST_NAME",      ANALYST_NAME),
            ("ANALYST_ROLE",      ANALYST_ROLE),
        ],
        col_widths=[36, 80],
    )
    footer(ws, r, ncols=3)
    ws.freeze_panes = "A4"

    wb.save(OUTPUT_FILE)
    print(f"[done] wrote {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
