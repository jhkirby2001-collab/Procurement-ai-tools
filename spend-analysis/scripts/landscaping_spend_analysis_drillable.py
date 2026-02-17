import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime

# City of Chicago colors
CHICAGO_BLUE = "1F4788"  # City of Chicago blue
CHICAGO_RED = "DC143C"   # Chicago flag red
LIGHT_BLUE = "C5D9F1"    # Light blue for alternating rows
WHITE = "FFFFFF"
GRAY = "D3D3D3"

def format_currency(value):
    """Format value as currency"""
    if pd.isna(value) or value is None:
        return "$0.00"
    return f"${value:,.2f}"

def create_header_style():
    """Create header style with City of Chicago blue"""
    return {
        'font': Font(name='Calibri', size=11, bold=True, color=WHITE),
        'fill': PatternFill(start_color=CHICAGO_BLUE, end_color=CHICAGO_BLUE, fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def create_data_style(is_alternate=False):
    """Create data cell style"""
    fill_color = LIGHT_BLUE if is_alternate else WHITE
    return {
        'font': Font(name='Calibri', size=10),
        'fill': PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color=GRAY),
            right=Side(style='thin', color=GRAY),
            top=Side(style='thin', color=GRAY),
            bottom=Side(style='thin', color=GRAY)
        )
    }

def apply_styles_to_sheet(ws, df, start_row=1, total_row_position=None):
    """Apply formatting to worksheet"""
    header_style = create_header_style()

    # Apply header styles
    for col_num in range(1, len(df.columns) + 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.font = header_style['font']
        cell.fill = header_style['fill']
        cell.alignment = header_style['alignment']
        cell.border = header_style['border']

    # Apply data styles
    for row_num in range(start_row + 1, start_row + len(df) + 1):
        is_alternate = (row_num - start_row) % 2 == 0

        # Check if this is a total/summary row
        is_total = total_row_position and row_num == total_row_position

        data_style = create_data_style(is_alternate and not is_total)

        for col_num in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_style['font']
            cell.fill = data_style['fill']
            cell.alignment = data_style['alignment']
            cell.border = data_style['border']

            # Make total row bold with gray background
            if is_total:
                cell.font = Font(name='Calibri', size=10, bold=True)
                cell.fill = PatternFill(start_color=GRAY, end_color=GRAY, fill_type='solid')

    # Adjust column widths
    for col_num, column in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_num)

        # Set width based on column content
        if 'Amount' in column or 'Total' in column or 'Spend' in column:
            ws.column_dimensions[col_letter].width = 18
        elif 'Vendor' in column or 'Description' in column:
            ws.column_dimensions[col_letter].width = 40
        elif 'Category' in column or 'Service' in column:
            ws.column_dimensions[col_letter].width = 35
        elif 'Department' in column:
            ws.column_dimensions[col_letter].width = 35
        elif 'Contract' in column or 'Specification' in column:
            ws.column_dimensions[col_letter].width = 18
        elif 'Date' in column:
            ws.column_dimensions[col_letter].width = 15
        else:
            ws.column_dimensions[col_letter].width = 15

# Read the structured data
input_path = "/workspaces/Spend_Analysis_and_Categorization/spend-analysis/data/processed/landscaping_structured.xlsx"
df = pd.read_excel(input_path)

# Format award date as string for better display
df['Award_Date_Display'] = pd.to_datetime(df['Award_Date']).dt.strftime('%m/%d/%Y')

# Create output workbook
output_path = "/workspaces/Spend_Analysis_and_Categorization/spend-analysis/outputs/LANDSCAPING_SPEND_ANALYSIS.xlsx"
writer = pd.ExcelWriter(output_path, engine='openpyxl')

print("Generating Drillable Landscaping Spend Analysis Report...")
print("=" * 80)

# Calculate total spend for percentages
total_spend = df['Award_Amount_Numeric'].sum()

# ============================================================================
# ANALYSIS 1: SPEND BY VENDOR (DETAILED)
# ============================================================================
print("\n1. Analyzing spend by vendor with full contract details...")

# Sort by vendor name and then by award amount descending
vendor_detailed = df[['Vendor name', 'Contract (PO) #', 'Specification #', 'Award_Date_Display',
                       'Service_Category', 'Award_Amount_Numeric', 'Department']].copy()
vendor_detailed = vendor_detailed.sort_values(['Vendor name', 'Award_Amount_Numeric'], ascending=[True, False])

# Rename columns
vendor_detailed.columns = ['Vendor Name', 'Contract Number', 'Specification Number', 'Award Date',
                           'Service Category', 'Award Amount', 'Department']

# Format currency
vendor_detailed['Award Amount'] = vendor_detailed['Award Amount'].apply(format_currency)

# Add grand total row only
total_row = pd.DataFrame({
    'Vendor Name': ['GRAND TOTAL'],
    'Contract Number': [f'{len(df)} Total Contracts'],
    'Specification Number': [''],
    'Award Date': [''],
    'Service Category': [''],
    'Award Amount': [format_currency(total_spend)],
    'Department': ['']
})

vendor_final = pd.concat([vendor_detailed, total_row], ignore_index=True)
vendor_final.to_excel(writer, sheet_name='Spend by Vendor', index=False)

# ============================================================================
# ANALYSIS 2: SPEND BY SERVICE CATEGORY (DETAILED)
# ============================================================================
print("2. Analyzing spend by service category with full contract details...")

service_detailed = df[['Service_Category', 'Vendor name', 'Contract (PO) #', 'Specification #',
                        'Award_Date_Display', 'Award_Amount_Numeric', 'Department']].copy()
service_detailed = service_detailed.sort_values(['Service_Category', 'Award_Amount_Numeric'], ascending=[True, False])

service_detailed.columns = ['Service Category', 'Vendor Name', 'Contract Number', 'Specification Number',
                            'Award Date', 'Award Amount', 'Department']

# Format currency
service_detailed['Award Amount'] = service_detailed['Award Amount'].apply(format_currency)

# Add grand total row only
total_row = pd.DataFrame({
    'Service Category': ['GRAND TOTAL'],
    'Vendor Name': [f'{len(df)} Total Contracts'],
    'Contract Number': [''],
    'Specification Number': [''],
    'Award Date': [''],
    'Award Amount': [format_currency(total_spend)],
    'Department': ['']
})

service_final = pd.concat([service_detailed, total_row], ignore_index=True)
service_final.to_excel(writer, sheet_name='Spend by Service', index=False)

# ============================================================================
# ANALYSIS 3: SPEND BY DEPARTMENT (DETAILED)
# ============================================================================
print("3. Analyzing spend by department with full contract details...")

dept_detailed = df[['Department', 'Vendor name', 'Contract (PO) #', 'Specification #',
                     'Award_Date_Display', 'Service_Category', 'Award_Amount_Numeric']].copy()
dept_detailed = dept_detailed.sort_values(['Department', 'Award_Amount_Numeric'], ascending=[True, False])

dept_detailed.columns = ['Department', 'Vendor Name', 'Contract Number', 'Specification Number',
                         'Award Date', 'Service Category', 'Award Amount']

# Format currency
dept_detailed['Award Amount'] = dept_detailed['Award Amount'].apply(format_currency)

# Add grand total row only
total_row = pd.DataFrame({
    'Department': ['GRAND TOTAL'],
    'Vendor Name': [f'{len(df)} Total Contracts'],
    'Contract Number': [''],
    'Specification Number': [''],
    'Award Date': [''],
    'Service Category': [''],
    'Award Amount': [format_currency(total_spend)]
})

dept_final = pd.concat([dept_detailed, total_row], ignore_index=True)
dept_final.to_excel(writer, sheet_name='Spend by Department', index=False)

# ============================================================================
# EXECUTIVE SUMMARY
# ============================================================================
print("4. Creating executive summary...")

# Calculate key metrics
total_contracts = len(df)
total_vendors = df['Vendor name'].nunique()
avg_contract_value = df['Award_Amount_Numeric'].mean()
median_contract_value = df['Award_Amount_Numeric'].median()
largest_contract = df['Award_Amount_Numeric'].max()
smallest_contract = df['Award_Amount_Numeric'].min()

# Top vendors
vendor_summary = df.groupby('Vendor name').agg({
    'Award_Amount_Numeric': 'sum',
    'Contract (PO) #': 'count'
}).reset_index()
vendor_summary = vendor_summary.sort_values('Award_Amount_Numeric', ascending=False)
top_5_vendors = vendor_summary.head(5).copy()
top_5_vendors_spend = top_5_vendors['Award_Amount_Numeric'].sum()
top_5_vendors_pct = (top_5_vendors_spend / total_spend * 100)

# Top service categories
service_summary = df.groupby('Service_Category').agg({
    'Award_Amount_Numeric': 'sum'
}).reset_index().sort_values('Award_Amount_Numeric', ascending=False)
top_3_services = service_summary.head(3).copy()

# Date range
date_range_start = df['Award_Date'].min()
date_range_end = df['Award_Date'].max()

# Create Executive Summary
summary_data = []
summary_data.append(['CITY OF CHICAGO', '', ''])
summary_data.append(['Landscaping Services Spend Analysis', '', ''])
summary_data.append(['Executive Summary', '', ''])
summary_data.append(['', '', ''])
summary_data.append([f'Report Date: {datetime.now().strftime("%B %d, %Y")}', '', ''])
summary_data.append([f'Analysis Period: {date_range_start.strftime("%B %Y")} to {date_range_end.strftime("%B %Y")}', '', ''])
summary_data.append(['', '', ''])
summary_data.append(['KEY FINDINGS', '', ''])
summary_data.append(['', '', ''])
summary_data.append(['OVERALL SPEND SUMMARY', '', ''])
summary_data.append(['Metric', 'Value', ''])
summary_data.append(['Total Spend', format_currency(total_spend), ''])
summary_data.append(['Number of Contracts', f'{total_contracts:,}', ''])
summary_data.append(['Number of Unique Vendors', f'{total_vendors:,}', ''])
summary_data.append(['Average Contract Value', format_currency(avg_contract_value), ''])
summary_data.append(['Median Contract Value', format_currency(median_contract_value), ''])
summary_data.append(['Largest Contract', format_currency(largest_contract), ''])
summary_data.append(['Smallest Contract', format_currency(smallest_contract), ''])
summary_data.append(['', '', ''])

summary_data.append(['TOP 5 VENDORS BY SPEND', '', ''])
summary_data.append(['Vendor', 'Total Spend', '% of Total'])
for idx, row in top_5_vendors.iterrows():
    vendor_pct = (row['Award_Amount_Numeric'] / total_spend * 100)
    summary_data.append([
        row['Vendor name'],
        format_currency(row['Award_Amount_Numeric']),
        f"{vendor_pct:.2f}%"
    ])
summary_data.append(['Top 5 Vendors Combined', format_currency(top_5_vendors_spend), f'{top_5_vendors_pct:.2f}%'])
summary_data.append(['', '', ''])

summary_data.append(['TOP 3 SERVICE CATEGORIES BY SPEND', '', ''])
summary_data.append(['Service Category', 'Total Spend', '% of Total'])
for idx, row in top_3_services.iterrows():
    service_pct = (row['Award_Amount_Numeric'] / total_spend * 100)
    summary_data.append([
        row['Service_Category'],
        format_currency(row['Award_Amount_Numeric']),
        f"{service_pct:.2f}%"
    ])
summary_data.append(['', '', ''])

summary_data.append(['KEY INSIGHTS', '', ''])
summary_data.append(['', '', ''])

# Generate insights
insights = []
insights.append(f"• The total landscaping spend amounts to {format_currency(total_spend)} across {total_contracts} contracts.")
insights.append(f"• The City works with {total_vendors} unique vendors for landscaping services.")
insights.append(f"• The top 5 vendors account for {top_5_vendors_pct:.1f}% of total spend, indicating {'high' if top_5_vendors_pct > 60 else 'moderate'} vendor concentration.")
insights.append(f"• {vendor_summary.iloc[0]['Vendor name']} is the largest vendor with {format_currency(vendor_summary.iloc[0]['Award_Amount_Numeric'])} ({(vendor_summary.iloc[0]['Award_Amount_Numeric']/total_spend*100):.1f}% of total spend).")
insights.append(f"• {service_summary.iloc[0]['Service_Category']} represents the largest spend category at {format_currency(service_summary.iloc[0]['Award_Amount_Numeric'])} ({(service_summary.iloc[0]['Award_Amount_Numeric']/total_spend*100):.1f}% of total).")
insights.append(f"• The top 3 service categories account for {(top_3_services['Award_Amount_Numeric'].sum()/total_spend*100):.1f}% of total landscaping spend.")

# Check for department concentration
dept_summary = df.groupby('Department').agg({
    'Award_Amount_Numeric': 'sum'
}).reset_index().sort_values('Award_Amount_Numeric', ascending=False)
top_dept_pct = (dept_summary.iloc[0]['Award_Amount_Numeric'] / total_spend * 100)
insights.append(f"• {dept_summary.iloc[0]['Department']} has the highest departmental spend at {format_currency(dept_summary.iloc[0]['Award_Amount_Numeric'])} ({top_dept_pct:.1f}% of total).")

for insight in insights:
    summary_data.append([insight, '', ''])

summary_data.append(['', '', ''])
summary_data.append(['DATA TRANSPARENCY', '', ''])
summary_data.append(['', '', ''])

transparency_notes = []
transparency_notes.append("• All analysis tabs display complete contract-level details with no summarization.")
transparency_notes.append("• Each row shows the actual vendor name, contract number, specification number, and award date.")
transparency_notes.append("• Users can sort, filter, and analyze data without relying on pre-calculated subtotals.")
transparency_notes.append("• Excel filters are available on all columns for custom analysis.")

for note in transparency_notes:
    summary_data.append([note, '', ''])

summary_data.append(['', '', ''])
summary_data.append(['RECOMMENDATIONS', '', ''])
summary_data.append(['', '', ''])

recommendations = []
if top_5_vendors_pct > 70:
    recommendations.append("• Consider diversifying vendor base to reduce dependency on top vendors.")
if (service_summary.iloc[0]['Award_Amount_Numeric'] / total_spend * 100) > 30:
    recommendations.append("• High concentration in one service category suggests opportunity for strategic sourcing.")
recommendations.append("• Review contracts with highest values for potential cost optimization opportunities.")
recommendations.append("• Analyze regional spend patterns to identify potential service consolidation opportunities.")
recommendations.append("• Use Excel pivot tables on the detail tabs to create custom summaries by any dimension.")

for rec in recommendations:
    summary_data.append([rec, '', ''])

summary_df = pd.DataFrame(summary_data, columns=['Column1', 'Column2', 'Column3'])
summary_df.to_excel(writer, sheet_name='Executive Summary', index=False, header=False)

# Save the workbook
writer.close()

# ============================================================================
# APPLY FORMATTING TO ALL SHEETS
# ============================================================================
print("5. Applying City of Chicago formatting...")

wb = openpyxl.load_workbook(output_path)

# Format Executive Summary
ws_summary = wb['Executive Summary']
ws_summary.column_dimensions['A'].width = 70
ws_summary.column_dimensions['B'].width = 25
ws_summary.column_dimensions['C'].width = 15

# Title formatting
ws_summary['A1'].font = Font(name='Calibri', size=24, bold=True, color=CHICAGO_BLUE)
ws_summary['A2'].font = Font(name='Calibri', size=18, bold=True, color=CHICAGO_RED)
ws_summary['A3'].font = Font(name='Calibri', size=14, bold=True, color=CHICAGO_BLUE)

# Section headers
section_rows = [8, 10, 20, 27, 34, 42]
for row in section_rows:
    cell = ws_summary.cell(row=row, column=1)
    cell.font = Font(name='Calibri', size=12, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=CHICAGO_BLUE, end_color=CHICAGO_BLUE, fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center')

# Format data tables in summary
for row in range(11, 19):  # Overall Summary table
    for col in range(1, 3):
        cell = ws_summary.cell(row=row, column=col)
        if row == 11:  # Header
            cell.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
            cell.fill = PatternFill(start_color=CHICAGO_BLUE, end_color=CHICAGO_BLUE, fill_type='solid')
        else:
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Format Spend by Vendor sheet
ws_vendor = wb['Spend by Vendor']
total_row_vendor = len(vendor_final) + 1
apply_styles_to_sheet(ws_vendor, vendor_final, start_row=1, total_row_position=total_row_vendor)

# Format Spend by Service sheet
ws_service = wb['Spend by Service']
total_row_service = len(service_final) + 1
apply_styles_to_sheet(ws_service, service_final, start_row=1, total_row_position=total_row_service)

# Format Spend by Department sheet
ws_dept = wb['Spend by Department']
total_row_dept = len(dept_final) + 1
apply_styles_to_sheet(ws_dept, dept_final, start_row=1, total_row_position=total_row_dept)

# Enable autofilter on all data sheets
ws_vendor.auto_filter.ref = ws_vendor.dimensions
ws_service.auto_filter.ref = ws_service.dimensions
ws_dept.auto_filter.ref = ws_dept.dimensions

# Freeze panes on data sheets (freeze header row)
ws_vendor.freeze_panes = 'A2'
ws_service.freeze_panes = 'A2'
ws_dept.freeze_panes = 'A2'

# Set sheet order - Executive Summary first
wb._sheets = [wb['Executive Summary']] + [sheet for sheet in wb._sheets if sheet.title != 'Executive Summary']

wb.save(output_path)

print("\n" + "=" * 80)
print("DRILLABLE REPORT GENERATION COMPLETE!")
print("=" * 80)
print(f"\nOutput file: {output_path}")
print("\nReport includes:")
print("  ✓ Executive Summary with key findings and recommendations")
print("  ✓ Spend by Vendor - ALL 176 contracts with full details (NO SUBTOTALS)")
print("  ✓ Spend by Service - ALL 176 contracts with full details (NO SUBTOTALS)")
print("  ✓ Spend by Department - ALL 176 contracts with full details (NO SUBTOTALS)")
print("  ✓ Every row shows actual vendor name, contract #, spec #, and award date")
print("  ✓ Excel autofilters enabled on all data tabs for custom analysis")
print("  ✓ City of Chicago color scheme applied")
print("  ✓ Proper number formatting with decimals and commas")
print("  ✓ Text wrapping and frozen header rows for readability")
print("  ✓ Fully transparent data - no assumptions needed!")
print("\n" + "=" * 80)
