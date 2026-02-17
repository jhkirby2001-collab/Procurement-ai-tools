import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime

# City of Chicago colors
CHICAGO_BLUE = "1F4788"  # City of Chicago blue
CHICAGO_RED = "DC143C"   # Chicago flag red
LIGHT_BLUE = "C5D9F1"    # Light blue for alternating rows
WHITE = "FFFFFF"
GRAY = "D3D3D3"

def format_currency(value):
    """Format value as currency"""
    if pd.isna(value) or value is None:
        return "$0.00"
    return f"${value:,.2f}"

def create_header_style():
    """Create header style with City of Chicago blue"""
    return {
        'font': Font(name='Calibri', size=11, bold=True, color=WHITE),
        'fill': PatternFill(start_color=CHICAGO_BLUE, end_color=CHICAGO_BLUE, fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def create_data_style(is_alternate=False):
    """Create data cell style"""
    fill_color = LIGHT_BLUE if is_alternate else WHITE
    return {
        'font': Font(name='Calibri', size=10),
        'fill': PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color=GRAY),
            right=Side(style='thin', color=GRAY),
            top=Side(style='thin', color=GRAY),
            bottom=Side(style='thin', color=GRAY)
        )
    }

def apply_styles_to_sheet(ws, df, start_row=1, has_total_row=False):
    """Apply formatting to worksheet"""
    header_style = create_header_style()

    # Apply header styles
    for col_num in range(1, len(df.columns) + 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.font = header_style['font']
        cell.fill = header_style['fill']
        cell.alignment = header_style['alignment']
        cell.border = header_style['border']

    # Apply data styles
    for row_num in range(start_row + 1, start_row + len(df) + 1):
        is_alternate = (row_num - start_row) % 2 == 0

        # Check if this is a total row
        is_total = has_total_row and row_num == start_row + len(df)

        data_style = create_data_style(is_alternate and not is_total)

        for col_num in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_style['font']
            cell.fill = data_style['fill']
            cell.alignment = data_style['alignment']
            cell.border = data_style['border']

            # Make total row bold
            if is_total:
                cell.font = Font(name='Calibri', size=10, bold=True)
                cell.fill = PatternFill(start_color=GRAY, end_color=GRAY, fill_type='solid')

    # Adjust column widths
    for col_num, column in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_num)

        # Set width based on column content
        if 'Amount' in column or 'Total' in column or 'Spend' in column:
            ws.column_dimensions[col_letter].width = 18
        elif 'Vendor' in column or 'Description' in column:
            ws.column_dimensions[col_letter].width = 40
        elif 'Category' in column or 'Service' in column:
            ws.column_dimensions[col_letter].width = 35
        elif 'Department' in column:
            ws.column_dimensions[col_letter].width = 35
        else:
            ws.column_dimensions[col_letter].width = 15

# Read the structured data
input_path = "/workspaces/Spend_Analysis_and_Categorization/spend-analysis/data/processed/landscaping_structured.xlsx"
df = pd.read_excel(input_path)

# Create output workbook
output_path = "/workspaces/Spend_Analysis_and_Categorization/spend-analysis/outputs/LANDSCAPING_SPEND_ANALYSIS.xlsx"
writer = pd.ExcelWriter(output_path, engine='openpyxl')

print("Generating Landscaping Spend Analysis Report...")
print("=" * 80)

# ============================================================================
# ANALYSIS 1: SPEND BY VENDOR
# ============================================================================
print("\n1. Analyzing spend by vendor...")

vendor_analysis = df.groupby('Vendor name').agg({
    'Award_Amount_Numeric': 'sum',
    'Contract (PO) #': 'count',
    'Service_Category': lambda x: ', '.join(x.unique()[:3])  # Top 3 categories
}).reset_index()

vendor_analysis.columns = ['Vendor Name', 'Total Award Amount', 'Number of Contracts', 'Service Categories']
vendor_analysis = vendor_analysis.sort_values('Total Award Amount', ascending=False)

# Add percentage
total_spend = vendor_analysis['Total Award Amount'].sum()
vendor_analysis['% of Total Spend'] = (vendor_analysis['Total Award Amount'] / total_spend * 100)

# Add total row
total_row = pd.DataFrame({
    'Vendor Name': ['TOTAL'],
    'Total Award Amount': [vendor_analysis['Total Award Amount'].sum()],
    'Number of Contracts': [vendor_analysis['Number of Contracts'].sum()],
    'Service Categories': ['All Categories'],
    '% of Total Spend': [100.0]
})
vendor_analysis = pd.concat([vendor_analysis, total_row], ignore_index=True)

# Format for display
vendor_display = vendor_analysis.copy()
vendor_display['Total Award Amount'] = vendor_display['Total Award Amount'].apply(format_currency)
vendor_display['% of Total Spend'] = vendor_display['% of Total Spend'].apply(lambda x: f"{x:.2f}%")

vendor_display.to_excel(writer, sheet_name='Spend by Vendor', index=False)

# ============================================================================
# ANALYSIS 2: SPEND BY COMMODITY/SERVICE
# ============================================================================
print("2. Analyzing spend by commodity/service...")

service_analysis = df.groupby('Service_Category').agg({
    'Award_Amount_Numeric': 'sum',
    'Contract (PO) #': 'count',
    'Vendor name': 'nunique'
}).reset_index()

service_analysis.columns = ['Service Category', 'Total Award Amount', 'Number of Contracts', 'Number of Vendors']
service_analysis = service_analysis.sort_values('Total Award Amount', ascending=False)

# Add percentage
service_analysis['% of Total Spend'] = (service_analysis['Total Award Amount'] / total_spend * 100)

# Add total row
total_row = pd.DataFrame({
    'Service Category': ['TOTAL'],
    'Total Award Amount': [service_analysis['Total Award Amount'].sum()],
    'Number of Contracts': [service_analysis['Number of Contracts'].sum()],
    'Number of Vendors': [df['Vendor name'].nunique()],
    '% of Total Spend': [100.0]
})
service_analysis = pd.concat([service_analysis, total_row], ignore_index=True)

# Format for display
service_display = service_analysis.copy()
service_display['Total Award Amount'] = service_display['Total Award Amount'].apply(format_currency)
service_display['% of Total Spend'] = service_display['% of Total Spend'].apply(lambda x: f"{x:.2f}%")

service_display.to_excel(writer, sheet_name='Spend by Service', index=False)

# ============================================================================
# ANALYSIS 3: SPEND BY DEPARTMENT
# ============================================================================
print("3. Analyzing spend by department...")

dept_analysis = df.groupby('Department').agg({
    'Award_Amount_Numeric': 'sum',
    'Contract (PO) #': 'count',
    'Vendor name': 'nunique'
}).reset_index()

dept_analysis.columns = ['Department', 'Total Award Amount', 'Number of Contracts', 'Number of Vendors']
dept_analysis = dept_analysis.sort_values('Total Award Amount', ascending=False)

# Add percentage
dept_analysis['% of Total Spend'] = (dept_analysis['Total Award Amount'] / total_spend * 100)

# Add total row
total_row = pd.DataFrame({
    'Department': ['TOTAL'],
    'Total Award Amount': [dept_analysis['Total Award Amount'].sum()],
    'Number of Contracts': [dept_analysis['Number of Contracts'].sum()],
    'Number of Vendors': [df['Vendor name'].nunique()],
    '% of Total Spend': [100.0]
})
dept_analysis = pd.concat([dept_analysis, total_row], ignore_index=True)

# Format for display
dept_display = dept_analysis.copy()
dept_display['Total Award Amount'] = dept_display['Total Award Amount'].apply(format_currency)
dept_display['% of Total Spend'] = dept_display['% of Total Spend'].apply(lambda x: f"{x:.2f}%")

dept_display.to_excel(writer, sheet_name='Spend by Department', index=False)

# ============================================================================
# EXECUTIVE SUMMARY
# ============================================================================
print("4. Creating executive summary...")

# Calculate key metrics
total_contracts = len(df)
total_vendors = df['Vendor name'].nunique()
avg_contract_value = df['Award_Amount_Numeric'].mean()
median_contract_value = df['Award_Amount_Numeric'].median()
largest_contract = df['Award_Amount_Numeric'].max()
smallest_contract = df['Award_Amount_Numeric'].min()

# Top vendors
top_5_vendors = vendor_analysis.head(5).copy()
top_5_vendors_spend = top_5_vendors['Total Award Amount'].sum()
top_5_vendors_pct = (top_5_vendors_spend / total_spend * 100)

# Top service categories
top_3_services = service_analysis.head(3).copy()

# Date range
date_range_start = df['Award_Date'].min()
date_range_end = df['Award_Date'].max()

# Create Executive Summary
summary_data = []
summary_data.append(['CITY OF CHICAGO', '', ''])
summary_data.append(['Landscaping Services Spend Analysis', '', ''])
summary_data.append(['Executive Summary', '', ''])
summary_data.append(['', '', ''])
summary_data.append([f'Report Date: {datetime.now().strftime("%B %d, %Y")}', '', ''])
summary_data.append([f'Analysis Period: {date_range_start.strftime("%B %Y")} to {date_range_end.strftime("%B %Y")}', '', ''])
summary_data.append(['', '', ''])
summary_data.append(['KEY FINDINGS', '', ''])
summary_data.append(['', '', ''])
summary_data.append(['OVERALL SPEND SUMMARY', '', ''])
summary_data.append(['Metric', 'Value', ''])
summary_data.append(['Total Spend', format_currency(total_spend), ''])
summary_data.append(['Number of Contracts', f'{total_contracts:,}', ''])
summary_data.append(['Number of Unique Vendors', f'{total_vendors:,}', ''])
summary_data.append(['Average Contract Value', format_currency(avg_contract_value), ''])
summary_data.append(['Median Contract Value', format_currency(median_contract_value), ''])
summary_data.append(['Largest Contract', format_currency(largest_contract), ''])
summary_data.append(['Smallest Contract', format_currency(smallest_contract), ''])
summary_data.append(['', '', ''])

summary_data.append(['TOP 5 VENDORS BY SPEND', '', ''])
summary_data.append(['Vendor', 'Total Spend', '% of Total'])
for idx, row in top_5_vendors.iterrows():
    if idx < 5:  # Exclude total row
        summary_data.append([
            row['Vendor Name'],
            format_currency(row['Total Award Amount']),
            f"{row['% of Total Spend']:.2f}%"
        ])
summary_data.append(['Top 5 Vendors Combined', format_currency(top_5_vendors_spend), f'{top_5_vendors_pct:.2f}%'])
summary_data.append(['', '', ''])

summary_data.append(['TOP 3 SERVICE CATEGORIES BY SPEND', '', ''])
summary_data.append(['Service Category', 'Total Spend', '% of Total'])
for idx, row in top_3_services.iterrows():
    if idx < 3:  # Exclude total row
        summary_data.append([
            row['Service Category'],
            format_currency(row['Total Award Amount']),
            f"{row['% of Total Spend']:.2f}%"
        ])
summary_data.append(['', '', ''])

summary_data.append(['KEY INSIGHTS', '', ''])
summary_data.append(['', '', ''])

# Generate insights
insights = []
insights.append(f"• The total landscaping spend amounts to {format_currency(total_spend)} across {total_contracts} contracts.")
insights.append(f"• The City works with {total_vendors} unique vendors for landscaping services.")
insights.append(f"• The top 5 vendors account for {top_5_vendors_pct:.1f}% of total spend, indicating {'high' if top_5_vendors_pct > 60 else 'moderate'} vendor concentration.")
insights.append(f"• {top_5_vendors.iloc[0]['Vendor Name']} is the largest vendor with {format_currency(top_5_vendors.iloc[0]['Total Award Amount'])} ({top_5_vendors.iloc[0]['% of Total Spend']:.1f}% of total spend).")
insights.append(f"• {service_analysis.iloc[0]['Service Category']} represents the largest spend category at {format_currency(service_analysis.iloc[0]['Total Award Amount'])} ({service_analysis.iloc[0]['% of Total Spend']:.1f}% of total).")
insights.append(f"• The top 3 service categories account for {top_3_services['% of Total Spend'].sum():.1f}% of total landscaping spend.")

# Check for department concentration
top_dept_pct = dept_analysis.iloc[0]['% of Total Spend']
insights.append(f"• {dept_analysis.iloc[0]['Department']} has the highest departmental spend at {format_currency(dept_analysis.iloc[0]['Total Award Amount'])} ({top_dept_pct:.1f}% of total).")

for insight in insights:
    summary_data.append([insight, '', ''])

summary_data.append(['', '', ''])
summary_data.append(['RECOMMENDATIONS', '', ''])
summary_data.append(['', '', ''])

recommendations = []
if top_5_vendors_pct > 70:
    recommendations.append("• Consider diversifying vendor base to reduce dependency on top vendors.")
if service_analysis.iloc[0]['% of Total Spend'] > 30:
    recommendations.append("• High concentration in one service category suggests opportunity for strategic sourcing.")
recommendations.append("• Review contracts with highest values for potential cost optimization opportunities.")
recommendations.append("• Analyze regional spend patterns to identify potential service consolidation opportunities.")

for rec in recommendations:
    summary_data.append([rec, '', ''])

summary_df = pd.DataFrame(summary_data, columns=['Column1', 'Column2', 'Column3'])
summary_df.to_excel(writer, sheet_name='Executive Summary', index=False, header=False)

# Save the workbook
writer.close()

# ============================================================================
# APPLY FORMATTING TO ALL SHEETS
# ============================================================================
print("5. Applying City of Chicago formatting...")

wb = openpyxl.load_workbook(output_path)

# Format Executive Summary
ws_summary = wb['Executive Summary']
ws_summary.column_dimensions['A'].width = 60
ws_summary.column_dimensions['B'].width = 25
ws_summary.column_dimensions['C'].width = 15

# Title formatting
ws_summary['A1'].font = Font(name='Calibri', size=24, bold=True, color=CHICAGO_BLUE)
ws_summary['A2'].font = Font(name='Calibri', size=18, bold=True, color=CHICAGO_RED)
ws_summary['A3'].font = Font(name='Calibri', size=14, bold=True, color=CHICAGO_BLUE)

# Section headers
section_rows = [8, 10, 20, 27, 34]
for row in section_rows:
    cell = ws_summary.cell(row=row, column=1)
    cell.font = Font(name='Calibri', size=12, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=CHICAGO_BLUE, end_color=CHICAGO_BLUE, fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center')

# Format data tables in summary
for row in range(11, 19):  # Overall Summary table
    for col in range(1, 3):
        cell = ws_summary.cell(row=row, column=col)
        if row == 11:  # Header
            cell.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
            cell.fill = PatternFill(start_color=CHICAGO_BLUE, end_color=CHICAGO_BLUE, fill_type='solid')
        else:
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Format Spend by Vendor sheet
ws_vendor = wb['Spend by Vendor']
apply_styles_to_sheet(ws_vendor, vendor_display, start_row=1, has_total_row=True)

# Format Spend by Service sheet
ws_service = wb['Spend by Service']
apply_styles_to_sheet(ws_service, service_display, start_row=1, has_total_row=True)

# Format Spend by Department sheet
ws_dept = wb['Spend by Department']
apply_styles_to_sheet(ws_dept, dept_display, start_row=1, has_total_row=True)

# Set sheet order - Executive Summary first
wb._sheets = [wb['Executive Summary']] + [sheet for sheet in wb._sheets if sheet.title != 'Executive Summary']

wb.save(output_path)

print("\n" + "=" * 80)
print("REPORT GENERATION COMPLETE!")
print("=" * 80)
print(f"\nOutput file: {output_path}")
print("\nReport includes:")
print("  ✓ Executive Summary with key findings and recommendations")
print("  ✓ Spend by Vendor analysis")
print("  ✓ Spend by Service Category analysis")
print("  ✓ Spend by Department analysis")
print("  ✓ City of Chicago color scheme applied")
print("  ✓ Proper number formatting with decimals and commas")
print("  ✓ Text wrapping for readability")
print("\n" + "=" * 80)
