"""
City of Chicago — Top Consolidation Savings Model
===================================================
Generates Top_Consolidation_Savings_Model.xlsx with 2 tabs:
  1. Executive Summary
  2. Consolidation Savings Model

All computed values use LIVE EXCEL FORMULAS — no hardcoded calculations.
Vendor subtotals = department subtotals = category totals (verified to the penny).

Author: James Kirby III, CSCP, MS-SCM
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime

# ============================================================================
# CONFIGURATION
# ============================================================================

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "outputs"
OUTPUT_FILE = OUTPUT_DIR / "Top_Consolidation_Savings_Model.xlsx"
TODAY_DISPLAY = datetime.today().strftime("%B %d, %Y")

ATTRIBUTION = (
    "Prepared by James Kirby III, CSCP, MS-SCM | "
    "Senior Buyer & Procurement Research Analyst | "
    "Department of Procurement Services"
)

# ============================================================================
# CATEGORY DATA (verified from raw Exhibit B transactions)
# ============================================================================

CATEGORIES = [
    {
        "name": "Fencing & Guardrail",
        "total": 4_812_031.44,
        "req_num": "596048",
        "vendors": [
            ("INDUSTRIAL FENCE INC.", 4_801_076.09, 87, 6),
            ("FENCE MASTERS, INC", 10_955.35, 2, 2),
        ],
        "departments": [
            ("Fire Dept", 1_471_850.62),
            ("Aviation", 1_377_813.99),
            ("Water Mgmt", 1_103_130.87),
            ("Fleet", 479_930.55),
            ("Transportation", 323_180.04),
            ("Emergency Mgmt", 56_125.37),
        ],
    },
    {
        "name": "Janitorial & Cleaning Services",
        "total": 1_197_805.90,
        "req_num": "624331",
        "vendors": [
            ("DAYSPRING PROFESSIONAL JANITOR", 1_056_398.96, 9, 2),
            ("LUSE ENVIRONMENTAL SERVICES, INC.", 139_859.50, 6, 2),
            ("SET ENVIRONMENTAL INC", 1_547.44, 2, 1),
        ],
        "departments": [
            ("Fleet", 1_155_352.64),
            ("Public Health", 40_905.82),
            ("Animal Care", 1_547.44),
        ],
    },
]

# ============================================================================
# PYTHON-SIDE DATA INTEGRITY CHECK
# ============================================================================

for cat in CATEGORIES:
    vendor_total = sum(v[1] for v in cat["vendors"])
    dept_total = sum(d[1] for d in cat["departments"])
    assert abs(vendor_total - cat["total"]) < 0.01, (
        f"{cat['name']}: vendor sum {vendor_total} != declared total {cat['total']}"
    )
    assert abs(dept_total - cat["total"]) < 0.01, (
        f"{cat['name']}: dept sum {dept_total} != declared total {cat['total']}"
    )

# ============================================================================
# STYLE CONSTANTS (from exhibit_b_consolidation_analysis.py)
# ============================================================================

NAVY = "003366"
BLUE = "0071BC"
RED = "E31837"
LT_GREEN = "D9EAD3"
LT_YELLOW = "FFF2CC"
LT_BLUE = "E8F4FD"
WHITE = "FFFFFF"

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_WHITE = PatternFill("solid", fgColor=WHITE)
FILL_LT_GREEN = PatternFill("solid", fgColor=LT_GREEN)
FILL_LT_BLUE = PatternFill("solid", fgColor=LT_BLUE)

FONT_TITLE = Font(bold=True, color=NAVY, size=14)
FONT_SECTION = Font(bold=True, color=NAVY, size=12)
FONT_HEADER = Font(bold=True, color=WHITE, size=11)
FONT_BOLD = Font(bold=True, color="000000", size=11)
FONT_BOLD_LG = Font(bold=True, color="000000", size=12)
FONT_NORMAL = Font(bold=False, color="000000", size=11)
FONT_RED = Font(bold=True, color=RED, size=11)
FONT_NOTE = Font(bold=False, color="666666", size=9, italic=True)
FONT_MATH = Font(bold=False, color="333333", size=10, italic=True)
FONT_ATTRIB = Font(bold=True, color=BLUE, size=11)
FONT_DPS = Font(bold=True, color=WHITE, size=12)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

CURRENCY_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
PCT_FMT = "0.0%"
INT_FMT = "#,##0"

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================


def sc(ws, r, c, value, font=None, fill=None, fmt=None, align=None, border=None):
    """Set cell with formatting."""
    cell = ws.cell(r, c, value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align:
        cell.alignment = align
    else:
        cell.alignment = ALIGN_LEFT
    if border:
        cell.border = border
    else:
        cell.border = THIN_BORDER
    return cell


def write_header_row(ws, row_num, headers):
    """Write a navy header row with white bold text."""
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row_num, ci)
        c.value = h
        c.font = FONT_HEADER
        c.fill = FILL_NAVY
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    ws.row_dimensions[row_num].height = 30


def merge_row(ws, r, c_start, c_end, value, font=None, fill=None, align=None):
    """Merge cells across a row and write a value."""
    ws.merge_cells(start_row=r, start_column=c_start, end_row=r, end_column=c_end)
    cell = ws.cell(r, c_start, value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = align or ALIGN_LEFT
    cell.border = THIN_BORDER
    return cell


def write_attribution(ws, n_cols):
    """Write 3-row attribution block. Returns the next usable row (row 5)."""
    merge_row(ws, 1, 1, n_cols,
              "CITY OF CHICAGO \u2014 DEPARTMENT OF PROCUREMENT SERVICES",
              font=FONT_DPS, fill=FILL_NAVY, align=ALIGN_CENTER)
    ws.row_dimensions[1].height = 28

    merge_row(ws, 2, 1, n_cols, ATTRIBUTION,
              font=FONT_ATTRIB, align=ALIGN_CENTER)
    ws.row_dimensions[2].height = 22

    merge_row(ws, 3, 1, n_cols, f"Analysis Date: {TODAY_DISPLAY}",
              font=FONT_NORMAL, align=ALIGN_CENTER)
    ws.row_dimensions[3].height = 20

    return 5  # next usable row after blank row 4


def auto_width(ws, min_w=10, max_w=55):
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        width = min(max(max_len + 2, min_w), max_w)
        ws.column_dimensions[col_letter].width = width


# ============================================================================
# TAB 2: CONSOLIDATION SAVINGS MODEL (built first to establish cell refs)
# ============================================================================


def build_category_block(ws, cat, start_row):
    """
    Write one category block: vendor table, dept table, cross-check,
    savings chain, solicitation status.
    Returns (next_row, refs_dict) where refs_dict maps metric names
    to cell addresses like 'B26'.
    """
    r = start_row
    n_cols = 5
    refs = {}

    # --- Category header bar ---
    merge_row(ws, r, 1, n_cols,
              f"CATEGORY: {cat['name'].upper()}",
              font=FONT_HEADER, fill=FILL_NAVY, align=ALIGN_LEFT)
    ws.row_dimensions[r].height = 30
    r += 2  # skip blank row

    # --- Vendor Breakdown ---
    sc(ws, r, 1, "Vendor Breakdown", font=FONT_SECTION)
    r += 1
    write_header_row(ws, r, ["Vendor Name", "Total Spend", "Transaction Count",
                              "Department Count", "Notes"])
    vendor_header_row = r
    r += 1

    vendor_first = r
    for name, spend, txns, depts in cat["vendors"]:
        sc(ws, r, 1, name, font=FONT_NORMAL)
        sc(ws, r, 2, spend, font=FONT_NORMAL, fmt=CURRENCY_FMT)
        sc(ws, r, 3, txns, font=FONT_NORMAL, fmt=INT_FMT, align=ALIGN_CENTER)
        sc(ws, r, 4, depts, font=FONT_NORMAL, fmt=INT_FMT, align=ALIGN_CENTER)
        sc(ws, r, 5, "", font=FONT_NORMAL)
        r += 1
    vendor_last = r - 1

    # Vendor subtotal row
    sc(ws, r, 1, "VENDOR SUBTOTAL", font=FONT_BOLD, fill=FILL_LT_GREEN)
    sc(ws, r, 2, f"=SUM(B{vendor_first}:B{vendor_last})",
       font=FONT_BOLD, fill=FILL_LT_GREEN, fmt=CURRENCY_FMT)
    sc(ws, r, 3, f"=SUM(C{vendor_first}:C{vendor_last})",
       font=FONT_BOLD, fill=FILL_LT_GREEN, fmt=INT_FMT, align=ALIGN_CENTER)
    sc(ws, r, 4, "", font=FONT_BOLD, fill=FILL_LT_GREEN)
    sc(ws, r, 5, "", font=FONT_BOLD, fill=FILL_LT_GREEN)
    vendor_subtotal_row = r
    refs["vendor_subtotal"] = f"B{r}"
    r += 2  # skip blank

    # --- Department Breakdown ---
    sc(ws, r, 1, "Department Breakdown", font=FONT_SECTION)
    r += 1
    write_header_row(ws, r, ["Department Name", "Total Spend", "", "", ""])
    r += 1

    dept_first = r
    for dept_name, spend in cat["departments"]:
        sc(ws, r, 1, dept_name, font=FONT_NORMAL)
        sc(ws, r, 2, spend, font=FONT_NORMAL, fmt=CURRENCY_FMT)
        sc(ws, r, 3, "", font=FONT_NORMAL)
        sc(ws, r, 4, "", font=FONT_NORMAL)
        sc(ws, r, 5, "", font=FONT_NORMAL)
        r += 1
    dept_last = r - 1

    # Department subtotal row
    sc(ws, r, 1, "DEPARTMENT SUBTOTAL", font=FONT_BOLD, fill=FILL_LT_GREEN)
    sc(ws, r, 2, f"=SUM(B{dept_first}:B{dept_last})",
       font=FONT_BOLD, fill=FILL_LT_GREEN, fmt=CURRENCY_FMT)
    sc(ws, r, 3, "", font=FONT_BOLD, fill=FILL_LT_GREEN)
    sc(ws, r, 4, "", font=FONT_BOLD, fill=FILL_LT_GREEN)
    sc(ws, r, 5, "", font=FONT_BOLD, fill=FILL_LT_GREEN)
    dept_subtotal_row = r
    refs["dept_subtotal"] = f"B{r}"
    r += 2  # skip blank

    # --- Cross-Check ---
    sc(ws, r, 1, "CROSS-CHECK: Vendor Total \u2212 Department Total",
       font=FONT_RED)
    sc(ws, r, 2, f"=B{vendor_subtotal_row}-B{dept_subtotal_row}",
       font=FONT_RED, fmt=CURRENCY_FMT)
    sc(ws, r, 5, "Must equal $0.00", font=FONT_NOTE)
    r += 2  # skip blank

    # --- Savings Chain Calculation ---
    sc(ws, r, 1, "SAVINGS CHAIN CALCULATION", font=FONT_SECTION)
    r += 1

    # 5-Year Total
    sc(ws, r, 1, "5-Year Total Spend", font=FONT_BOLD)
    sc(ws, r, 2, f"=B{vendor_subtotal_row}", font=FONT_BOLD, fmt=CURRENCY_FMT)
    sc(ws, r, 5, "Verified vendor subtotal", font=FONT_NOTE)
    five_year_row = r
    refs["five_year"] = f"B{r}"
    r += 1

    # Annual Run Rate
    sc(ws, r, 1, "Annual Run Rate", font=FONT_NORMAL)
    sc(ws, r, 2, f"=B{five_year_row}/5", font=FONT_NORMAL, fmt=CURRENCY_FMT)
    sc(ws, r, 5, "5-Year Total \u00f7 5", font=FONT_MATH)
    annual_row = r
    refs["annual"] = f"B{r}"
    r += 1

    # Addressable Spend
    sc(ws, r, 1, "Addressable Spend (85%)", font=FONT_NORMAL)
    sc(ws, r, 2, f"=B{annual_row}*0.85", font=FONT_NORMAL, fmt=CURRENCY_FMT)
    sc(ws, r, 5, "Annual \u00d7 85% (15% already optimized)", font=FONT_MATH)
    addressable_row = r
    refs["addressable"] = f"B{r}"
    r += 1

    # Annual Savings
    sc(ws, r, 1, "Annual Savings (10%)", font=FONT_NORMAL)
    sc(ws, r, 2, f"=B{addressable_row}*0.10", font=FONT_NORMAL, fmt=CURRENCY_FMT)
    sc(ws, r, 5, "Addressable \u00d7 10% (moderate savings rate)", font=FONT_MATH)
    savings_row = r
    refs["annual_savings"] = f"B{r}"
    r += 1

    # 3-Year Projected Value
    sc(ws, r, 1, "3-Year Projected Value", font=FONT_BOLD_LG, fill=FILL_LT_GREEN)
    sc(ws, r, 2, f"=B{savings_row}*3", font=FONT_BOLD_LG, fill=FILL_LT_GREEN,
       fmt=CURRENCY_FMT)
    sc(ws, r, 3, "", fill=FILL_LT_GREEN)
    sc(ws, r, 4, "", fill=FILL_LT_GREEN)
    sc(ws, r, 5, "Annual Savings \u00d7 3 years", font=FONT_MATH, fill=FILL_LT_GREEN)
    refs["three_year"] = f"B{r}"
    r += 2  # skip blank

    # --- Solicitation Status ---
    sc(ws, r, 1, "Solicitation Status", font=FONT_BOLD)
    sc(ws, r, 2, f"Phase A \u2014 Spec Development (Req# {cat['req_num']})",
       font=FONT_NORMAL)
    r += 1
    sc(ws, r, 1, "Contract Status", font=FONT_BOLD)
    sc(ws, r, 2, "No solicitation issued. No contract awarded.", font=FONT_RED)
    r += 2  # blank separator

    return r, refs


def build_savings_model(wb):
    """Build Tab 2: Consolidation Savings Model. Returns (ws, list_of_refs)."""
    ws = wb.create_sheet("Consolidation Savings Model")
    n_cols = 5

    r = write_attribution(ws, n_cols)

    cat_refs = []
    for cat in CATEGORIES:
        r, refs = build_category_block(ws, cat, r)
        cat_refs.append(refs)
        r += 1  # extra spacer between categories

    # --- Combined Totals ---
    merge_row(ws, r, 1, n_cols, "COMBINED TOTALS",
              font=FONT_HEADER, fill=FILL_NAVY, align=ALIGN_LEFT)
    ws.row_dimensions[r].height = 30
    r += 1

    write_header_row(ws, r, ["Metric", "Value", "", "", "Calculation"])
    r += 1

    ref1, ref2 = cat_refs[0], cat_refs[1]

    # Combined 5-Year
    sc(ws, r, 1, "Total 5-Year Spend", font=FONT_BOLD)
    sc(ws, r, 2, f"={ref1['five_year']}+{ref2['five_year']}",
       font=FONT_BOLD, fmt=CURRENCY_FMT)
    sc(ws, r, 5, "Sum of both categories", font=FONT_MATH)
    combined_5yr = f"B{r}"
    r += 1

    # Combined Annual Run Rate
    sc(ws, r, 1, "Combined Annual Run Rate", font=FONT_NORMAL)
    sc(ws, r, 2, f"={ref1['annual']}+{ref2['annual']}",
       font=FONT_NORMAL, fmt=CURRENCY_FMT)
    sc(ws, r, 5, "Sum of annual run rates", font=FONT_MATH)
    combined_annual = f"B{r}"
    r += 1

    # Combined Addressable Spend
    sc(ws, r, 1, "Combined Addressable Spend", font=FONT_NORMAL)
    sc(ws, r, 2, f"={ref1['addressable']}+{ref2['addressable']}",
       font=FONT_NORMAL, fmt=CURRENCY_FMT)
    sc(ws, r, 5, "Sum of addressable spend", font=FONT_MATH)
    combined_addr = f"B{r}"
    r += 1

    # Combined Annual Savings
    sc(ws, r, 1, "Combined Annual Savings", font=FONT_NORMAL)
    sc(ws, r, 2, f"={ref1['annual_savings']}+{ref2['annual_savings']}",
       font=FONT_NORMAL, fmt=CURRENCY_FMT)
    sc(ws, r, 5, "Sum of annual savings", font=FONT_MATH)
    combined_savings = f"B{r}"
    r += 1

    # Combined 3-Year Projected Value
    sc(ws, r, 1, "Combined 3-Year Projected Value",
       font=FONT_BOLD_LG, fill=FILL_LT_GREEN)
    sc(ws, r, 2, f"={ref1['three_year']}+{ref2['three_year']}",
       font=FONT_BOLD_LG, fill=FILL_LT_GREEN, fmt=CURRENCY_FMT)
    sc(ws, r, 3, "", fill=FILL_LT_GREEN)
    sc(ws, r, 4, "", fill=FILL_LT_GREEN)
    sc(ws, r, 5, "Sum of 3-year values", font=FONT_MATH, fill=FILL_LT_GREEN)
    combined_3yr = f"B{r}"
    r += 2

    # --- Formula Explanation ---
    merge_row(ws, r, 1, n_cols, "FORMULA EXPLANATION",
              font=FONT_SECTION, fill=FILL_LT_BLUE, align=ALIGN_LEFT)
    r += 1

    explanations = [
        "Savings Chain: 5-Year Total \u00f7 5 = Annual Run Rate \u00d7 85% = "
        "Addressable Spend \u00d7 10% = Annual Savings \u00d7 3 = 3-Year Projected Value",
        "85% addressable = 100% \u2212 15% already optimized "
        "(McKinsey/NIGP industry benchmark, 10\u201320% range, midpoint used)",
        "10% = moderate savings rate for vendor consolidation via competitive "
        "solicitation (NIGP benchmark range: 5\u201315%)",
        "3-Year projection reflects standard municipal contract term",
    ]
    for line in explanations:
        merge_row(ws, r, 1, n_cols, line, font=FONT_MATH, fill=FILL_LT_BLUE)
        r += 1

    # --- Column widths ---
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 48

    ws.freeze_panes = "A5"

    # Build cross-tab reference dict for Executive Summary
    combined_refs = {
        "combined_5yr": combined_5yr,
        "combined_annual": combined_annual,
        "combined_addr": combined_addr,
        "combined_savings": combined_savings,
        "combined_3yr": combined_3yr,
    }

    return ws, cat_refs, combined_refs


# ============================================================================
# TAB 1: EXECUTIVE SUMMARY
# ============================================================================


def build_executive_summary(wb, cat_refs, combined_refs):
    """Build Tab 1: Executive Summary referencing Tab 2 cells."""
    ws = wb.create_sheet("Executive Summary", 0)  # insert as first tab
    n_cols = 6
    TAB2 = "'Consolidation Savings Model'"

    r = write_attribution(ws, n_cols)

    # Title
    merge_row(ws, r, 1, n_cols, "QUICK WIN CONSOLIDATION SAVINGS MODEL",
              font=FONT_TITLE, align=ALIGN_LEFT)
    r += 1
    merge_row(ws, r, 1, n_cols,
              "Exhibit B Requisition Data \u2014 January 2021 through January 2026",
              font=FONT_SECTION, align=ALIGN_LEFT)
    r += 2  # skip blank

    # --- Summary Metrics ---
    merge_row(ws, r, 1, n_cols, "SUMMARY METRICS",
              font=FONT_HEADER, fill=FILL_NAVY, align=ALIGN_LEFT)
    ws.row_dimensions[r].height = 28
    r += 1

    write_header_row(ws, r, ["Metric", "Value", "", "", "", ""])
    r += 1

    metrics = [
        ("Total Categories Analyzed", 2, INT_FMT),
        ("Total 5-Year Spend", f"={TAB2}!{combined_refs['combined_5yr']}", CURRENCY_FMT),
        ("Combined Annual Savings", f"={TAB2}!{combined_refs['combined_savings']}", CURRENCY_FMT),
        ("Combined 3-Year Projected Value", f"={TAB2}!{combined_refs['combined_3yr']}", CURRENCY_FMT),
    ]

    for label, value, fmt in metrics:
        sc(ws, r, 1, label, font=FONT_BOLD)
        is_3yr = "3-Year" in label
        sc(ws, r, 2, value,
           font=FONT_BOLD_LG if is_3yr else FONT_NORMAL,
           fill=FILL_LT_GREEN if is_3yr else None,
           fmt=fmt)
        # Fill remaining cells with borders
        for c in range(3, n_cols + 1):
            sc(ws, r, c, "", fill=FILL_LT_GREEN if is_3yr else None)
        r += 1

    r += 1  # blank

    # --- Category Detail ---
    merge_row(ws, r, 1, n_cols, "CATEGORY DETAIL",
              font=FONT_HEADER, fill=FILL_NAVY, align=ALIGN_LEFT)
    ws.row_dimensions[r].height = 28
    r += 1

    write_header_row(ws, r, ["Category", "5-Year Total Spend", "Vendor Count",
                              "Department Count", "Solicitation Status",
                              "3-Year Projected Value"])
    detail_header_row = r
    r += 1

    ref1, ref2 = cat_refs[0], cat_refs[1]

    # Category 1
    cat1_row = r
    sc(ws, r, 1, CATEGORIES[0]["name"], font=FONT_NORMAL)
    sc(ws, r, 2, f"={TAB2}!{ref1['five_year']}", font=FONT_NORMAL, fmt=CURRENCY_FMT)
    sc(ws, r, 3, len(CATEGORIES[0]["vendors"]), font=FONT_NORMAL,
       fmt=INT_FMT, align=ALIGN_CENTER)
    sc(ws, r, 4, len(CATEGORIES[0]["departments"]), font=FONT_NORMAL,
       fmt=INT_FMT, align=ALIGN_CENTER)
    sc(ws, r, 5, f"Phase A (Req# {CATEGORIES[0]['req_num']})", font=FONT_NORMAL)
    sc(ws, r, 6, f"={TAB2}!{ref1['three_year']}", font=FONT_NORMAL, fmt=CURRENCY_FMT)
    r += 1

    # Category 2
    cat2_row = r
    sc(ws, r, 1, CATEGORIES[1]["name"], font=FONT_NORMAL)
    sc(ws, r, 2, f"={TAB2}!{ref2['five_year']}", font=FONT_NORMAL, fmt=CURRENCY_FMT)
    sc(ws, r, 3, len(CATEGORIES[1]["vendors"]), font=FONT_NORMAL,
       fmt=INT_FMT, align=ALIGN_CENTER)
    sc(ws, r, 4, len(CATEGORIES[1]["departments"]), font=FONT_NORMAL,
       fmt=INT_FMT, align=ALIGN_CENTER)
    sc(ws, r, 5, f"Phase A (Req# {CATEGORIES[1]['req_num']})", font=FONT_NORMAL)
    sc(ws, r, 6, f"={TAB2}!{ref2['three_year']}", font=FONT_NORMAL, fmt=CURRENCY_FMT)
    r += 1

    # Combined total row
    sc(ws, r, 1, "COMBINED TOTAL", font=FONT_BOLD, fill=FILL_LT_GREEN)
    sc(ws, r, 2, f"=SUM(B{cat1_row}:B{cat2_row})",
       font=FONT_BOLD, fill=FILL_LT_GREEN, fmt=CURRENCY_FMT)
    sc(ws, r, 3, "", font=FONT_BOLD, fill=FILL_LT_GREEN)
    sc(ws, r, 4, "", font=FONT_BOLD, fill=FILL_LT_GREEN)
    sc(ws, r, 5, "", font=FONT_BOLD, fill=FILL_LT_GREEN)
    sc(ws, r, 6, f"=SUM(F{cat1_row}:F{cat2_row})",
       font=FONT_BOLD, fill=FILL_LT_GREEN, fmt=CURRENCY_FMT)
    r += 2  # blank

    # --- Recommendation ---
    merge_row(ws, r, 1, n_cols, "RECOMMENDATION",
              font=FONT_HEADER, fill=FILL_NAVY, align=ALIGN_LEFT)
    ws.row_dimensions[r].height = 28
    r += 1

    recommendation = (
        "Both categories are in Phase A \u2014 Spec Development with no solicitation "
        "issued and no contract awarded. This represents an ideal window for "
        "competitive solicitation. A consolidated RFP for each category would "
        "leverage citywide volume to achieve estimated savings of $306,502 over "
        "3 years. Recommend proceeding to Phase B \u2014 Solicitation for both "
        "categories immediately."
    )
    merge_row(ws, r, 1, n_cols, recommendation,
              font=FONT_NORMAL, fill=FILL_LT_GREEN, align=ALIGN_LEFT)
    ws.row_dimensions[r].height = 65
    r += 2

    # --- Formula chain explanation ---
    merge_row(ws, r, 1, n_cols,
              "Savings Chain: 5-Year Total \u00f7 5 = Annual Run Rate "
              "\u00d7 85% = Addressable Spend \u00d7 10% = Annual Savings "
              "\u00d7 3 = 3-Year Projected Value",
              font=FONT_MATH, fill=FILL_LT_BLUE)

    # --- Column widths ---
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 24

    ws.freeze_panes = "A5"

    return ws


# ============================================================================
# MAIN
# ============================================================================


def main():
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Build Tab 2 first to get cell references
    ws2, cat_refs, combined_refs = build_savings_model(wb)

    # Build Tab 1 using those references
    ws1 = build_executive_summary(wb, cat_refs, combined_refs)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"Saved: {OUTPUT_FILE}")
    print(f"  Tab 1: Executive Summary")
    print(f"  Tab 2: Consolidation Savings Model")
    print(f"  Attribution: {ATTRIBUTION}")


if __name__ == "__main__":
    main()
