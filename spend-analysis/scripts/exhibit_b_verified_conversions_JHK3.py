#!/usr/bin/env python3
"""
Exhibit B -> Awarded Contract: Verified Conversions
====================================================
Strips the analysis down to one question: which Exhibit B (maverick) spend
provably moved onto an awarded contract, and what is the evidence?

Only conversions that can be independently proven are included:
  - the award is linked to the Exhibit B by requisition ID, or by vendor
    plus commodity overlap, AND
  - the award is dated AFTER the Exhibit B request.

Three tabs: Summary | Verified Conversions (the proof) | Backing Data.

Author: James H. Kirby III, CSCP, MS-SCM
Usage:  python exhibit_b_verified_conversions_JHK3.py [source.xlsx] [output.xlsx]
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).resolve().parent))
from exhibit_b_conversion_analysis_JHK3 import (  # noqa: E402
    DEFAULT_SRC, load, match, s,
)

OUT = Path(__file__).resolve().parents[2] / "outputs" / "Exhibit_B_Verified_Conversions_JHK3.xlsx"

NAVY, RED, TINT, WHITE, GREEN, AMBER = "002F6C", "DA291C", "D6EEF9", "FFFFFF", "1E7B34", "B26B00"
FONT = "Arial"
MONEY = '$#,##0.00;($#,##0.00);"-"'
MONEY0 = '$#,##0;($#,##0);"-"'
thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

# Whole-population context for the Verified Conversions tab. These describe the FULL
# Exhibit B report, not just the converted rows, so they cannot be derived from this
# workbook - they are entered as stated constants and labelled with their source.
TOTAL_EB_SPEND = 53461392.14
TOTAL_EB_REQS = 730
TOTAL_SOURCE = ("Source: Exhibit-B Report tab of the supplied workbook - all 730 requisitions, "
                "Jan 5 - Aug 20, 2026. Blue figures are entered values; every percentage below "
                "is calculated from them.")
BLUE = "0000FF"

DEPT_SHORT = {
    "DEPARTMENT OF FLEET AND FACILITY MANAGEMENT": "2FM",
    "CHICAGO DEPARTMENT OF AVIATION": "CDA",
    "CHICAGO POLICE DEPARTMENT": "CPD",
    "CHICAGO DEPARTMENT OF PUBLIC HEALTH": "CDPH",
    "DEPARTMENT OF PROCUREMENT SERVICES": "DPS",
    "DEPARTMENT OF WATER MANAGEMENT": "DWM",
    "DEPARTMENT OF STREETS AND SANITATION": "DSS",
    "DEPARTMENT OF ASSETS INFORMATION AND SERVICES": "AIS",
}


def short_depts(series):
    """Every department in the group, not just the first - a sourcing effort can span several."""
    names = sorted({s(x).strip() for x in series if s(x).strip()})
    return ", ".join(DEPT_SHORT.get(n, n[:18]) for n in names)


def banner(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, bold=True, size=14, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = Font(name=FONT, italic=True, size=9, color="444444")
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[2].height = 16


def header(ws, row, ncols):
    for i in range(1, ncols + 1):
        c = ws.cell(row=row, column=i)
        c.font = Font(name=FONT, bold=True, size=9, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[row].height = 30


def put(ws, df, start, money=(), dates=(), wrap=()):
    for j, col in enumerate(df.columns, 1):
        ws.cell(row=start, column=j, value=col)
    header(ws, start, len(df.columns))
    for i, (_, r) in enumerate(df.iterrows()):
        rw = start + 1 + i
        for j, col in enumerate(df.columns, 1):
            v = r[col]
            if isinstance(v, pd.Timestamp):
                v = v.to_pydatetime() if pd.notna(v) else None
            elif pd.isna(v):
                v = ""
            c = ws.cell(row=rw, column=j, value=v)
            c.font = Font(name=FONT, size=9)
            c.border = BOX
            c.alignment = Alignment(vertical="top", wrap_text=col in wrap)
            if col in money:
                c.number_format = MONEY
            if col in dates:
                c.number_format = "yyyy-mm-dd"
            if i % 2 == 1:
                c.fill = PatternFill("solid", fgColor=TINT)
    return start + len(df)


def widths(ws, spec):
    for k, v in spec.items():
        ws.column_dimensions[k].width = v


# ------------------------------------------------------------------ build ---
def build_proof(res):
    """One row per sourcing effort that verifiably converted."""
    c = res[res["CONVERTED"]].copy()
    rows = []
    for key, g in c.groupby("SOURCING_KEY"):
        f = g.iloc[0]
        depts = short_depts(g["Department Description"])
        rows.append({
            # Identity comes from the AWARDED CONTRACT, which names the commodity cleanly.
            # Exhibit B descriptions are invoice-level free text - often a single invoice number
            # and a dollar figure that is NOT the group total, which reads as a contradiction.
            "What Was Being Bought": s(f["AW_DESC"])[:95],
            "Exhibit B Dept(s)": depts,
            "Exhibit B Reqs": len(g),
            "Exhibit B Spend": round(g["Requisition Amount"].sum(), 2),
            "First Exhibit B": g["EB_Date"].min(),
            "Awarded Contract": s(f["AW_CONTRACT_ID"]),
            "Awarded Vendor": s(f["AW_VENDOR"]),
            "Award Date": f["AW_DATE"],
            "Award Method": s(f["AW_TYPE"]),
            "Proof": ("Requisition ID link" if f["tier"] == "A"
                      else f"Vendor + commodity match ({f['basis'].split('similarity ')[-1].rstrip(')')})"),
            "Sourcing Key": key,
        })
    df = pd.DataFrame(rows).sort_values("Exhibit B Spend", ascending=False).reset_index(drop=True)
    df.insert(0, "#", range(1, len(df) + 1))
    return df


def tab_summary(wb, proof, res):
    ws = wb.create_sheet("Summary")
    widths(ws, {"A": 3, "B": 46, "C": 20, "D": 62})
    eb_lo, eb_hi = res["EB_Date"].min(), res["EB_Date"].max()
    banner(ws, "EXHIBIT B SPEND VERIFIED AS MOVED TO AN AWARDED CONTRACT",
           f"Exhibit B activity {eb_lo:%b %-d} - {eb_hi:%b %-d, %Y}   |   "
           f"only conversions that can be independently proven", 4)
    n = len(proof)
    last = 5 + n  # proof table data rows live at rows 6..(5+n) on the proof tab
    P = f"'Verified Conversions'!"

    r = 4
    ws.cell(row=r, column=2, value="THE ANSWER").font = Font(name=FONT, bold=True, size=12, color=NAVY)
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r + 1, end_column=4)
    n_bid = int((proof["Award Method"] == "BID").sum())
    c = ws.cell(row=r, column=2, value=(
        f"Yes - {n} Exhibit B items moved onto an awarded contract this year, covering "
        f"{int(proof['Exhibit B Reqs'].sum())} requisitions and "
        f"${proof['Exhibit B Spend'].sum():,.2f} of maverick spend. All {n} are wins: the spend is now "
        f"under a contract vehicle with negotiated terms and an audit trail, instead of being authorized "
        f"purchase by purchase as a contract exception. {n_bid} were competitively bid; the remainder "
        f"were awarded by sole source or emergency, which are legitimate procurement methods with their "
        f"own justification process - not maverick spend."))
    c.font = Font(name=FONT, size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.fill = PatternFill("solid", fgColor=TINT)
    c.border = BOX
    ws.row_dimensions[r].height = 30
    ws.row_dimensions[r + 1].height = 30
    r += 3

    ws.cell(row=r, column=2, value="THE NUMBERS").font = Font(name=FONT, bold=True, size=12, color=NAVY)
    r += 1
    for j, h in enumerate(["Measure", "Value", "How it is counted"], 2):
        ws.cell(row=r, column=j, value=h)
    header(ws, r, 4)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=WHITE)
    r += 1
    lines = [
        ("Exhibit B items converted", f'=COUNTA({P}$B$6:$B${last})', "#,##0",
         "One per sourcing effort - the commodity being put under contract"),
        ("Requisitions behind them", f'=SUM({P}$D$6:$D${last})', "#,##0",
         "Individual Exhibit B requisitions rolled into those items. Two carry a $0 amount, "
         "so 20 of the 22 contribute dollars - the spend total is unaffected"),
        ("Maverick spend eliminated", f'=SUM({P}$E$6:$E${last})', MONEY,
         "Exhibit B dollars now covered by an awarded contract"),
        # Distinct count without array functions: each value contributes 1/(its occurrence count).
        ("Distinct contracts they landed on",
         f'=SUMPRODUCT(1/COUNTIF({P}$G$6:$G${last},{P}$G$6:$G${last}))', "#,##0",
         "Fewer than the item count - separate Exhibit B items can converge on one contract"),
        ("   awarded by competitive bid", f'=COUNTIF({P}$J$6:$J${last},"BID")', "#,##0",
         "Competitively solicited"),
        ("   awarded by sole source or emergency",
         f'=COUNTA({P}$J$6:$J${last})-COUNTIF({P}$J$6:$J${last},"BID")', "#,##0",
         "Also a contract vehicle - justified, documented, auditable. Still a win"),
        ("Proven by requisition ID", f'=COUNTIF({P}$K$6:$K${last},"Requisition ID link")', "#,##0",
         "Hardest evidence - same requisition number in both systems"),
        ("Proven by vendor + commodity", f'=COUNTA({P}$K$6:$K${last})-COUNTIF({P}$K$6:$K${last},"Requisition ID link")',
         "#,##0", "Same vendor, matching commodity, award dated after"),
    ]
    for label, formula, fmt, how in lines:
        ws.cell(row=r, column=2, value=label).font = Font(name=FONT, size=10)
        v = ws.cell(row=r, column=3, value=formula)
        v.number_format = fmt
        v.font = Font(name=FONT, size=11, bold=True, color=NAVY)
        v.alignment = Alignment(horizontal="center")
        h = ws.cell(row=r, column=4, value=how)
        h.font = Font(name=FONT, size=8, color="555555")
        h.alignment = Alignment(wrap_text=True, vertical="center")
        for cc in range(2, 5):
            ws.cell(row=r, column=cc).border = BOX
        ws.row_dimensions[r].height = 24
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="WHAT THE PROOF MEANS").font = Font(name=FONT, bold=True, size=12, color=NAVY)
    r += 1
    for label, body in [
        ("Requisition ID link",
         "The new-contract requisition number the department wrote on the Exhibit B form appears as the "
         "originating requisition on the awarded contract. Same number, both systems. This is the hardest "
         "evidence available and covers most of the conversions listed."),
        ("Vendor + commodity match",
         "No matching requisition number, but the same vendor holds the new contract, the commodity "
         "descriptions overlap, and the award is dated after the Exhibit B request. Strong, but one step "
         "softer than an ID match - the similarity score is shown so it can be judged individually."),
        ("Award method is shown as information, not as a pass or fail",
         "Every item listed is a win - the spend left Exhibit B and is now on a contract. The award "
         "method is reported because it tells you how the contract was procured, but a sole-source or "
         "emergency award is a legitimate, documented procurement vehicle. It is not maverick spend."),
        ("Timing test applied to every row",
         "Every conversion listed has its contract awarded AFTER the Exhibit B request. Matches where the "
         "contract already existed are excluded - those are not conversions."),
    ]:
        ws.cell(row=r, column=2, value=label).font = Font(name=FONT, bold=True, size=9, color=RED)
        b = ws.cell(row=r, column=3, value=body)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        b.font = Font(name=FONT, size=9)
        b.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 40
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="ONE CAVEAT").font = Font(name=FONT, bold=True, size=11, color=NAVY)
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r + 1, end_column=4)
    c = ws.cell(row=r, column=2, value=(
        f"This is what can be PROVEN from the two tabs supplied - {len(res)} Exhibit B requisitions and the "
        f"awarded contract list. The two systems share no reliable common key, so a conversion that left no "
        f"traceable link would not appear here. Treat this as the floor, not the ceiling."))
    c.font = Font(name=FONT, size=9, italic=True)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.border = BOX
    ws.sheet_view.showGridLines = False
    return ws


def tab_proof(wb, proof):
    ws = wb.create_sheet("Verified Conversions")
    banner(ws, "VERIFIED CONVERSIONS - MAVERICK SPEND NOW UNDER CONTRACT",
           "Each row is one Exhibit B item proven to have moved onto an awarded contract - all are "
           "wins. The item is named from the awarded contract; the Exhibit B requisitions behind it "
           "are on the Backing Data tab. The Proof column states the evidence.", 11)
    df = proof.drop(columns=["Sourcing Key"])
    ws.cell(row=4, column=1, value="EXHIBIT B SIDE").font = Font(name=FONT, bold=True, size=9, color=WHITE)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=6)
    ws.cell(row=4, column=1).fill = PatternFill("solid", fgColor="41B6E6")
    ws.cell(row=4, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=4, column=7, value="AWARDED CONTRACT SIDE").font = Font(name=FONT, bold=True, size=9, color=WHITE)
    ws.merge_cells(start_row=4, start_column=7, end_row=4, end_column=11)
    ws.cell(row=4, column=7).fill = PatternFill("solid", fgColor=GREEN)
    ws.cell(row=4, column=7).alignment = Alignment(horizontal="center")

    end = put(ws, df, 5, money=("Exhibit B Spend",),
              dates=("First Exhibit B", "Award Date"), wrap=("Exhibit B Item", "Proof"))
    tr = end + 1
    ws.cell(row=tr, column=2, value="TOTAL")
    ws.cell(row=tr, column=4, value=f"=SUM(D6:D{end})").number_format = "#,##0"
    ws.cell(row=tr, column=5, value=f"=SUM(E6:E{end})").number_format = MONEY
    for cc in range(1, 12):
        c = ws.cell(row=tr, column=cc)
        c.fill = PatternFill("solid", fgColor=RED)
        c.font = Font(name=FONT, bold=True, size=10, color=WHITE)
        c.border = BOX
    for rw in range(6, end + 1):
        cc = ws.cell(row=rw, column=10)
        cc.font = Font(name=FONT, size=9, bold=True, color=GREEN)
        cc.alignment = Alignment(horizontal="center")
    widths(ws, {"A": 5, "B": 56, "C": 16, "D": 10, "E": 15, "F": 13, "G": 14,
                "H": 32, "I": 12, "J": 17, "K": 30})
    for rw in range(6, end + 1):
        ws.row_dimensions[rw].height = 42
    ws.freeze_panes = "A6"
    context_block(ws, tr, end)
    return ws


def context_block(ws, total_row, last_data_row):
    """Small detached block below the main table: what converted, against the whole year.

    Columns used: B measure | C amount | D requisitions | E percent.
    """
    r = total_row + 3
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    c = ws.cell(row=r, column=2,
                value="CONTEXT - HOW THIS COMPARES TO ALL EXHIBIT B SPEND THIS YEAR")
    c.font = Font(name=FONT, bold=True, size=11, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[r].height = 22
    r += 1

    for j, h in enumerate(["", "Amount", "Requisitions", "% of Total"], 2):
        x = ws.cell(row=r, column=j, value=h)
        x.font = Font(name=FONT, bold=True, size=9, color=WHITE)
        x.fill = PatternFill("solid", fgColor="41B6E6")
        x.alignment = Alignment(horizontal="center", vertical="center")
        x.border = BOX
    r += 1

    tot_r, conv_r, out_r = r, r + 1, r + 2
    rows = [
        ("Total Exhibit B spend, Jan 5 - Aug 20, 2026", TOTAL_EB_SPEND, TOTAL_EB_REQS,
         f"=C{tot_r}/$C${tot_r}", True),
        ("Verified as moved onto an awarded contract", f"=E{total_row}", f"=D{total_row}",
         f"=C{conv_r}/$C${tot_r}", False),
        ("Still being bought outside a contract", f"=C{tot_r}-C{conv_r}", f"=D{tot_r}-D{conv_r}",
         f"=C{out_r}/$C${tot_r}", False),
    ]
    for label, amt, reqs, pct, entered in rows:
        ws.cell(row=r, column=2, value=label).font = Font(name=FONT, size=10,
                                                          bold=(r == out_r))
        a = ws.cell(row=r, column=3, value=amt)
        a.number_format = MONEY
        a.font = Font(name=FONT, size=10, bold=True,
                      color=BLUE if entered else (RED if r == out_r else NAVY))
        q = ws.cell(row=r, column=4, value=reqs)
        q.number_format = "#,##0"
        q.font = Font(name=FONT, size=10, color=BLUE if entered else "000000")
        q.alignment = Alignment(horizontal="center")
        pc = ws.cell(row=r, column=5, value=pct)
        pc.number_format = "0.00%"
        pc.font = Font(name=FONT, size=11, bold=True,
                       color=RED if r == out_r else (NAVY if not entered else "000000"))
        pc.alignment = Alignment(horizontal="center")
        for cc in range(2, 6):
            ws.cell(row=r, column=cc).border = BOX
            if r == out_r:
                ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=TINT)
        r += 1

    r += 1
    ws.cell(row=r, column=2,
            value="Exhibit B spend for every $1 moved onto a contract").font = \
        Font(name=FONT, bold=True, size=10, color=NAVY)
    ratio = ws.cell(row=r, column=3, value=f"=C{tot_r}/C{conv_r}")
    ratio.number_format = '"$"#,##0'
    ratio.font = Font(name=FONT, bold=True, size=12, color=RED)
    ratio.alignment = Alignment(horizontal="center")
    for cc in range(2, 4):
        ws.cell(row=r, column=cc).border = BOX
        ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=TINT)
    r += 2

    for text, italic, size in [
        ("Only conversions provable against the awarded contract file are counted here. The two "
         "systems share no common key, so the true figure may be modestly higher - but not by an "
         "order of magnitude.", True, 9),
        (TOTAL_SOURCE, True, 8),
    ]:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        x = ws.cell(row=r, column=2, value=text)
        x.font = Font(name=FONT, italic=italic, size=size, color="555555")
        x.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 26
        r += 1
    return ws


def tab_backing(wb, res, aw, proof):
    ws = wb.create_sheet("Backing Data")
    banner(ws, "BACKING DATA - EVERY ROW BEHIND THE PROOF",
           "Section 1: the Exhibit B requisitions that converted - descriptions here are invoice-level "
           "free text, so a dollar figure inside a description is that invoice, not the row total.  "
           "Section 2: the awarded contracts they landed on. Every figure above traces to these rows.", 11)
    c = res[res["CONVERTED"]].copy().sort_values(["SOURCING_KEY", "EB_Date"])
    eb_df = pd.DataFrame({
        "Sourcing Key": c["SOURCING_KEY"],
        "Exhibit B Req#": c["Requisition Number"],
        "Exhibit B Date": c["EB_Date"],
        "Department": c["Department Description"].map(s),
        "Exhibit B Vendor": c["Vendor Name"].map(s),
        "Description": c["Requisition Description"].map(lambda x: s(x)[:110]),
        "Amount": c["Requisition Amount"].round(2),
        "Justification": c["Justification"].map(lambda x: s(x)[:70]),
        "New Contract Req# on form": c["Requisition# for New Contract/Mod"].map(s),
        "Matched Contract": c["AW_CONTRACT_ID"].map(s),
        "Evidence": c["basis"].map(s),
    })
    ws.cell(row=4, column=1, value="SECTION 1 - EXHIBIT B REQUISITIONS THAT CONVERTED "
                                   f"({len(eb_df)} rows)").font = Font(name=FONT, bold=True, size=10, color=NAVY)
    end1 = put(ws, eb_df, 5, money=("Amount",), dates=("Exhibit B Date",),
               wrap=("Description", "Evidence", "Justification"))
    t1 = end1 + 1
    ws.cell(row=t1, column=6, value="SECTION 1 TOTAL")
    ws.cell(row=t1, column=7, value=f"=SUM(G6:G{end1})").number_format = MONEY
    for cc in range(1, 12):
        x = ws.cell(row=t1, column=cc)
        x.fill = PatternFill("solid", fgColor=RED)
        x.font = Font(name=FONT, bold=True, size=9, color=WHITE)
        x.border = BOX

    start2 = t1 + 3
    ids = set(proof["Awarded Contract"])
    a = aw[aw["CID_s"].isin(ids)].copy()
    aw_df = pd.DataFrame({
        "Contract ID": a["CID_s"],
        "PO": a["PO"],
        "Requisition": a["REQ"],
        "Spec": a["SPEC"].map(s),
        "Description": a["DESCRIPTION"].map(lambda x: s(x)[:110]),
        "Vendor": a["VENDOR"].map(s),
        "Department": a["DEPT"].map(s),
        "Award Date": a["AWDDATE"],
        "Award Amount": a["AMOUNT"],
        "Procurement Type": a["PROCUREMENT_TYPE"].map(s),
        "Contract Type": a["CONTRACT_TYPE"].map(s),
    })
    ws.cell(row=start2 - 1, column=1,
            value=f"SECTION 2 - AWARDED CONTRACTS THEY LANDED ON ({len(aw_df)} contracts)").font = \
        Font(name=FONT, bold=True, size=10, color=NAVY)
    put(ws, aw_df, start2, money=("Award Amount",), dates=("Award Date",), wrap=("Description",))
    widths(ws, {"A": 14, "B": 15, "C": 14, "D": 30, "E": 26, "F": 54, "G": 13,
                "H": 40, "I": 24, "J": 17, "K": 34})
    return ws


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SRC
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else OUT
    out.parent.mkdir(parents=True, exist_ok=True)

    eb, aw = load(src)
    res = match(eb, aw)
    proof = build_proof(res)

    wb = Workbook()
    wb.remove(wb.active)
    tab_summary(wb, proof, res)
    tab_proof(wb, proof)
    tab_backing(wb, res, aw, proof)
    wb.save(out)

    print(f"{len(proof)} verified conversions | "
          f"{int(proof['Exhibit B Reqs'].sum())} requisitions | "
          f"${proof['Exhibit B Spend'].sum():,.2f}")
    print("award methods: " + ", ".join(
        f"{k} {v}" for k, v in proof["Award Method"].value_counts().items()))
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
