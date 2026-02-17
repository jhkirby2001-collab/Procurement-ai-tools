"""
Add Departments/Using Areas Table and Chart to Landscaping Consolidation Workbook
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

# Load the landscaping data
input_path = "/workspaces/Spend_Analysis_and_Categorization/spend-analysis/data/processed/landscaping_structured.xlsx"
df = pd.read_excel(input_path)

# Filter to 2015-2025
df['Award_Date'] = pd.to_datetime(df['Award_Date'])
df['Award_Year'] = df['Award_Date'].dt.year
df = df[(df['Award_Year'] >= 2015) & (df['Award_Year'] <= 2025)].copy()

# Calculate department spend
dept_spend = df.groupby('Department').agg({
    'Award_Amount_Numeric': 'sum',
    'Contract (PO) #': 'count'
}).reset_index()
dept_spend.columns = ['Department', 'Total Spend', 'Contract Count']
dept_spend = dept_spend.sort_values('Total Spend', ascending=False)
total_spend = dept_spend['Total Spend'].sum()
dept_spend['% of Total'] = (dept_spend['Total Spend'] / total_spend * 100)

print("="*80)
print("ADDING DEPARTMENTS/USING AREAS ANALYSIS")
print("="*80)
print(f"\nFound {len(dept_spend)} departments/areas using landscaping services")
print(f"Total spend: ${total_spend:,.2f}")
print("\nTop 5 departments:")
for idx, row in dept_spend.head(5).iterrows():
    print(f"  {row['Department']}: ${row['Total Spend']:,.2f} ({row['% of Total']:.1f}%)")

# Load existing workbook
wb_path = "/workspaces/Spend_Analysis_and_Categorization/spend-analysis/outputs/new landscape analysis.xlsx"
wb = load_workbook(wb_path)

# Define styles
header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')

# ============================================================================
# UPDATE SOURCE DATA SHEET - ADD USING AREAS TABLE
# ============================================================================
print("\n1. Adding USING AREAS/DEPARTMENTS table to Source Data sheet...")
ws_source = wb["Source Data"]

# Find the row where we added the departments table previously and replace it
# Or add new table at the end - let's find the last row with data
last_row = ws_source.max_row

# Add space
row = last_row + 3

# USING AREAS/DEPARTMENTS TABLE
ws_source[f'A{row}'] = 'USING AREAS/DEPARTMENTS'
ws_source[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
ws_source[f'A{row}'].fill = header_fill
ws_source.merge_cells(f'A{row}:D{row}')
row += 1

dept_headers = ['Rank', 'Department/Area', 'Total Spend', '% of Total']
for col, header in enumerate(dept_headers, 1):
    cell = ws_source.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border_thin
row += 1

start_data_row = row
for idx, (_, dept_row) in enumerate(dept_spend.iterrows(), 1):
    ws_source.cell(row=row, column=1, value=idx).alignment = center_align
    ws_source.cell(row=row, column=2, value=dept_row['Department']).alignment = left_align
    ws_source.cell(row=row, column=3, value=dept_row['Total Spend']).number_format = '$#,##0'
    ws_source.cell(row=row, column=4, value=dept_row['% of Total'] / 100).number_format = '0.0%'

    for col in range(1, 5):
        ws_source.cell(row=row, column=col).border = border_thin
    row += 1

print(f"   ✓ Added {len(dept_spend)} departments/areas to table")

# Adjust column widths
ws_source.column_dimensions['A'].width = 8
ws_source.column_dimensions['B'].width = 50
ws_source.column_dimensions['C'].width = 18
ws_source.column_dimensions['D'].width = 12

# ============================================================================
# UPDATE CHARTS SHEET - ADD TOP 10 DEPARTMENTS CHART
# ============================================================================
print("\n2. Adding Top 10 Departments bar chart to Charts sheet...")
ws_charts = wb["Charts"]

# Find next available row for data
chart_data_row = 40

# Add data for top 10 departments chart
ws_charts[f'A{chart_data_row}'] = 'Top 10 Using Departments/Areas by Spend'
ws_charts[f'A{chart_data_row}'].font = Font(name='Calibri', size=12, bold=True)

dept_chart_headers = ['Department/Area', 'Total Spend (Millions)']
row = chart_data_row + 1
for col, header in enumerate(dept_chart_headers, 1):
    cell = ws_charts.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border_thin
row += 1

top10_depts = dept_spend.head(10)
for _, dept_row in top10_depts.iterrows():
    ws_charts.cell(row=row, column=1, value=dept_row['Department']).alignment = left_align
    ws_charts.cell(row=row, column=2, value=dept_row['Total Spend'] / 1_000_000).number_format = '0.0'

    for col in range(1, 3):
        ws_charts.cell(row=row, column=col).border = border_thin
    row += 1

# Create horizontal bar chart
bar_depts = BarChart()
bar_depts.type = "bar"  # Horizontal bars
bar_depts.title = "Top 10 Departments/Areas by Landscaping Spend"
bar_depts.y_axis.title = "Department/Area"
bar_depts.x_axis.title = "Total Spend ($ Millions)"
bar_depts.style = 12
bar_depts.height = 15
bar_depts.width = 18

dept_labels = Reference(ws_charts, min_col=1, min_row=chart_data_row+2, max_row=chart_data_row+11)
dept_values = Reference(ws_charts, min_col=2, min_row=chart_data_row+1, max_row=chart_data_row+11)
bar_depts.add_data(dept_values, titles_from_data=True)
bar_depts.set_categories(dept_labels)

# Add data labels
bar_depts.dataLabels = DataLabelList()
bar_depts.dataLabels.showVal = True

ws_charts.add_chart(bar_depts, "E40")

print(f"   ✓ Added Top 10 Departments horizontal bar chart")

# Adjust column widths for chart data
ws_charts.column_dimensions['A'].width = 50

# Save workbook
wb.save(wb_path)

print("\n" + "="*80)
print("DEPARTMENTS/USING AREAS ANALYSIS ADDED")
print("="*80)
print(f"\nUpdated file: {wb_path}")
print("\nChanges made:")
print(f"  ✓ Source Data: Added USING AREAS/DEPARTMENTS table ({len(dept_spend)} entries)")
print("  ✓ Charts: Added Top 10 Departments horizontal bar chart")
print("\nDepartment breakdown:")
for idx, row in dept_spend.iterrows():
    print(f"  {row['Department']}: ${row['Total Spend']:,.0f} ({row['% of Total']:.1f}%)")
print("\n" + "="*80)
